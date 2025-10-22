#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
深度分析石滩区分区计量.xlsx表格
"""

import openpyxl
from openpyxl.utils import get_column_letter
import json
from datetime import datetime

def analyze_excel_structure(file_path):
    """
    深度分析Excel文件结构
    """
    print("=" * 80)
    print("石滩区分区计量.xlsx 深度分析报告")
    print("=" * 80)
    print(f"分析时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"文件路径: {file_path}")
    print()
    
    try:
        # 加载工作簿
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        print("=" * 80)
        print("1. 工作簿基本信息")
        print("=" * 80)
        print(f"工作表数量: {len(wb.sheetnames)}")
        print(f"工作表名称: {', '.join(wb.sheetnames)}")
        print()
        
        # 分析每个工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            print("=" * 80)
            print(f"2. 工作表分析: [{sheet_name}]")
            print("=" * 80)
            
            # 基本信息
            print(f"\n2.1 基本信息:")
            print(f"  - 最大行数: {ws.max_row}")
            print(f"  - 最大列数: {ws.max_column}")
            print(f"  - 数据范围: A1:{get_column_letter(ws.max_column)}{ws.max_row}")
            
            # 表头分析
            print(f"\n2.2 表头结构:")
            headers = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(1, col)
                if cell.value:
                    headers.append({
                        'column': get_column_letter(col),
                        'index': col,
                        'name': str(cell.value),
                        'type': type(cell.value).__name__
                    })
            
            if headers:
                print(f"  共 {len(headers)} 列:")
                for h in headers:
                    print(f"    列{h['column']} (第{h['index']}列): {h['name']}")
            else:
                print("  未找到表头")
            
            # 数据类型分析
            print(f"\n2.3 数据类型分析:")
            if len(headers) > 0 and ws.max_row > 1:
                # 分析前10行数据
                sample_rows = min(10, ws.max_row - 1)
                print(f"  (基于前{sample_rows}行数据样本)")
                
                for h in headers:
                    col_idx = h['index']
                    col_name = h['name']
                    
                    # 收集该列的数据类型
                    types = {}
                    non_empty = 0
                    sample_values = []
                    
                    for row in range(2, min(12, ws.max_row + 1)):
                        cell = ws.cell(row, col_idx)
                        if cell.value is not None:
                            non_empty += 1
                            type_name = type(cell.value).__name__
                            types[type_name] = types.get(type_name, 0) + 1
                            if len(sample_values) < 3:
                                sample_values.append(str(cell.value))
                    
                    print(f"\n  列 [{col_name}]:")
                    print(f"    - 非空单元格: {non_empty}/{sample_rows}")
                    print(f"    - 数据类型: {', '.join([f'{k}({v})' for k, v in types.items()])}")
                    if sample_values:
                        print(f"    - 样本值: {', '.join(sample_values[:3])}")
            
            # 数据范围分析
            print(f"\n2.4 数据内容分析:")
            if ws.max_row > 1:
                print(f"  - 数据行数: {ws.max_row - 1} 行（不含表头）")
                
                # 检查是否有日期列
                date_columns = []
                for h in headers:
                    if '日期' in h['name'] or '时间' in h['name'] or 'date' in h['name'].lower():
                        date_columns.append(h)
                
                if date_columns:
                    print(f"  - 发现日期列: {', '.join([d['name'] for d in date_columns])}")
                    
                    # 分析日期范围
                    for date_col in date_columns:
                        col_idx = date_col['index']
                        dates = []
                        
                        for row in range(2, ws.max_row + 1):
                            cell = ws.cell(row, col_idx)
                            if cell.value:
                                dates.append(cell.value)
                        
                        if dates:
                            print(f"\n  列 [{date_col['name']}] 日期范围:")
                            print(f"    - 最早: {min(dates)}")
                            print(f"    - 最晚: {max(dates)}")
                            print(f"    - 总计: {len(dates)} 条记录")
                
                # 检查是否有数值列
                numeric_columns = []
                for h in headers:
                    col_idx = h['index']
                    # 检查第2行数据类型
                    cell = ws.cell(2, col_idx)
                    if isinstance(cell.value, (int, float)):
                        numeric_columns.append(h)
                
                if numeric_columns:
                    print(f"\n  - 发现数值列: {', '.join([n['name'] for n in numeric_columns])}")
                    
                    # 分析数值范围
                    for num_col in numeric_columns[:5]:  # 只分析前5个数值列
                        col_idx = num_col['index']
                        values = []
                        
                        for row in range(2, ws.max_row + 1):
                            cell = ws.cell(row, col_idx)
                            if isinstance(cell.value, (int, float)):
                                values.append(cell.value)
                        
                        if values:
                            print(f"\n  列 [{num_col['name']}] 统计:")
                            print(f"    - 最小值: {min(values)}")
                            print(f"    - 最大值: {max(values)}")
                            print(f"    - 平均值: {sum(values)/len(values):.2f}")
                            print(f"    - 非空记录: {len(values)}")
            
            # 合并单元格分析
            print(f"\n2.5 合并单元格:")
            merged_cells = list(ws.merged_cells.ranges)
            if merged_cells:
                print(f"  共 {len(merged_cells)} 个合并区域:")
                for i, merged in enumerate(merged_cells[:10], 1):  # 只显示前10个
                    print(f"    {i}. {merged}")
                if len(merged_cells) > 10:
                    print(f"    ... 还有 {len(merged_cells) - 10} 个")
            else:
                print("  无合并单元格")
            
            # 公式分析
            print(f"\n2.6 公式分析:")
            formula_count = 0
            formula_samples = []
            
            for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row)):
                for cell in row:
                    if cell.data_type == 'f':  # 公式单元格
                        formula_count += 1
                        if len(formula_samples) < 5:
                            formula_samples.append({
                                'cell': cell.coordinate,
                                'formula': cell.value
                            })
            
            if formula_count > 0:
                print(f"  共 {formula_count} 个公式单元格（前20行）")
                if formula_samples:
                    print("  样本公式:")
                    for sample in formula_samples:
                        print(f"    {sample['cell']}: {sample['formula']}")
            else:
                print("  未发现公式")
            
            # 数据完整性分析
            print(f"\n2.7 数据完整性:")
            if ws.max_row > 1 and len(headers) > 0:
                empty_cells_per_col = {}
                
                for h in headers:
                    col_idx = h['index']
                    empty_count = 0
                    
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row, col_idx)
                        if cell.value is None or str(cell.value).strip() == '':
                            empty_count += 1
                    
                    empty_cells_per_col[h['name']] = {
                        'empty': empty_count,
                        'total': ws.max_row - 1,
                        'completeness': ((ws.max_row - 1 - empty_count) / (ws.max_row - 1) * 100) if ws.max_row > 1 else 0
                    }
                
                print("  各列完整度:")
                for col_name, stats in empty_cells_per_col.items():
                    print(f"    {col_name}: {stats['completeness']:.1f}% ({stats['total'] - stats['empty']}/{stats['total']})")
            
            print()
        
        # 总结
        print("=" * 80)
        print("3. 分析总结")
        print("=" * 80)
        
        # 判断表格用途
        print("\n3.1 表格用途推断:")
        if '分区计量' in file_path:
            print("  这是一个分区计量统计表格，可能用于:")
            print("  - 记录不同区域的水表计量数据")
            print("  - 统计各分区的用水情况")
            print("  - 进行分区用水量对比分析")
        
        print("\n3.2 数据特点:")
        ws = wb[wb.sheetnames[0]]
        if ws.max_row > 100:
            print(f"  - 数据量较大 ({ws.max_row - 1} 行)")
        elif ws.max_row > 10:
            print(f"  - 数据量适中 ({ws.max_row - 1} 行)")
        else:
            print(f"  - 数据量较小 ({ws.max_row - 1} 行)")
        
        print("\n3.3 建议:")
        print("  - 可以用于历史数据分析")
        print("  - 可以与实时采集数据对比")
        print("  - 可以生成统计报表")
        
        wb.close()
        
        print("\n" + "=" * 80)
        print("分析完成")
        print("=" * 80)
        
        return True
    
    except Exception as e:
        print(f"\n错误: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    file_path = "excel_exports/石滩区分区计量.xlsx"
    analyze_excel_structure(file_path)

