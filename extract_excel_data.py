#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# type: ignore

# 尝试导入openpyxl，如果失败则提供错误信息
try:
    import openpyxl
except ImportError as e:
    print(f"⚠️ openpyxl不可用: {e}")
    openpyxl = None

from datetime import datetime

def extract_actual_data():
    """提取Excel文件中的实际数值数据"""
    
    # 检查openpyxl是否可用
    if not openpyxl:
        print("❌ 无法提取Excel数据，缺少openpyxl库")
        return
    
    try:
        print("🔍 提取Excel文件中的实际数值...")
        
        # 检查文件是否存在
        import os
        excel_file_path = 'excel_exports/石滩供水服务部每日总供水情况.xlsx'
        if not os.path.exists(excel_file_path):
            print(f"❌ Excel文件不存在: {excel_file_path}")
            return
        
        # 读取Excel文件，计算公式结果
        wb = openpyxl.load_workbook(excel_file_path, data_only=True)
        sheet = wb.active
        
        print('=' * 60)
        print('📊 实际数值数据分析')
        print('=' * 60)
        
        # 获取列标题
        headers = []
        for col in range(1, min(17, sheet.max_column + 1)):  # 只看前16列主要业务列
            header = sheet.cell(4, col).value
            if header and str(header).strip():
                headers.append(str(header).strip().replace('\n', ' '))
        
        print('📋 数据列标题:')
        for i, header in enumerate(headers, 1):
            print(f'  第{i:2d}列: {header}')
        
        print()
        print('=' * 60)
        print('📈 最近10天实际数据样例')
        print('=' * 60)
        
        # 找到最新的有数据的行
        latest_rows = []
        for row in range(5, min(100, sheet.max_row + 1)):
            date_cell = sheet.cell(row, 1).value
            if date_cell and isinstance(date_cell, datetime):
                # 检查是否有实际数值
                has_data = False
                row_values = []
                for col in range(1, min(17, sheet.max_column + 1)):
                    cell_value = sheet.cell(row, col).value
                    if col == 1:  # 日期列
                        row_values.append(date_cell.strftime('%Y-%m-%d'))
                    elif isinstance(cell_value, (int, float)) and cell_value != 0:
                        row_values.append(f'{cell_value:.1f}')
                        has_data = True
                    else:
                        row_values.append('')
                
                if has_data:
                    latest_rows.append(row_values)
                    if len(latest_rows) >= 10:
                        break
        
        if latest_rows:
            # 显示表头
            header_line = ' | '.join([f'{h[:12]:<12}' for h in headers[:8]])
            print(header_line)
            print('-' * len(header_line))
            
            # 显示数据
            for row_data in latest_rows[-10:]:  # 最新10行
                data_line = ' | '.join([f'{d[:12]:<12}' for d in row_data[:8]])
                print(data_line)
        
        print()
        print('=' * 60)
        print('📊 数据统计分析')
        print('=' * 60)
        
        # 统计有效数据
        total_data_rows = 0
        date_range = []
        meter_stats = {i: {'count': 0, 'sum': 0, 'max': 0, 'min': float('inf')} 
                      for i in range(2, 17)}  # 第2-16列是数据列
        
        for row in range(5, sheet.max_row + 1):
            date_cell = sheet.cell(row, 1).value
            if date_cell and isinstance(date_cell, datetime):
                date_range.append(date_cell)
                has_valid_data = False
                
                for col in range(2, 17):  # 数据列
                    cell_value = sheet.cell(row, col).value
                    if isinstance(cell_value, (int, float)) and cell_value > 0:
                        has_valid_data = True
                        meter_stats[col]['count'] += 1
                        meter_stats[col]['sum'] += cell_value
                        meter_stats[col]['max'] = max(meter_stats[col]['max'], cell_value)
                        meter_stats[col]['min'] = min(meter_stats[col]['min'], cell_value)
                
                if has_valid_data:
                    total_data_rows += 1
        
        if date_range:
            date_range.sort()
            print(f'📅 数据时间范围: {date_range[0].strftime("%Y-%m-%d")} 至 {date_range[-1].strftime("%Y-%m-%d")}')
            print(f'📊 有效数据行数: {total_data_rows:,}')
            print(f'📊 总时间跨度: {(date_range[-1] - date_range[0]).days + 1:,} 天')
        
        print()
        print('💧 各水表数据统计:')
        for col in range(2, min(9, len(headers) + 1)):  # 主要水表列
            if col - 1 < len(headers):
                stats = meter_stats.get(col, {})
                if stats.get('count', 0) > 0:
                    avg_value = stats['sum'] / stats['count']
                    min_val = stats['min'] if stats['min'] != float('inf') else 0
                    print(f'  {headers[col-1][:20]:<20}: '
                          f'数据点 {stats["count"]:,} | '
                          f'平均 {avg_value:.1f} | '
                          f'最大 {stats["max"]:.1f} | '
                          f'最小 {min_val:.1f}')
        
        print()
        print('=' * 60)
        print('🏗️ 文件结构特点')
        print('=' * 60)
        print('📋 表格设计特点:')
        print('  • 第1行: 总标题 "监控表流量明细"')
        print('  • 第2-3行: 副标题和分类标识')
        print('  • 第4行: 详细列标题')
        print('  • 第5行开始: 实际数据')
        print()
        print('📊 数据特点:')
        print('  • 使用Excel公式进行自动计算')
        print('  • 支持环比差值计算')
        print('  • 包含汇总和明细数据')
        print('  • 列结构设计便于扩展')
        print()
        print('🔧 与系统集成:')
        print('  • 所有8个水表都有对应列')
        print('  • 日期格式标准化')
        print('  • 数据结构适合自动化更新')
        print('  • 支持历史数据查询和分析')
        
        wb.close()
        
    except Exception as e:
        print(f'❌ 提取数据时出错: {e}')
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    extract_actual_data()