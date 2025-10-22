#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
检查9月统计表的公式
"""

import openpyxl
from openpyxl.utils import get_column_letter

def main():
    excel_path = "excel_exports/石滩区分区计量.xlsx"
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb["石滩区"]
    
    print("="*100)
    print("[Checking September Summary Formulas]")
    print("="*100)
    
    # 查找9月的统计表
    september_row = None
    for row in range(1, ws.max_row + 1):
        for col in range(1, 11):
            cell_value = ws.cell(row, col).value
            if cell_value and isinstance(cell_value, str) and '2025年9月' in cell_value:
                september_row = row
                print(f"\n[OK] Found September at row {september_row}")
                break
        if september_row:
            break
    
    if not september_row:
        print("[ERROR] September not found!")
        return
    
    # 分析表头
    header_row = september_row + 1
    print(f"\n[Headers] Row {header_row}:")
    headers = {}
    for col in range(1, 11):
        val = ws.cell(header_row, col).value
        if val:
            headers[col] = val
            print(f"  Col {get_column_letter(col)}: {val}")
    
    # 检查数据行
    print(f"\n[Data Rows] Starting from row {september_row + 2}:")
    
    for row_offset in range(2, 10):  # 检查前8行数据
        row = september_row + row_offset
        print(f"\n  Row {row}:")
        
        # 第一列（日期）
        date_val = ws.cell(row, 1).value
        print(f"    A{row}: {date_val}")
        
        # 检查所有列
        for col in range(2, 11):
            cell = ws.cell(row, col)
            col_letter = get_column_letter(col)
            col_name = headers.get(col, f"Col{col}")
            
            if cell.value is not None:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"    {col_letter}{row} ({col_name}): FORMULA = {cell.value}")
                else:
                    print(f"    {col_letter}{row} ({col_name}): VALUE = {cell.value}")
        
        # 如果遇到总计行，停止
        if date_val and isinstance(date_val, str) and '总计' in date_val:
            break
    
    # 查找总计行
    print(f"\n[Looking for Total Row]...")
    for row_offset in range(2, 35):
        row = september_row + row_offset
        first_cell = ws.cell(row, 1).value
        if first_cell and isinstance(first_cell, str) and ('总计' in first_cell or '合计' in first_cell):
            print(f"\n[Total Row] Row {row}:")
            for col in range(1, 11):
                cell = ws.cell(row, col)
                col_letter = get_column_letter(col)
                col_name = headers.get(col, f"Col{col}")
                
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        print(f"    {col_letter}{row} ({col_name}): FORMULA = {cell.value}")
                    else:
                        print(f"    {col_letter}{row} ({col_name}): VALUE = {cell.value}")
            break
    
    print(f"\n{'='*100}")
    print("[OK] Analysis Complete!")
    print(f"{'='*100}")

if __name__ == "__main__":
    main()

