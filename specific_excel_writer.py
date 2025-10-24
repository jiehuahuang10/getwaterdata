#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from datetime import datetime, timedelta
import time
import os

class SpecificExcelWriter:
    """专门用于写入石滩供水服务部每日总供水情况.xlsx的类"""
    
    def __init__(self, excel_path='excel_exports/石滩供水服务部每日总供水情况.xlsx'):
        self.excel_path = excel_path
        
        # 水表名称映射：系统名称 -> (Excel列号, Excel列名)
        self.meter_mapping = {
            '荔新大道DN1200流量计': (7, '荔新大道'),
            '新城大道医院DN800流量计': (8, '新城大道'),
            '三江新总表DN800（2190066）': (9, '三江新总表'),
            '宁西总表DN1200': (12, '宁西2总表'),
            '沙庄总表': (13, '沙庄总表'),
            '如丰大道600监控表': (14, '如丰大道600监控表'),
            '三棵树600监控表': (15, '三棵树600监控表'),
            '2501200108': (16, '中山西路DN300流量计')
        }
        
        print(f"初始化SpecificExcelWriter，目标文件：{self.excel_path}")
        print(f"水表映射关系：{len(self.meter_mapping)}个水表")
    
    def load_workbook_with_retry(self, max_retries=3):
        """带重试的工作簿加载"""
        for attempt in range(max_retries):
            try:
                wb = openpyxl.load_workbook(self.excel_path)
                return wb
            except PermissionError:
                if attempt < max_retries - 1:
                    print(f"文件被占用，等待2秒后重试... (尝试 {attempt + 1}/{max_retries})")
                    time.sleep(2)
                else:
                    raise
            except Exception as e:
                print(f"加载Excel文件失败: {e}")
                raise
    
    def save_workbook_with_retry(self, wb, max_retries=3):
        """带重试的工作簿保存"""
        for attempt in range(max_retries):
            try:
                wb.save(self.excel_path)
                print(f"Excel文件保存成功：{self.excel_path}")
                return True
            except PermissionError:
                if attempt < max_retries - 1:
                    print(f"文件被占用，等待2秒后重试保存... (尝试 {attempt + 1}/{max_retries})")
                    time.sleep(2)
                else:
                    print("文件被占用，无法保存。请关闭Excel文件后重试。")
                    return False
            except Exception as e:
                print(f"保存Excel文件失败: {e}")
                return False
    
    def find_date_row(self, sheet, target_date):
        """查找指定日期的行号"""
        target_date_str = target_date if isinstance(target_date, str) else target_date.strftime('%Y-%m-%d')
        
        # 从第5行开始查找（前4行是标题）
        for row in range(5, sheet.max_row + 1):
            cell_value = sheet.cell(row, 1).value  # 第1列是日期列
            
            if cell_value:
                if isinstance(cell_value, datetime):
                    date_str = cell_value.strftime('%Y-%m-%d')
                elif isinstance(cell_value, str):
                    date_str = cell_value
                else:
                    continue
                
                if date_str == target_date_str:
                    return row
        
        return None
    
    def insert_new_date_row(self, sheet, target_date):
        """插入新的日期行（按时间顺序）"""
        target_date_obj = datetime.strptime(target_date, '%Y-%m-%d') if isinstance(target_date, str) else target_date
        target_date_str = target_date_obj.strftime('%Y-%m-%d')
        
        # 找到插入位置
        insert_row = None
        for row in range(5, sheet.max_row + 1):
            cell_value = sheet.cell(row, 1).value
            if cell_value:
                if isinstance(cell_value, datetime):
                    existing_date = cell_value
                elif isinstance(cell_value, str):
                    try:
                        existing_date = datetime.strptime(cell_value, '%Y-%m-%d')
                    except:
                        continue
                else:
                    continue
                
                if target_date_obj < existing_date:
                    insert_row = row
                    break
        
        if insert_row is None:
            # 插入到最后
            insert_row = sheet.max_row + 1
        
        # 插入新行
        sheet.insert_rows(insert_row)
        
        # 设置日期
        sheet.cell(insert_row, 1, target_date_str)
        
        print(f"在第{insert_row}行插入新日期：{target_date_str}")
        return insert_row
    
    def write_water_data(self, target_date, water_data):
        """
        将水表数据写入Excel文件
        
        Args:
            target_date: 目标日期 (str 'YYYY-MM-DD' 或 datetime对象)
            water_data: 水表数据字典，格式：
                {
                    '荔新大道DN1200流量计': 12345.67,
                    '新城大道医院DN800流量计': 23456.78,
                    ...
                }
        """
        print(f"开始写入数据到Excel文件...")
        print(f"目标日期：{target_date}")
        print(f"水表数据：{len(water_data)}个水表")
        
        try:
            # 加载工作簿
            wb = self.load_workbook_with_retry()
            sheet = wb.active
            
            # 查找目标日期的行
            target_row = self.find_date_row(sheet, target_date)
            
            if target_row:
                print(f"找到现有日期行：第{target_row}行，将更新数据")
            else:
                print(f"未找到日期行，将插入新行")
                target_row = self.insert_new_date_row(sheet, target_date)
            
            # 写入水表数据
            updated_count = 0
            for system_name, value in water_data.items():
                if system_name in self.meter_mapping:
                    col_num, excel_name = self.meter_mapping[system_name]
                    
                    # 写入数据
                    if value is not None:
                        sheet.cell(target_row, col_num, value)
                        print(f"  [OK] {excel_name}(第{col_num}列): {value}")
                    else:
                        sheet.cell(target_row, col_num, None)  # 空白
                        print(f"  [EMPTY] {excel_name}(第{col_num}列): 空白")
                    
                    updated_count += 1
                else:
                    print(f"  [SKIP] 未找到映射：{system_name}")
            
            # 保存文件
            success = self.save_workbook_with_retry(wb)
            wb.close()
            
            if success:
                print(f"数据写入完成！更新了 {updated_count} 个水表的数据")
                return True
            else:
                print("数据写入失败：无法保存文件")
                return False
            
        except Exception as e:
            print(f"写入Excel数据时出错: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def get_existing_dates(self):
        """获取Excel文件中已有的日期列表"""
        try:
            wb = self.load_workbook_with_retry()
            sheet = wb.active
            
            dates = []
            for row in range(5, sheet.max_row + 1):
                cell_value = sheet.cell(row, 1).value
                if cell_value:
                    if isinstance(cell_value, datetime):
                        dates.append(cell_value.strftime('%Y-%m-%d'))
                    elif isinstance(cell_value, str):
                        dates.append(cell_value)
            
            wb.close()
            return sorted(dates)
            
        except Exception as e:
            print(f"读取现有日期时出错: {e}")
            return []

def test_specific_excel_writer():
    """测试函数"""
    writer = SpecificExcelWriter()
    
    # 测试数据
    test_data = {
        '荔新大道DN1200流量计': 12345.67,
        '新城大道医院DN800流量计': 23456.78,
        '三江新总表DN800（2190066）': 34567.89,
        '宁西总表DN1200': 45678.90,
        '沙庄总表': 56789.01,
        '如丰大道600监控表': 67890.12,
        '三棵树600监控表': 78901.23,
        '2501200108': 89012.34
    }
    
    # 测试写入今天的数据
    today = datetime.now().strftime('%Y-%m-%d')
    print(f"测试写入日期：{today}")
    
    result = writer.write_water_data(today, test_data)
    print(f"写入结果：{'成功' if result else '失败'}")

if __name__ == '__main__':
    test_specific_excel_writer()
