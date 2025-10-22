#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查日期单元格
"""

import openpyxl
from datetime import datetime

excel_path = "excel_exports/石滩区分区计量.xlsx"

wb = openpyxl.load_workbook(excel_path)
ws = wb["石滩区"]

print("检查第45-65行的A列和E列:")

for row in range(45, 66):
    a_val = ws.cell(row, 1).value
    e_val = ws.cell(row, 5).value
    
    if a_val or e_val:
        print(f"第{row}行:")
        if a_val:
            print(f"  A列: {a_val} (类型: {type(a_val).__name__})")
            # 如果是日期类型
            if isinstance(a_val, datetime):
                print(f"    日期: {a_val.strftime('%Y年%m月')}")
        if e_val:
            print(f"  E列: {e_val} (类型: {type(e_val).__name__})")
            if isinstance(e_val, datetime):
                print(f"    日期: {e_val.strftime('%Y年%m月')}")

wb.close()

