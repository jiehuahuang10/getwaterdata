#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel导出功能模块
"""

# 优先尝试导入pandas，如果失败则使用备用方案
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    print("⚠️ pandas不可用，将使用备用的Excel处理方案")
    PANDAS_AVAILABLE = False
    pd = None

from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 如果pandas可用，才导入dataframe_to_rows
if PANDAS_AVAILABLE:
    try:
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("⚠️ openpyxl.utils.dataframe不可用")
        PANDAS_AVAILABLE = False

def calculate_yesterday():
    """计算昨天的日期"""
    yesterday = datetime.now() - timedelta(days=1)
    return yesterday.strftime('%Y-%m-%d')

def extract_horizontal_data(water_data):
    """提取水表数据并按横向格式组织（水表作为列，日期作为行）"""
    if not water_data or 'data' not in water_data or 'rows' not in water_data['data']:
        return [], []
    
    rows = water_data['data']['rows']
    
    # 收集所有水表信息
    meters = []
    all_dates = set()
    
    for row in rows:
        if isinstance(row, dict):
            meter_name = row.get('Name', 'N/A')
            meters.append({
                'id': row.get('ID', 'N/A'),
                'name': meter_name,
                'diameter': row.get('MeterDiameter', 'N/A'),
                'data': row
            })
            
            # 收集所有日期
            date_columns = [key for key in row.keys() if key.startswith('202')]
            all_dates.update(date_columns)
    
    # 按日期排序
    sorted_dates = sorted(list(all_dates))
    
    # 构建横向数据结构
    horizontal_data = []
    for date in sorted_dates:
        row_data = {'日期': date}
        for meter in meters:
            value = meter['data'].get(date, None)
            row_data[meter['name']] = value if value is not None else ''
        horizontal_data.append(row_data)
    
    return horizontal_data, meters

def extract_yesterday_data(water_data):
    """从水务数据中提取昨天的用水量（保留原有功能用于兼容）"""
    yesterday = calculate_yesterday()
    
    if not water_data or 'data' not in water_data or 'rows' not in water_data['data']:
        return []
    
    rows = water_data['data']['rows']
    yesterday_data = []
    
    for i, row in enumerate(rows, 1):
        if isinstance(row, dict):
            meter_id = row.get('ID', 'N/A')
            meter_name = row.get('Name', 'N/A')
            meter_diameter = row.get('MeterDiameter', 'N/A')
            
            # 查找昨天的数据
            yesterday_value = row.get(yesterday, None)
            
            # 如果没有昨天的数据，尝试查找最近的数据
            if yesterday_value is None:
                # 查找所有日期列
                date_columns = [key for key in row.keys() if key.startswith('202')]
                date_columns.sort(reverse=True)  # 按日期倒序
                
                # 取最近的数据作为昨天的数据
                for date_col in date_columns:
                    if row.get(date_col) is not None:
                        yesterday_value = row.get(date_col)
                        yesterday = date_col  # 更新实际日期
                        break
            
            yesterday_data.append({
                '序号': i,
                '水表ID': meter_id,
                '水表名称': meter_name,
                '管径': meter_diameter,
                '日期': yesterday,
                '用水量': yesterday_value if yesterday_value is not None else 'N/A',
                '单位': '立方米'
            })
    
    return yesterday_data

def create_horizontal_excel(horizontal_data, meters, filename):
    """创建横向格式的Excel文件（水表作为列，日期作为行）"""
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "水表用水量数据"
    
    # 定义样式
    header_font = Font(name='Microsoft YaHei', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name='Microsoft YaHei', size=11)
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    
    # 第一行：日期标题
    ws['A1'] = "日期"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = header_alignment
    ws['A1'].border = border
    
    # 第一行：水表名称作为列标题
    for col_idx, meter in enumerate(meters, 2):
        cell = ws.cell(row=1, column=col_idx, value=meter['name'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 添加数据行
    for row_idx, row_data in enumerate(horizontal_data, 2):
        # 日期列
        date_cell = ws.cell(row=row_idx, column=1, value=row_data['日期'])
        date_cell.font = data_font
        date_cell.alignment = data_alignment
        date_cell.border = border
        
        # 各水表数据列
        for col_idx, meter in enumerate(meters, 2):
            value = row_data.get(meter['name'], '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
            # 数值格式化
            if isinstance(value, (int, float)) and value != '':
                cell.number_format = '#,##0.00'
                # 大数值高亮
                if value > 100000:
                    cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                elif value > 50000:
                    cell.fill = PatternFill(start_color='FFF2E6', end_color='FFF2E6', fill_type='solid')
    
    # 自动调整列宽
    ws.column_dimensions['A'].width = 15  # 日期列
    for col_idx in range(2, len(meters) + 2):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 20  # 水表数据列
    
    # 设置行高
    for row in range(1, len(horizontal_data) + 2):
        ws.row_dimensions[row].height = 25
    
    # 保存文件
    wb.save(filename)
    return filename

def create_styled_excel(data, filename):
    """创建带样式的Excel文件（原有纵向格式，保留兼容性）"""
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "水表用水量数据"
    
    # 定义样式
    header_font = Font(name='Microsoft YaHei', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name='Microsoft YaHei', size=11)
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    
    # 添加标题
    title_cell = ws['A1']
    title_cell.value = f"水务数据获取系统 - 水表用水量报表"
    title_cell.font = Font(name='Microsoft YaHei', size=16, bold=True, color='4472C4')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:G1')
    
    # 添加生成时间
    time_cell = ws['A2']
    time_cell.value = f"生成时间: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}"
    time_cell.font = Font(name='Microsoft YaHei', size=10, color='666666')
    time_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A2:G2')
    
    # 添加空行
    ws['A3'] = ""
    
    # 添加表头
    headers = ['序号', '水表ID', '水表名称', '管径', '日期', '用水量', '单位']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 添加数据
    for row_idx, row_data in enumerate(data, 5):
        for col_idx, (key, value) in enumerate(row_data.items(), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
            # 特殊处理用水量列，添加数值格式
            if key == '用水量' and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                if value > 100000:
                    cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                elif value > 50000:
                    cell.fill = PatternFill(start_color='FFF2E6', end_color='FFF2E6', fill_type='solid')
    
    # 自动调整列宽
    column_widths = {
        'A': 8,   # 序号
        'B': 18,  # 水表ID
        'C': 25,  # 水表名称
        'D': 12,  # 管径
        'E': 15,  # 日期
        'F': 15,  # 用水量
        'G': 10   # 单位
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 设置行高
    for row in range(1, len(data) + 5):
        ws.row_dimensions[row].height = 25
    
    # 添加统计信息
    stats_row = len(data) + 6
    ws[f'A{stats_row}'] = "统计信息:"
    ws[f'A{stats_row}'].font = Font(name='Microsoft YaHei', size=11, bold=True, color='4472C4')
    
    total_meters = len(data)
    valid_data_count = sum(1 for item in data if item['用水量'] != 'N/A')
    
    ws[f'A{stats_row + 1}'] = f"水表总数: {total_meters} 个"
    ws[f'A{stats_row + 2}'] = f"有效数据: {valid_data_count} 个"
    ws[f'A{stats_row + 3}'] = f"数据完整率: {(valid_data_count/total_meters*100):.1f}%"
    
    # 保存文件
    wb.save(filename)
    return filename

def export_to_excel(water_data, output_dir="excel_exports"):
    """导出水务数据到Excel文件（横向格式）"""
    try:
        # 创建输出目录
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # 提取横向数据格式
        horizontal_data, meters = extract_horizontal_data(water_data)
        
        if not horizontal_data or not meters:
            return None, "没有可导出的数据"
        
        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        date_range = water_data.get('date_range', {})
        start_date = date_range.get('start', '').replace('-', '')
        end_date = date_range.get('end', '').replace('-', '')
        
        if start_date and end_date:
            filename = os.path.join(output_dir, f"水表数据_{start_date}至{end_date}_{timestamp}.xlsx")
        else:
            filename = os.path.join(output_dir, f"水表数据_{timestamp}.xlsx")
        
        # 创建横向Excel文件
        create_horizontal_excel(horizontal_data, meters, filename)
        
        return filename, f"成功导出 {len(meters)} 个水表的数据，共 {len(horizontal_data)} 天"
        
    except Exception as e:
        return None, f"导出失败: {str(e)}"

def export_simple_csv(water_data, output_dir="excel_exports"):
    """导出简单的CSV文件（备用方案）"""
    try:
        # 创建输出目录
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # 提取昨天的数据
        yesterday_data = extract_yesterday_data(water_data)
        
        if not yesterday_data:
            return None, "没有可导出的数据"
        
        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        yesterday = calculate_yesterday().replace('-', '')
        
        if PANDAS_AVAILABLE:
            # 使用pandas导出CSV
            df = pd.DataFrame(yesterday_data)
            filename = os.path.join(output_dir, f"水表用水量_{yesterday}_{timestamp}.csv")
            df.to_csv(filename, index=False, encoding='utf-8-sig')
        else:
            # 使用内置csv模块
            import csv
            filename = os.path.join(output_dir, f"水表用水量_{yesterday}_{timestamp}.csv")
            
            with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
                if yesterday_data:
                    fieldnames = yesterday_data[0].keys()
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(yesterday_data)
        
        return filename, f"成功导出 {len(yesterday_data)} 个水表的数据"
        
    except Exception as e:
        return None, f"导出失败: {str(e)}"

def test_excel_export():
    """测试Excel导出功能"""
    # 模拟测试数据
    test_data = {
        'data': {
            'rows': [
                {
                    'ID': '1261181000263',
                    'Name': '荔新大道DN1200流量计',
                    'MeterDiameter': 'DN1200',
                    '2025-08-15': 135824.0,
                    '2025-08-14': 135596.0,
                    '2025-08-13': 139265.0
                },
                {
                    'ID': '1261181000300',
                    'Name': '新城大道医院DN800流量计',
                    'MeterDiameter': 'DN800',
                    '2025-08-15': 16501.0,
                    '2025-08-14': 15976.0,
                    '2025-08-13': 16952.0
                },
                {
                    'ID': '1262330402331',
                    'Name': '宁西总表DN1200',
                    'MeterDiameter': 'DN1200',
                    '2025-08-15': 113211.0,
                    '2025-08-14': 112823.0,
                    '2025-08-13': 116525.0
                }
            ]
        },
        'date_range': {
            'start': '2025-08-13',
            'end': '2025-08-15',
            'description': '测试数据范围'
        }
    }
    
    print("🧪 测试横向Excel导出功能...")
    filename, message = export_to_excel(test_data)
    
    if filename:
        print(f"✅ {message}")
        print(f"📄 文件路径: {filename}")
        return True
    else:
        print(f"❌ {message}")
        return False

def update_excel_with_date(excel_file_path, water_data, target_date):
    """向现有Excel文件中添加指定日期的数据"""
    import time
    
    try:
        from openpyxl import load_workbook
        
        # 检查文件是否存在
        if not os.path.exists(excel_file_path):
            return False, "Excel文件不存在"
        
        # 检查Excel临时文件，表明文件可能被占用
        temp_file = os.path.join(os.path.dirname(excel_file_path), f"~${os.path.basename(excel_file_path)}")
        if os.path.exists(temp_file):
            return False, f"Excel文件正被使用中，请关闭Excel程序后重试。临时文件: {temp_file}"
        
        # 尝试加载现有工作簿，处理权限错误
        max_retries = 3
        wb = None
        for attempt in range(max_retries):
            try:
                wb = load_workbook(excel_file_path)
                break
            except PermissionError as e:
                if attempt < max_retries - 1:
                    print(f"⚠️ 文件被占用，等待重试... ({attempt + 1}/{max_retries})")
                    time.sleep(1)
                    continue
                else:
                    return False, f"Excel文件被占用无法访问，请关闭Excel程序后重试: {str(e)}"
        
        if not wb:
            return False, "无法加载Excel文件"
            
        ws = wb.active
        
        # 检查是否已存在该日期的数据
        existing_dates = []
        for row in range(2, ws.max_row + 1):  # 从第2行开始（第1行是表头）
            date_cell = ws.cell(row=row, column=1)
            if date_cell.value:
                existing_dates.append(str(date_cell.value))
        
        if target_date in existing_dates:
            print(f"⚠️ 日期 {target_date} 已存在，将覆盖现有数据")
            # 删除现有日期的行
            for row in range(2, ws.max_row + 1):
                date_cell = ws.cell(row=row, column=1)
                if date_cell.value and str(date_cell.value) == target_date:
                    ws.delete_rows(row)
                    print(f"🗑️ 已删除现有的 {target_date} 数据")
                    break
        
        # 提取指定日期的数据
        if not water_data or 'data' not in water_data or 'rows' not in water_data['data']:
            return False, "水务数据格式不正确"
        
        rows = water_data['data']['rows']
        date_data = {}
        
        # 收集指定日期的所有水表数据（包括空值）
        for row in rows:
            if isinstance(row, dict):
                meter_name = row.get('Name', 'N/A')
                # 检查日期键是否存在，允许空值（None）
                if target_date in row:
                    value = row.get(target_date)
                    date_data[meter_name] = value  # 可能是None、数字或其他值
                    print(f"📊 水表 {meter_name}: {value if value is not None else '(空白)'}")
        
        if not date_data:
            return False, f"未找到日期 {target_date} 的任何数据记录"
        
        print(f"🔍 从数据中找到的水表: {list(date_data.keys())}")
        
        # 获取表头信息（水表名称）
        header_meters = []
        for col in range(2, ws.max_column + 1):  # 从第2列开始（第1列是日期）
            header_cell = ws.cell(row=1, column=col)
            if header_cell.value:
                header_meters.append(header_cell.value)
        
        print(f"🔍 Excel表头中的水表: {header_meters}")
        
        # 检查名称匹配情况
        matched_meters = []
        for header_meter in header_meters:
            if header_meter in date_data:
                matched_meters.append(header_meter)
            else:
                print(f"⚠️ Excel表头水表 '{header_meter}' 在数据中未找到匹配")
        
        print(f"🎯 匹配的水表数量: {len(matched_meters)}/{len(header_meters)}")
        
        # 找到合适的插入位置（按日期排序）
        insert_row = 2
        for row in range(2, ws.max_row + 1):
            date_cell = ws.cell(row=row, column=1)
            if date_cell.value and str(date_cell.value) > target_date:
                insert_row = row
                break
            insert_row = row + 1
        
        # 如果需要在中间插入，先移动现有行
        if insert_row <= ws.max_row:
            ws.insert_rows(insert_row)
        
        # 定义样式
        data_font = Font(name='Microsoft YaHei', size=11)
        data_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )
        
        # 添加日期
        date_cell = ws.cell(row=insert_row, column=1, value=target_date)
        date_cell.font = data_font
        date_cell.alignment = data_alignment
        date_cell.border = border
        
        # 添加各水表数据
        for col_idx, meter_name in enumerate(header_meters, 2):
            value = date_data.get(meter_name, None)
            # 如果值是None，在Excel中显示为空白
            display_value = value if value is not None else ""
            cell = ws.cell(row=insert_row, column=col_idx, value=display_value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
            # 如果是空白值，可以添加特殊样式
            if value is None:
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # 浅灰色背景
            
            # 数值格式化和高亮
            if isinstance(value, (int, float)) and value != '':
                cell.number_format = '#,##0.00'
                if value > 100000:
                    cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                elif value > 50000:
                    cell.fill = PatternFill(start_color='FFF2E6', end_color='FFF2E6', fill_type='solid')
        
        # 保存文件（处理权限问题）
        max_save_retries = 3
        for attempt in range(max_save_retries):
            try:
                wb.save(excel_file_path)
                break
            except PermissionError as e:
                if attempt < max_save_retries - 1:
                    print(f"⚠️ 保存文件被占用，等待重试... ({attempt + 1}/{max_save_retries})")
                    time.sleep(1)
                    continue
                else:
                    return False, f"保存Excel文件被占用，请关闭Excel程序后重试: {str(e)}"
        
        return True, f"成功添加日期 {target_date} 的数据，包含 {len(date_data)} 个水表数据"
        
    except Exception as e:
        return False, f"更新Excel文件失败: {str(e)}"

def get_excel_existing_dates(excel_file_path):
    """获取Excel文件中已存在的日期列表"""
    try:
        from openpyxl import load_workbook
        
        if not os.path.exists(excel_file_path):
            return []
        
        wb = load_workbook(excel_file_path)
        ws = wb.active
        
        existing_dates = []
        for row in range(2, ws.max_row + 1):  # 从第2行开始
            date_cell = ws.cell(row=row, column=1)
            if date_cell.value:
                existing_dates.append(str(date_cell.value))
        
        return sorted(existing_dates)
        
    except Exception as e:
        print(f"读取Excel文件日期失败: {str(e)}")
        return []

if __name__ == "__main__":
    test_excel_export()
