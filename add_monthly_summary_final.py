#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
在每月1号自动添加统计表 - 最终版本
根据实际表格结构实现
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import copy
from datetime import date

def add_summary_to_september():
    """
    在9月工作表底部添加统计表
    """
    excel_path = "excel_exports/石滩区分区计量.xlsx"
    
    print("=" * 80)
    print("[START] 在9月工作表添加统计表")
    print("=" * 80)
    
    # 加载工作簿
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["石滩2025年9月"]
    
    print(f"[INFO] 当前工作表: {ws.title}")
    print(f"[INFO] 当前最大行: {ws.max_row}")
    
    # 找到现有的统计行（第11行是合计行）
    summary_template_row = 11
    print(f"[INFO] 使用第{summary_template_row}行作为模板")
    
    # 显示模板行内容
    print("\n[INFO] 模板行内容:")
    for col in range(1, 11):
        val = ws.cell(summary_template_row, col).value
        if val:
            print(f"  列{col}: {val}")
    
    # 在最后一行下面添加新的统计表
    # 根据分析，统计表包含：
    # - 第1行：空行（分隔）
    # - 第2行：标题行（合并单元格）
    # - 第3-6行：数据行（荔城、嘉园、石滩）
    # - 第7行：合计行
    
    start_row = ws.max_row + 2  # 留一行空行
    
    print(f"\n[ADD] 新统计表起始行: {start_row}")
    
    # 1. 添加标题行（类似第2行）
    title_row = start_row
    title_cell = ws.cell(title_row, 2)
    title_cell.value = "石滩10月分区计量统计"  # 下个月的标题
    
    # 复制第2行的格式
    source_title = ws.cell(2, 2)
    title_cell.font = copy.copy(source_title.font)
    title_cell.alignment = copy.copy(source_title.alignment)
    title_cell.fill = copy.copy(source_title.fill)
    
    # 合并单元格 B到G
    ws.merge_cells(f'B{title_row}:G{title_row}')
    
    print(f"[OK] 添加标题行: 第{title_row}行")
    
    # 2. 添加空行
    start_row += 1
    
    # 3. 添加时间行（第4-5行的格式）
    time_row1 = start_row + 1
    ws.cell(time_row1, 2).value = "数据时间"
    ws.cell(time_row1, 3).value = "待填写"
    
    time_row2 = start_row + 2
    ws.cell(time_row2, 2).value = "数据读取值时间"
    ws.cell(time_row2, 3).value = "待填写"
    
    # 4. 添加表头行（第7行的格式）
    header_row = start_row + 4
    headers = ["", "", "售水量", "供水量", "损耗水量", "水损耗(百分比)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col)
        cell.value = header
        # 复制第7行的格式
        source_cell = ws.cell(7, col)
        cell.font = copy.copy(source_cell.font)
        cell.alignment = copy.copy(source_cell.alignment)
        cell.border = copy.copy(source_cell.border)
        cell.fill = copy.copy(source_cell.fill)
    
    print(f"[OK] 添加表头行: 第{header_row}行")
    
    # 5. 添加数据行（荔城、嘉园、石滩）
    data_rows = [
        ("荔城", 0, 0, 0, 0),
        ("嘉园", 0, 0, 0, 0),
        ("石滩", 0, 0, 0, 0)
    ]
    
    current_row = header_row + 1
    for area_name, val1, val2, val3, val4 in data_rows:
        ws.cell(current_row, 2).value = area_name
        ws.cell(current_row, 3).value = val1
        ws.cell(current_row, 4).value = val2
        ws.cell(current_row, 5).value = val3
        ws.cell(current_row, 6).value = val4
        
        # 复制格式
        for col in range(2, 7):
            cell = ws.cell(current_row, col)
            source_cell = ws.cell(8, col)  # 使用第8行（荔城行）作为模板
            cell.font = copy.copy(source_cell.font)
            cell.alignment = copy.copy(source_cell.alignment)
            cell.border = copy.copy(source_cell.border)
            if col >= 3:  # 数字列
                cell.number_format = source_cell.number_format
        
        print(f"[OK] 添加数据行: 第{current_row}行 - {area_name}")
        current_row += 1
    
    # 6. 添加合计行（第11行的格式）
    summary_row = current_row
    ws.cell(summary_row, 2).value = "合计"
    ws.cell(summary_row, 3).value = 0
    ws.cell(summary_row, 4).value = 0
    ws.cell(summary_row, 5).value = 0
    ws.cell(summary_row, 6).value = 0
    
    # 复制第11行的格式
    for col in range(2, 7):
        cell = ws.cell(summary_row, col)
        source_cell = ws.cell(11, col)
        cell.font = copy.copy(source_cell.font)
        cell.alignment = copy.copy(source_cell.alignment)
        cell.border = copy.copy(source_cell.border)
        cell.fill = copy.copy(source_cell.fill)
        if col >= 3:
            cell.number_format = source_cell.number_format
    
    print(f"[OK] 添加合计行: 第{summary_row}行")
    
    # 7. 添加下半部分（监控表供水量）
    # 这部分从第14行开始
    separator_row = summary_row + 3
    ws.cell(separator_row, 3).value = "监控表供水量"
    
    # 复制第14行格式
    source_cell = ws.cell(14, 3)
    ws.cell(separator_row, 3).font = copy.copy(source_cell.font)
    ws.cell(separator_row, 3).alignment = copy.copy(source_cell.alignment)
    
    # 添加表头（第15行）
    header2_row = separator_row + 1
    headers2 = ["", "荔新大道", "宁西总表（插入式）DN1200", "如丰大道", "新城大道医院NB", "三樵仔", "供水量", "售水量", "损耗水量", "水损耗（百分比）"]
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(header2_row, col)
        cell.value = header
        # 复制第15行格式
        source_cell = ws.cell(15, col)
        cell.font = copy.copy(source_cell.font)
        cell.alignment = copy.copy(source_cell.alignment)
        cell.border = copy.copy(source_cell.border)
        cell.fill = copy.copy(source_cell.fill)
    
    print(f"[OK] 添加监控表表头: 第{header2_row}行")
    
    # 添加数据行（3行）
    data2_rows = [
        ("1区", 0, 0, 0, 0, 0, 0, 0, 0, 0),
        ("2区", 0, 0, 0, 0, 0, 0, 0, 0, 0),
        ("3区", 0, 0, 0, 0, 0, 0, 0, 0, 0)
    ]
    
    current_row = header2_row + 1
    for row_data in data2_rows:
        for col, val in enumerate(row_data, 1):
            ws.cell(current_row, col).value = val
            # 复制第16行格式
            source_cell = ws.cell(16, col)
            ws.cell(current_row, col).font = copy.copy(source_cell.font)
            ws.cell(current_row, col).alignment = copy.copy(source_cell.alignment)
            ws.cell(current_row, col).border = copy.copy(source_cell.border)
            if col >= 2 and isinstance(val, (int, float)):
                ws.cell(current_row, col).number_format = source_cell.number_format
        
        print(f"[OK] 添加监控表数据行: 第{current_row}行")
        current_row += 1
    
    # 保存到新文件
    new_path = "excel_exports/石滩区分区计量_已添加统计表.xlsx"
    print(f"\n[SAVE] 保存工作簿到新文件...")
    wb.save(new_path)
    print(f"[SUCCESS] 成功保存到: {new_path}")
    
    print("\n" + "=" * 80)
    print("[COMPLETE] 统计表添加完成！")
    print("=" * 80)
    print(f"新统计表位置: 第{title_row}行 到 第{current_row-1}行")
    print(f"总共添加: {current_row - title_row} 行")
    
    return True

if __name__ == "__main__":
    try:
        add_summary_to_september()
        print("\n[SUCCESS] 测试成功！请打开Excel文件查看效果。")
    except Exception as e:
        print(f"\n[ERROR] 测试失败: {e}")
        import traceback
        traceback.print_exc()

