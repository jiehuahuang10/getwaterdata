#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
from datetime import datetime

data_file = "excel_exports/石滩供水服务部每日总供水情况.xlsx"
wb = openpyxl.load_workbook(data_file, data_only=True)
ws = wb.worksheets[2]

print("Checking for 2025-09-25 data...")

target_date = datetime(2025, 9, 25)

for row in range(2800, 2900):
    date_val = ws.cell(row, 1).value
    if isinstance(date_val, datetime) and date_val == target_date:
        print(f"\nFound 2025-09-25 at row {row}")
        print(f"  Col7 (荔新大道): {ws.cell(row, 7).value}")
        print(f"  Col12 (宁西2总表): {ws.cell(row, 12).value}")
        print(f"  Col14 (如丰大道): {ws.cell(row, 14).value}")
        
        # 检查是否有非零值
        has_data = False
        for col in [7, 12, 14, 8, 15]:
            val = ws.cell(row, col).value
            if val and isinstance(val, (int, float)) and val != 0:
                has_data = True
                break
        
        print(f"  Has non-zero data: {has_data}")
        break

