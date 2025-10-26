#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl

excel_path = 'excel_exports/石滩区分区计量.xlsx'

wb = openpyxl.load_workbook(excel_path)
ws = wb['石滩区']

print("=== 检查第25-43行的所有单元格 ===\n")

for row_idx in range(25, min(44, ws.max_row + 1)):
    print(f"\n--- 第{row_idx}行 ---")
    has_data = False
    for col_idx in range(1, min(15, ws.max_column + 1)):
        cell = ws.cell(row_idx, col_idx)
        if cell.value:
            has_data = True
            print(f"  列{col_idx}: {repr(cell.value)} (类型: {type(cell.value).__name__})")
    if not has_data:
        print("  (空行)")

# 检查合并单元格
print("\n\n=== 检查合并单元格 ===")
for merged_range in ws.merged_cells.ranges:
    cell = ws.cell(merged_range.min_row, merged_range.min_col)
    if cell.value:
        print(f"合并区域 {merged_range}: {repr(cell.value)}")

wb.close()

