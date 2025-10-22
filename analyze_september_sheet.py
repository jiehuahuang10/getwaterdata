#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
详细分析9月工作表
"""

import openpyxl

excel_path = "excel_exports/石滩区分区计量.xlsx"

print("=" * 80)
print("详细分析9月工作表")
print("=" * 80)

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb["石滩2025年9月"]

print(f"工作表: {ws.title}")
print(f"行数: {ws.max_row}")
print(f"列数: {ws.max_column}")
print()

# 显示所有非空单元格
print("所有非空单元格:")
non_empty_count = 0

for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        if cell.value is not None:
            non_empty_count += 1
            # 只显示前50个
            if non_empty_count <= 50:
                print(f"  [{row},{col}] = {cell.value} (类型: {type(cell.value).__name__})")

print(f"\n总计非空单元格: {non_empty_count}")

# 检查合并单元格
print(f"\n合并单元格区域:")
for merged in ws.merged_cells.ranges:
    print(f"  {merged}")
    # 获取合并区域的值
    min_row, min_col, max_row, max_col = merged.min_row, merged.min_col, merged.max_row, merged.max_col
    value = ws.cell(min_row, min_col).value
    if value:
        print(f"    值: {value}")

wb.close()

