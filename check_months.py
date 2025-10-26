#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import re

excel_path = 'excel_exports/石滩区分区计量表.xlsx'
wb = openpyxl.load_workbook(excel_path)
ws = wb['石滩区']

print(f"工作表总行数: {ws.max_row}")
print("\n=== 搜索包含'年'和'月'的单元格 ===\n")

found_months = []

# 从后往前搜索
for row in range(ws.max_row, max(0, ws.max_row - 30), -1):
    for col in range(1, 10):
        cell_value = ws.cell(row, col).value
        if cell_value and isinstance(cell_value, str):
            if '年' in cell_value and '月' in cell_value:
                match = re.search(r'(\d{4})年(\d{1,2})月', cell_value)
                if match:
                    print(f"第{row}行, 第{col}列: {cell_value}")
                    if cell_value not in found_months:
                        found_months.append(cell_value)

print(f"\n找到的月份: {found_months}")
print(f"最后月份: {found_months[0] if found_months else '未找到'}")

wb.close()

