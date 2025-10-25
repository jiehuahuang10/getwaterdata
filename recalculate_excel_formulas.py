#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
强制重新计算 Excel 文件中的所有公式并保存
这个脚本会让 openpyxl 重新计算公式并缓存结果
"""
import openpyxl
from openpyxl.utils import get_column_letter
import os

EXCEL_PATH = "excel_exports/石滩供水服务部每日总供水情况.xlsx"

print("=" * 80)
print("强制重新计算 Excel 公式")
print("=" * 80)

if not os.path.exists(EXCEL_PATH):
    print(f"错误: 文件不存在: {EXCEL_PATH}")
    exit(1)

# 备份原文件
backup_path = EXCEL_PATH.replace('.xlsx', '_backup.xlsx')
import shutil
shutil.copy2(EXCEL_PATH, backup_path)
print(f"[OK] 已备份原文件到: {backup_path}")

# 加载工作簿（不使用 data_only，这样可以读取公式）
print(f"\n[>>] 正在加载Excel文件...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
ws = wb.active

print(f"[INFO] 工作表信息:")
print(f"   - 总行数: {ws.max_row}")
print(f"   - 总列数: {ws.max_column}")

# 统计公式单元格数量
formula_count = 0
print(f"\n[>>] 正在扫描公式单元格...")

for row in ws.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            formula_count += 1

print(f"[OK] 找到 {formula_count} 个公式单元格")

# 设置工作簿为需要重新计算
wb.calculation.calcMode = "auto"
wb.calculation.fullCalcOnLoad = True

print(f"\n[>>] 正在保存Excel文件（触发公式重新计算）...")
temp_path = EXCEL_PATH.replace('.xlsx', '_temp.xlsx')
try:
    wb.save(temp_path)
    wb.close()
    print(f"[OK] 已保存到临时文件: {temp_path}")
    
    # 关闭工作簿后，替换原文件
    print(f"[>>] 正在替换原文件...")
    if os.path.exists(EXCEL_PATH):
        os.remove(EXCEL_PATH)
    os.rename(temp_path, EXCEL_PATH)
except PermissionError as e:
    print(f"[ERROR] 权限错误: {e}")
    print(f"[ERROR] 请关闭所有打开此Excel文件的程序（WPS、Excel等），然后重新运行此脚本")
    wb.close()
    exit(1)

print(f"[OK] Excel文件已保存！")

# 验证：重新读取并检查数据
print(f"\n[>>] 验证：重新读取数据...")
wb_verify = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws_verify = wb_verify.active

rows = list(ws_verify.iter_rows(values_only=True))
if len(rows) > 4:
    header = rows[3]
    data_row = rows[4]
    print(f"\n表头前5列: {header[:5]}")
    print(f"数据前5列: {data_row[:5]}")
    
    # 检查是否还有 None
    none_count = sum(1 for cell in data_row[:10] if cell is None)
    if none_count > 0:
        print(f"\n[WARN] 警告: 前10列中仍有 {none_count} 个空值")
        print(f"       这可能是因为公式引用的单元格也是空的")
    else:
        print(f"\n[OK] 数据读取成功！所有公式都已计算")

wb_verify.close()

print("\n" + "=" * 80)
print("[OK] 完成！请刷新网页查看效果")
print("=" * 80)

