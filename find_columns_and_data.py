#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
查找数据源文件的列映射和2025年数据
"""

import openpyxl
from datetime import datetime, timedelta

def main():
    data_file = "excel_exports/石滩供水服务部每日总供水情况.xlsx"
    
    print("="*100)
    print("[Step 1] Loading file...")
    print("="*100)
    
    wb = openpyxl.load_workbook(data_file, data_only=True)
    
    # 尝试所有工作表
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n[Analyzing Sheet] {sheet_name}")
        print(f"  Rows: {ws.max_row}, Columns: {ws.max_column}")
        
        # 读取第一行作为表头
        print(f"\n  [Headers]:")
        headers = {}
        for col in range(1, min(50, ws.max_column + 1)):
            val = ws.cell(1, col).value
            if val:
                headers[col] = str(val).strip()
                if col <= 15:  # 只显示前15列
                    print(f"    Col {col}: {val}")
        
        # 查找目标列
        print(f"\n  [Looking for Target Columns]:")
        target_columns = {
            "荔新大道": ["荔新大道", "荔新"],
            "宁西2总表": ["宁西2总表", "宁西2", "宁西总表"],
            "如丰大道600监控表": ["如丰大道600", "如丰大道", "如丰"],
            "新城大道医院NB": ["新城大道医院NB", "新城大道", "新城"],
            "三棵树600监控表": ["三棵树600", "三棵树", "三棵竹"]
        }
        
        found_mapping = {}
        for target_name, search_terms in target_columns.items():
            for col, header in headers.items():
                for term in search_terms:
                    if term in header:
                        found_mapping[target_name] = col
                        print(f"    [OK] {target_name} -> Column {col} ({header})")
                        break
                if target_name in found_mapping:
                    break
        
        # 如果找到了列映射，继续查找2025年数据
        if found_mapping:
            print(f"\n  [Searching for 2025 Data]...")
            
            # 从后往前找（最新数据在后面）
            for row in range(ws.max_row, max(1, ws.max_row - 500), -1):
                date_val = ws.cell(row, 1).value
                
                if isinstance(date_val, datetime):
                    if date_val.year == 2025:
                        print(f"    [Found] Row {row}: {date_val.strftime('%Y-%m-%d')}")
                        
                        # 显示这一行的数据
                        print(f"      Data:")
                        for target_name, col_idx in found_mapping.items():
                            val = ws.cell(row, col_idx).value
                            print(f"        {target_name}: {val}")
                        
                        # 只显示一行示例
                        break
            
            # 查找8月25日到9月24日的数据范围
            print(f"\n  [Looking for Aug 25 - Sep 24, 2025]...")
            
            start_date = datetime(2025, 8, 25)
            end_date = datetime(2025, 9, 24)
            
            date_rows = []
            for row in range(2, ws.max_row + 1):
                date_val = ws.cell(row, 1).value
                
                if isinstance(date_val, datetime):
                    if start_date <= date_val <= end_date:
                        date_rows.append({
                            'row': row,
                            'date': date_val
                        })
            
            if date_rows:
                print(f"    [OK] Found {len(date_rows)} days of data")
                print(f"    First: Row {date_rows[0]['row']} - {date_rows[0]['date'].strftime('%Y-%m-%d')}")
                print(f"    Last: Row {date_rows[-1]['row']} - {date_rows[-1]['date'].strftime('%Y-%m-%d')}")
                
                # 计算一个列的总和作为示例
                first_target = list(found_mapping.keys())[0]
                first_col = found_mapping[first_target]
                
                total = 0
                for item in date_rows:
                    val = ws.cell(item['row'], first_col).value
                    if val and isinstance(val, (int, float)):
                        total += val
                
                print(f"\n    [Example Calculation]")
                print(f"    {first_target} (Column {first_col})")
                print(f"    Sum from 2025-08-25 to 2025-09-24: {total:,.0f}")
            else:
                print(f"    [Info] No data found in Aug-Sep 2025")
                
                # 查找最新的日期
                print(f"\n  [Finding Latest Date]...")
                latest_date = None
                latest_row = None
                
                for row in range(ws.max_row, max(1, ws.max_row - 100), -1):
                    date_val = ws.cell(row, 1).value
                    if isinstance(date_val, datetime):
                        latest_date = date_val
                        latest_row = row
                        break
                
                if latest_date:
                    print(f"    Latest date: Row {latest_row} - {latest_date.strftime('%Y-%m-%d')}")
    
    print(f"\n{'='*100}")
    print("[Complete]")
    print(f"{'='*100}")

if __name__ == "__main__":
    main()

