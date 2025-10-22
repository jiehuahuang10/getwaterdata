#!/usr:bin/env python
# -*- coding: utf-8 -*-
"""
检查真实的9月统计表（第58行）
"""

import openpyxl
from openpyxl.utils import get_column_letter

def main():
    excel_path = "excel_exports/石滩区分区计量.xlsx"
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb["石滩区"]
    
    print("="*100)
    print("[Checking September (Row 58)]")
    print("="*100)
    
    start_row = 58
    
    # 查看标题行
    print(f"\n[Title Row] Row {start_row}:")
    for col in range(1, 11):
        val = ws.cell(start_row, col).value
        print(f"  Col {col}: {val} (type: {type(val).__name__})")
    
    # 查看表头行
    header_row = start_row + 1
    print(f"\n[Header Row] Row {header_row}:")
    headers = {}
    for col in range(1, 11):
        val = ws.cell(header_row, col).value
        if val:
            headers[col] = val
            print(f"  Col {get_column_letter(col)} ({col}): {val}")
    
    # 查看数据行
    print(f"\n[Data Rows]:")
    for row_offset in range(2, 8):  # 检查前6行数据
        row = start_row + row_offset
        print(f"\n  Row {row}:")
        
        has_data = False
        for col in range(1, 11):
            cell = ws.cell(row, col)
            col_letter = get_column_letter(col)
            col_name = headers.get(col, f"Col{col}")
            
            if cell.value is not None:
                has_data = True
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"    {col_letter}{row} ({col_name}): FORMULA = {cell.value}")
                else:
                    val_str = str(cell.value)[:50]  # 限制长度
                    print(f"    {col_letter}{row} ({col_name}): VALUE = {val_str}")
        
        if not has_data:
            print(f"    (empty row)")
            break
    
    # 查找总计行
    print(f"\n[Looking for Total Row]...")
    for row_offset in range(2, 10):
        row = start_row + row_offset
        first_val = ws.cell(row, 1).value
        if first_val and isinstance(first_val, str) and ('总计' in first_val or '合计' in first_val):
            print(f"\n[Total Row] Row {row}:")
            for col in range(1, 11):
                cell = ws.cell(row, col)
                col_letter = get_column_letter(col)
                col_name = headers.get(col, f"Col{col}")
                
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        print(f"    {col_letter}{row} ({col_name}): FORMULA = {cell.value}")
                    else:
                        print(f"    {col_letter}{row} ({col_name}): VALUE = {cell.value}")
            break
    
    print(f"\n{'='*100}")
    print("[OK] Complete!")
    print(f"{'='*100}")

if __name__ == "__main__":
    main()

