#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
详细读取表头和数据
"""

import openpyxl
from datetime import datetime

def main():
    data_file = "excel_exports/石滩供水服务部每日总供水情况.xlsx"
    
    print("="*100)
    print("[Reading Data Source File]")
    print("="*100)
    
    wb = openpyxl.load_workbook(data_file, data_only=True)
    
    # 重点分析第3个工作表（索引2）
    ws = wb.worksheets[2]  # 第3个工作表
    
    print(f"\n[Sheet] {ws.title}")
    print(f"  Total Rows: {ws.max_row}")
    print(f"  Total Columns: {ws.max_column}")
    
    # 读取前10行，查看表头结构
    print(f"\n[First 10 Rows]:")
    for row in range(1, 11):
        print(f"\n  Row {row}:")
        row_data = []
        for col in range(1, min(20, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if val:
                row_data.append(f"Col{col}={val}")
        if row_data:
            for item in row_data[:10]:  # 只显示前10个有值的列
                print(f"    {item}")
        else:
            print(f"    (empty)")
    
    # 尝试找到真正的表头行
    print(f"\n[Looking for Header Row with Date Column]...")
    header_row = None
    date_col = None
    
    for row in range(1, 20):
        for col in range(1, 10):
            val = ws.cell(row, col).value
            if val and isinstance(val, str):
                if '日期' in val or '时间' in val or 'date' in val.lower():
                    header_row = row
                    date_col = col
                    print(f"  [Found] Header row: {row}, Date column: {col}")
                    break
        if header_row:
            break
    
    # 如果没找到，假设从某行开始是日期
    if not header_row:
        print(f"  [Info] No explicit header found, checking for datetime values...")
        for row in range(1, 20):
            val = ws.cell(row, 1).value
            if isinstance(val, datetime):
                header_row = row - 1
                date_col = 1
                print(f"  [Found] Data starts at row {row}, assuming header at row {header_row}")
                break
    
    if header_row and date_col:
        # 读取表头
        print(f"\n[Headers at Row {header_row}]:")
        headers = {}
        for col in range(1, min(30, ws.max_column + 1)):
            val = ws.cell(header_row, col).value
            if val:
                headers[col] = str(val).strip()
                print(f"  Col {col}: {val}")
        
        # 查找数据行
        data_start_row = header_row + 1
        print(f"\n[Data Sample from Row {data_start_row}]:")
        
        for row in range(data_start_row, min(data_start_row + 5, ws.max_row + 1)):
            date_val = ws.cell(row, date_col).value
            print(f"\n  Row {row}:")
            print(f"    Date: {date_val}")
            
            # 显示前几列的数据
            for col in range(1, min(10, ws.max_column + 1)):
                val = ws.cell(row, col).value
                header = headers.get(col, f"Col{col}")
                if val is not None:
                    print(f"    {header}: {val}")
        
        # 查找2025年数据
        print(f"\n[Searching for 2025 Data]...")
        found_2025 = False
        
        for row in range(data_start_row, min(data_start_row + 3000, ws.max_row + 1)):
            date_val = ws.cell(row, date_col).value
            
            if isinstance(date_val, datetime) and date_val.year == 2025:
                if not found_2025:
                    print(f"  [Found] First 2025 data at row {row}: {date_val.strftime('%Y-%m-%d')}")
                    found_2025 = True
                
                # 显示前几行2025年的数据
                if row < data_start_row + 3000:
                    # 继续找到9月的数据
                    if date_val.month == 9 and date_val.day == 1:
                        print(f"  [Found] 2025-09-01 at row {row}")
                        
                        # 显示这一行的数据
                        print(f"\n  Data at 2025-09-01:")
                        for col, header in list(headers.items())[:10]:
                            val = ws.cell(row, col).value
                            print(f"    {header}: {val}")
                        break
    
    print(f"\n{'='*100}")
    print("[Complete]")
    print(f"{'='*100}")

if __name__ == "__main__":
    main()

