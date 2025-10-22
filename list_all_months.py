#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
列出所有月度统计表
"""

import openpyxl
import re
from datetime import datetime, timedelta

def main():
    excel_path = "excel_exports/石滩区分区计量.xlsx"
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb["石滩区"]
    
    print("="*100)
    print("[Listing All Monthly Summary Tables]")
    print("="*100)
    
    print(f"\nTotal rows: {ws.max_row}")
    
    # 遍历所有行，查找月份标题
    found_months = []
    
    for row in range(1, ws.max_row + 1):
        for col in range(1, 11):
            cell_value = ws.cell(row, col).value
            
            # 方式1：文本格式 "2025年9月"
            if cell_value and isinstance(cell_value, str):
                match = re.match(r'(\d{4})年(\d{1,2})月', cell_value)
                if match:
                    found_months.append({
                        'row': row,
                        'col': col,
                        'value': cell_value,
                        'type': 'text'
                    })
                    print(f"\n[Found] Row {row}, Col {col}: {cell_value} (text)")
                    break
            
            # 方式2：整数格式（Excel日期序列号）
            elif cell_value and isinstance(cell_value, int):
                if 40000 < cell_value < 50000:  # 合理的日期范围
                    try:
                        base = datetime(1899, 12, 30)
                        date_obj = base + timedelta(days=cell_value)
                        
                        # 检查是否在单元格A列或左侧列，且看起来像标题行
                        if col <= 6:
                            # 检查这一行的其他单元格是否也像标题
                            next_cell = ws.cell(row + 1, 2).value
                            if next_cell and isinstance(next_cell, str):
                                found_months.append({
                                    'row': row,
                                    'col': col,
                                    'value': cell_value,
                                    'date': date_obj,
                                    'type': 'date_serial'
                                })
                                print(f"\n[Found] Row {row}, Col {col}: {cell_value} (date serial = {date_obj.year}年{date_obj.month}月)")
                                break
                    except:
                        pass
    
    print(f"\n{'='*100}")
    print(f"[Summary] Found {len(found_months)} monthly summary tables")
    print(f"{'='*100}")
    
    # 分析第一个找到的表
    if found_months:
        first = found_months[0]
        print(f"\n[Analyzing First Table]")
        print(f"Title Row: {first['row']}")
        print(f"Value: {first['value']} ({first['type']})")
        
        # 检查标题行下面的表头
        header_row = first['row'] + 1
        print(f"\nHeader Row: {header_row}")
        for col in range(1, 11):
            val = ws.cell(header_row, col).value
            if val:
                print(f"  Col {col}: {val}")
        
        # 检查第一行数据
        data_row = first['row'] + 2
        print(f"\nFirst Data Row: {data_row}")
        for col in range(1, 11):
            cell = ws.cell(data_row, col)
            if cell.value is not None:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"  Col {col}: FORMULA = {cell.value}")
                else:
                    print(f"  Col {col}: VALUE = {cell.value} (type: {type(cell.value).__name__})")

if __name__ == "__main__":
    main()

