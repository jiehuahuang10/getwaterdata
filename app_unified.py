#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
水务数据管理系统 - 统一版
包含两个功能：
1. 月度统计表添加
2. 水务数据获取
"""

from flask import Flask, render_template, jsonify, request, redirect, url_for
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import copy
from datetime import datetime, timedelta
import os
import re

app = Flask(__name__)

# Excel文件路径
EXCEL_PATH = "excel_exports/石滩区分区计量.xlsx"
DATA_SOURCE_PATH = "excel_exports/石滩供水服务部每日总供水情况.xlsx"

# ==================== 辅助函数 ====================

def calculate_recent_7days():
    """计算最近7天的日期范围"""
    today = datetime.now()
    end_date = today - timedelta(days=1)  # 昨天
    start_date = end_date - timedelta(days=7)  # 昨天往前推7天
    
    return start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')

def sync_excel_to_github(file_path, commit_message):
    """
    将 Excel 文件同步到 GitHub
    
    参数:
        file_path: 文件路径
        commit_message: 提交信息
    
    返回:
        dict: {'success': bool, 'message': str}
    """
    import subprocess
    
    print("=" * 80)
    print("[GITHUB SYNC] Starting synchronization process...")
    print(f"[GITHUB SYNC] File: {file_path}")
    print(f"[GITHUB SYNC] Commit message: {commit_message}")
    print("=" * 80)
    
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            print(f"[GITHUB SYNC] ERROR: File does not exist: {file_path}")
            return {'success': False, 'message': f'文件不存在: {file_path}'}
        
        print(f"[GITHUB SYNC] File exists: {file_path}")
        
        # 获取 GitHub Token
        github_token = os.environ.get('GITHUB_TOKEN')
        if not github_token:
            print("[GITHUB SYNC] ERROR: GITHUB_TOKEN not configured")
            return {'success': False, 'message': '未配置 GITHUB_TOKEN 环境变量'}
        
        print("[GITHUB SYNC] GITHUB_TOKEN found")
        
        # 配置 Git 用户信息
        print("[GITHUB SYNC] Configuring Git user info...")
        subprocess.run(['git', 'config', 'user.email', 'render-bot@getwaterdata.com'], check=False)
        subprocess.run(['git', 'config', 'user.name', 'Render Auto Sync'], check=False)
        
        # 配置 Git remote（先检查是否存在，不存在则添加）
        repo_url = f'https://{github_token}@github.com/jiehuahuang10/getwaterdata.git'
        
        # 检查 remote 是否存在
        print("[GITHUB SYNC] Checking Git remote...")
        check_remote = subprocess.run(
            ['git', 'remote', 'get-url', 'origin'],
            capture_output=True,
            text=True
        )
        
        if check_remote.returncode != 0:
            # remote 不存在，添加它
            print("[GITHUB SYNC] Remote 'origin' not found, adding it...")
            subprocess.run(['git', 'remote', 'add', 'origin', repo_url], check=False)
        else:
            # remote 存在，更新 URL
            print("[GITHUB SYNC] Remote 'origin' exists, updating URL...")
            subprocess.run(['git', 'remote', 'set-url', 'origin', repo_url], check=False)
        
        # 确保在 main 分支上
        print("[GITHUB SYNC] Checking out main branch...")
        checkout_result = subprocess.run(
            ['git', 'checkout', 'main'],
            capture_output=True,
            text=True,
            timeout=10
        )
        if checkout_result.returncode != 0:
            print(f"[GITHUB SYNC] Checkout stderr: {checkout_result.stderr}")
        
        # Git 操作
        print(f"[GITHUB SYNC] Adding file: {file_path}")
        
        # 先检查文件状态
        status_result = subprocess.run(
            ['git', 'status', '--porcelain', file_path],
            capture_output=True,
            text=True
        )
        print(f"[GITHUB SYNC] Git status for file: {status_result.stdout.strip()}")
        
        # 强制添加文件
        add_result = subprocess.run(
            ['git', 'add', '-f', file_path],
            capture_output=True,
            text=True
        )
        print(f"[GITHUB SYNC] Add result: {add_result.returncode}")
        if add_result.stderr:
            print(f"[GITHUB SYNC] Add stderr: {add_result.stderr}")
        
        # 再次检查状态
        status_result2 = subprocess.run(
            ['git', 'status', '--porcelain', file_path],
            capture_output=True,
            text=True
        )
        print(f"[GITHUB SYNC] Git status after add: {status_result2.stdout.strip()}")
        
        print("[GITHUB SYNC] Committing changes...")
        result = subprocess.run(
            ['git', 'commit', '-m', commit_message],
            capture_output=True,
            text=True
        )
        print(f"[GITHUB SYNC] Commit result: {result.returncode}")
        if result.stdout:
            print(f"[GITHUB SYNC] Commit stdout: {result.stdout}")
        if result.stderr:
            print(f"[GITHUB SYNC] Commit stderr: {result.stderr}")
        
        # 如果没有变化，返回成功（没有需要提交的内容）
        if 'nothing to commit' in result.stdout or 'nothing to commit' in result.stderr:
            print("[GITHUB SYNC] No changes to commit")
            return {'success': True, 'message': '文件无变化，无需同步'}
        
        # 推送到 GitHub
        print("[GITHUB SYNC] Pushing to GitHub...")
        push_result = subprocess.run(
            ['git', 'push', 'origin', 'main'],
            capture_output=True,
            text=True,
            timeout=30
        )
        print(f"[GITHUB SYNC] Push result: {push_result.returncode}")
        if push_result.stdout:
            print(f"[GITHUB SYNC] Push stdout: {push_result.stdout}")
        if push_result.stderr:
            print(f"[GITHUB SYNC] Push stderr: {push_result.stderr}")
        
        if push_result.returncode != 0:
            print(f"[GITHUB SYNC] ERROR: Push failed")
            return {
                'success': False,
                'message': f'推送失败: {push_result.stderr}'
            }
        
        print("[GITHUB SYNC] SUCCESS: Synced to GitHub!")
        print("=" * 80)
        return {
            'success': True,
            'message': '成功同步到 GitHub'
        }
        
    except subprocess.TimeoutExpired:
        return {'success': False, 'message': 'Git 操作超时'}
    except Exception as e:
        return {'success': False, 'message': f'同步失败: {str(e)}'}

# ==================== 首页：功能选择 ====================

@app.route('/')
def index():
    """首页 - 显示三个功能入口"""
    print("=" * 50)
    print("ROOT ROUTE CALLED - RETURNING index_unified.html")
    print("=" * 50)
    # 强制返回主页模板，添加多层缓存控制
    response = app.make_response(render_template('index_unified.html'))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/test')
def test():
    """测试路由 - 验证部署"""
    return jsonify({
        'status': 'ok',
        'message': '新版本已部署！主页应该显示三个功能卡片。',
        'version': '2.0',
        'routes': ['/', '/summary', '/data', '/auto_update']
    })

# ==================== 功能1：月度统计表添加 ====================

@app.route('/summary')
def summary_page():
    """月度统计表页面"""
    return render_template('add_summary.html')

@app.route('/auto_update')
def auto_update_page():
    """自动更新Excel数据页面"""
    return render_template('auto_update.html')

@app.route('/get_info')
def get_info():
    """获取Excel文件信息"""
    try:
        if not os.path.exists(EXCEL_PATH):
            return jsonify({
                'success': False,
                'message': f'Excel文件不存在: {EXCEL_PATH}'
            })
        
        wb = openpyxl.load_workbook(EXCEL_PATH)
        
        if "石滩区" not in wb.sheetnames:
            return jsonify({
                'success': False,
                'message': '工作表"石滩区"不存在'
            })
        
        ws = wb["石滩区"]
        
        # 查找最后一个月份（从单元格中查找）
        last_month = None
        all_months = []  # 存储所有找到的月份 (year, month, month_str)
        
        # 遍历所有单元格，查找月份标题
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    # 情况1: 文本格式的月份 "2025年6月"
                    if isinstance(cell.value, str):
                        match = re.search(r'(\d{4})年(\d{1,2})月', cell.value)
                        if match:
                            year = int(match.group(1))
                            month = int(match.group(2))
                            month_str = f"{year}年{month}月"
                            if (year, month, month_str) not in all_months:
                                all_months.append((year, month, month_str))
                    
                    # 情况2: Excel日期序列号（整数，且在合理范围内）
                    elif isinstance(cell.value, (int, float)):
                        # Excel日期序列号通常在 40000-50000 之间（2009-2037年）
                        if 40000 <= cell.value <= 50000:
                            try:
                                from datetime import datetime, timedelta
                                excel_start = datetime(1899, 12, 30)
                                actual_date = excel_start + timedelta(days=int(cell.value))
                                # 只取每月1号的日期作为月份标题
                                if actual_date.day == 1:
                                    year = actual_date.year
                                    month = actual_date.month
                                    month_str = f"{year}年{month}月"
                                    if (year, month, month_str) not in all_months:
                                        all_months.append((year, month, month_str))
                            except:
                                pass
        
        # 按年月排序，取最新的
        if all_months:
            all_months.sort(reverse=True)  # 降序排列
            last_month = all_months[0][2]  # 取最新月份的字符串
        
        # 获取总行数
        total_rows = ws.max_row
        
        wb.close()
        
        return jsonify({
            'success': True,
            'excel_path': os.path.basename(EXCEL_PATH),
            'total_rows': ws.max_row,
            'last_month': last_month
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'读取Excel文件失败: {str(e)}'
        })

def prepare_git_before_modify():
    """在修改文件前准备 Git 环境"""
    import subprocess
    
    try:
        print("[GIT PREP] Preparing Git environment before file modification...")
        
        # 获取 GitHub Token
        github_token = os.environ.get('GITHUB_TOKEN')
        if not github_token:
            print("[GIT PREP] No GITHUB_TOKEN, skipping Git sync")
            return
        
        # 配置 Git remote
        repo_url = f'https://{github_token}@github.com/jiehuahuang10/getwaterdata.git'
        
        check_remote = subprocess.run(
            ['git', 'remote', 'get-url', 'origin'],
            capture_output=True,
            text=True
        )
        
        if check_remote.returncode != 0:
            subprocess.run(['git', 'remote', 'add', 'origin', repo_url], check=False)
        else:
            subprocess.run(['git', 'remote', 'set-url', 'origin', repo_url], check=False)
        
        # 切换到 main 分支
        subprocess.run(['git', 'checkout', 'main'], capture_output=True, timeout=10)
        
        # 拉取最新代码
        print("[GIT PREP] Pulling latest changes from GitHub...")
        pull_result = subprocess.run(
            ['git', 'pull', 'origin', 'main', '--rebase'],
            capture_output=True,
            text=True,
            timeout=30
        )
        print(f"[GIT PREP] Pull completed: {pull_result.returncode}")
        
    except Exception as e:
        print(f"[GIT PREP] Error: {str(e)}")

@app.route('/add_summary', methods=['POST'])
def add_summary():
    """添加月度统计表"""
    try:
        data = request.get_json()
        month_offset = data.get('month_offset', 1)
        sale_values = data.get('sale_values', [])
        
        # 在修改文件前，先拉取最新代码
        prepare_git_before_modify()
        
        # 导入添加函数
        from add_summary_web import add_monthly_summary_to_main
        
        result = add_monthly_summary_to_main(
            month_offset=month_offset,
            use_real_data=True,
            sale_values=sale_values
        )
        
        # 如果成功，自动同步到 GitHub
        if result.get('success'):
            result['download_url'] = '/download_excel/石滩区分区计量.xlsx'
            
            # 尝试自动同步到 GitHub
            sync_result = sync_excel_to_github(
                'excel_exports/石滩区分区计量.xlsx',
                f"自动更新: 添加{result.get('month', '月度')}统计表"
            )
            
            if sync_result['success']:
                result['message'] += ' | ✅ 已自动同步到GitHub'
                result['github_synced'] = True
            else:
                result['message'] += f" | ⚠️ GitHub同步失败: {sync_result['message']}"
                result['github_synced'] = False
        
        return jsonify(result)
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'添加统计表失败: {str(e)}'
        })

@app.route('/download_excel/<filename>')
def download_excel(filename):
    """下载 Excel 文件"""
    try:
        from flask import send_file
        file_path = f'excel_exports/{filename}'
        
        if not os.path.exists(file_path):
            return jsonify({
                'success': False,
                'message': f'文件不存在: {filename}'
            }), 404
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'下载失败: {str(e)}'
        }), 500

# ==================== 功能2：水务数据获取 ====================

@app.route('/data')
def data_page():
    """水务数据获取页面 - 使用原有的完整界面"""
    # 计算日期范围
    start_date, end_date = calculate_recent_7days()
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    return render_template('index.html', 
                         current_time=current_time,
                         start_date=start_date,
                         end_date=end_date)

# 导入水务数据获取的所有功能路由
# 这些路由来自 web_app_fixed.py
import json
import glob
import threading
import time

# 全局变量存储任务状态
task_status = {
    'running': False,
    'progress': 0,
    'message': '准备就绪',
    'data': None,
    'error': None,
    'start_time': None,
    'end_time': None
}

@app.route('/get_data')
def get_data():
    """获取现有数据（兼容原有接口）"""
    # 查找最新的JSON文件
    try:
        json_files = glob.glob('WEB_COMPLETE_8_METERS_*.json')
        if json_files:
            json_files.sort(reverse=True)
            latest_file = json_files[0]
            with open(latest_file, 'r', encoding='utf-8') as f:
                file_data = json.load(f)
            
            # 前端逻辑分析:
            # 1. 第1610行: if (result.success && result.data) displayData(result.data)
            # 2. 第1147行: if (!data.data || !data.data.rows)
            # 
            # 数据流: result.data 传给 displayData, 然后访问 data.data.rows
            # 所以需要: result.data.data.rows 存在
            # 即: result = {success: true, data: {data: {rows: [...]}}}
            #
            # JSON文件是: {success: true, data: {rows: [...]}}
            # 所以需要包装一层
            # 调试：打印返回的数据结构
            result = {
                'success': True,
                'data': file_data  # 包含整个JSON,这样result.data.data.rows就能访问到
            }
            print(f"DEBUG: 返回数据有success? {'success' in result}")
            print(f"DEBUG: success值: {result.get('success')}")
            print(f"DEBUG: 有data? {'data' in result}")
            print(f"DEBUG: data有data? {'data' in result.get('data', {})}")
            return jsonify(result)
        else:
            return jsonify({'success': False, 'message': '暂无数据'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/status')
def get_status():
    """获取任务状态"""
    return jsonify(task_status)

@app.route('/start_task', methods=['POST'])
def start_task():
    """启动数据获取任务"""
    global task_status
    
    if task_status['running']:
        return jsonify({'success': False, 'message': '任务正在运行中'})
    
    # 重置状态
    task_status = {
        'running': True,
        'progress': 0,
        'message': '任务已启动...',
        'data': None,
        'error': None,
        'start_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'end_time': None
    }
    
    # 在后台线程运行任务
    thread = threading.Thread(target=run_data_task)
    thread.daemon = True
    thread.start()
    
    return jsonify({'success': True, 'message': '任务已启动'})

def run_data_task():
    """执行数据获取任务 - 从水务系统获取真实数据"""
    global task_status
    
    try:
        task_status['running'] = True
        task_status['progress'] = 0
        task_status['message'] = '开始获取数据...'
        task_status['error'] = None
        task_status['start_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        task_status['data'] = None
        
        # 更新进度
        task_status['progress'] = 10
        task_status['message'] = '正在登录系统...'
        time.sleep(1)
        
        # 直接调用数据获取函数
        task_status['progress'] = 30
        task_status['message'] = '正在获取数据...'
        
        # 导入必要的库
        import requests
        from bs4 import BeautifulSoup
        import hashlib
        
        session = requests.Session()
        
        # 登录
        task_status['progress'] = 40
        task_status['message'] = '正在登录水务系统...'
        
        login_url = "http://axwater.dmas.cn/Login.aspx"
        login_page = session.get(login_url)
        
        soup = BeautifulSoup(login_page.text, 'html.parser')
        form = soup.find('form')
        form_data = {}
        
        for input_elem in form.find_all('input'):
            name = input_elem.get('name')
            value = input_elem.get('value', '')
            if name:
                form_data[name] = value
        
        username = '13509288500'
        password = '288500'
        form_data['user'] = username
        form_data['pwd'] = hashlib.md5(password.encode('utf-8')).hexdigest()
        
        login_response = session.post(login_url, data=form_data)
        
        if "window.location='frmMain.aspx'" in login_response.text:
            task_status['progress'] = 60
            task_status['message'] = '登录成功，正在访问报表页面...'
            
            # 跳转到主页面
            main_url = "http://axwater.dmas.cn/frmMain.aspx"
            main_response = session.get(main_url)
            
            if main_response.status_code == 200:
                # 访问报表页面
                task_status['progress'] = 70
                task_status['message'] = '正在获取水表数据...'
                time.sleep(1)
                
                report_url = "http://axwater.dmas.cn/reports/FluxRpt.aspx"
                report_response = session.get(report_url)
                
                if '登录超时' not in report_response.text:
                    # 获取数据
                    task_status['progress'] = 80
                    task_status['message'] = '正在调用数据API...'
                    
                    start_date, end_date = calculate_recent_7days()
                    
                    meter_ids = [
                        '1261181000263', '1261181000300', '1262330402331',
                        '2190066', '2190493', '2501200108', '2520005', '2520006'
                    ]
                    
                    formatted_node_ids = "'" + "','".join(meter_ids) + "'"
                    
                    api_url = "http://axwater.dmas.cn/reports/ashx/getRptWaterYield.ashx"
                    api_params = {
                        'nodeId': formatted_node_ids,
                        'startDate': start_date,
                        'endDate': end_date,
                        'meterType': '-1',
                        'statisticsType': 'flux',
                        'type': 'dayRpt'
                    }
                    
                    api_headers = {
                        'Referer': report_url,
                        'X-Requested-With': 'XMLHttpRequest',
                        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                    }
                    
                    api_response = session.post(api_url, data=api_params, headers=api_headers)
                    
                    if api_response.text and len(api_response.text) > 10:
                        try:
                            data = api_response.json()
                            
                            # 保存数据
                            timestamp = time.strftime('%Y%m%d_%H%M%S')
                            filename = f"WEB_COMPLETE_8_METERS_{timestamp}.json"
                            
                            output_data = {
                                'timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
                                'source': 'web_interface_scraper',
                                'success': True,
                                'data_type': 'json',
                                'calculation_date': datetime.now().strftime('%Y-%m-%d'),
                                'date_range': {
                                    'start': start_date,
                                    'end': end_date,
                                    'description': '昨天往前推7天的数据'
                                },
                                'meter_count': len(meter_ids),
                                'data': data,
                                'note': f'通过Web界面获取的{datetime.now().strftime("%Y年%m月%d日")}最新数据'
                            }
                            
                            with open(filename, 'w', encoding='utf-8') as f:
                                json.dump(output_data, f, ensure_ascii=False, indent=2)
                            
                            task_status['data'] = output_data
                            task_status['progress'] = 100
                            task_status['message'] = f'数据获取成功！共获取{len(data.get("rows", []))}个水表数据'
                            
                        except json.JSONDecodeError:
                            task_status['error'] = 'API返回的不是有效的JSON数据'
                            task_status['message'] = '数据解析失败'
                    else:
                        task_status['error'] = 'API返回空响应'
                        task_status['message'] = '数据获取失败'
                else:
                    task_status['error'] = '访问报表页面时显示登录超时'
                    task_status['message'] = '会话已过期'
            else:
                task_status['error'] = '无法访问主页面'
                task_status['message'] = '系统访问失败'
        else:
            task_status['error'] = '登录失败，请检查账号密码'
            task_status['message'] = '登录失败'
            
    except Exception as e:
        task_status['error'] = str(e)
        task_status['message'] = f'❌ 获取失败: {str(e)}'
    finally:
        task_status['running'] = False
        task_status['end_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

@app.route('/history')
def history_page():
    """历史数据页面"""
    return render_template('history.html')

@app.route('/get_history')
def get_history():
    """获取历史数据文件列表"""
    try:
        json_files = glob.glob('WEB_COMPLETE_8_METERS_*.json')
        json_files.sort(reverse=True)
        
        history_data = []
        for file in json_files[:20]:  # 最多返回20条
            try:
                with open(file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    history_data.append({
                        'file': file,
                        'date': data.get('export_time', '未知'),
                        'count': len(data.get('data', []))
                    })
            except:
                continue
        
        return jsonify({'success': True, 'data': history_data})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/task_status')
def task_status_route():
    """获取任务状态（兼容路由）"""
    return jsonify(task_status)

@app.route('/export_excel', methods=['POST'])
def export_excel():
    """导出Excel（功能暂不可用）"""
    return jsonify({'success': False, 'message': 'Excel导出功能需要额外的依赖包'})

@app.route('/update_specific_excel', methods=['POST'])
def update_specific_excel():
    """更新特定Excel（功能暂不可用）"""
    return jsonify({'success': False, 'message': 'Excel更新功能需要额外的依赖包'})

@app.route('/get_excel_files')
def get_excel_files():
    """获取Excel文件列表"""
    try:
        excel_files = glob.glob('excel_exports/*.xlsx')
        files_info = []
        for file in excel_files:
            files_info.append({
                'name': os.path.basename(file),
                'path': file
            })
        return jsonify({'success': True, 'files': files_info})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_available_dates')
def get_available_dates():
    """获取可用的日期列表"""
    try:
        # 返回最近30天的日期
        dates = []
        for i in range(30):
            date = datetime.now() - timedelta(days=i)
            dates.append(date.strftime('%Y-%m-%d'))
        return jsonify({'success': True, 'dates': dates})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/update_excel_date', methods=['POST'])
def update_excel_date():
    """按日期更新Excel（功能暂不可用）"""
    return jsonify({'success': False, 'message': 'Excel更新功能需要额外的依赖包'})

# ==================== 功能3：自动更新Excel ====================

@app.route('/execute_auto_update', methods=['POST'])
def execute_auto_update():
    """执行自动更新Excel任务"""
    try:
        data = request.get_json()
        target_date = data.get('date')
        
        if not target_date:
            return jsonify({'success': False, 'message': '请提供目标日期'})
        
        # 导入更新模块
        from integrated_excel_updater import update_excel_with_real_data
        
        # 执行更新
        result = update_excel_with_real_data(target_date)
        
        # 如果更新成功，重新保存Excel以保留公式
        if result.get('success'):
            try:
                excel_path = os.path.abspath(DATA_SOURCE_PATH)
                print(f"[INFO] 重新保存Excel文件以保留公式: {excel_path}")
                
                # 以保留公式的方式重新加载和保存
                wb = openpyxl.load_workbook(excel_path, data_only=False)
                wb.calculation.calcMode = 'auto'
                wb.calculation.fullCalcOnLoad = True
                wb.save(excel_path)
                wb.close()
                
                print("[SUCCESS] Excel文件已重新保存，公式已保留")
                result['formula_preserved'] = True
                result['note'] = '公式已保留，将在Excel中打开时自动计算'
                
            except Exception as save_error:
                print(f"[WARNING] 重新保存失败: {save_error}")
                result['formula_preserved'] = False
                result['save_warning'] = f'数据已更新，但重新保存失败: {str(save_error)}'
        
        return jsonify(result)
        
    except ImportError as e:
        return jsonify({
            'success': False,
            'message': f'缺少必要的模块: {str(e)}'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'更新失败: {str(e)}'
        })

# ==================== 功能4：在线查看Excel表格 ====================

@app.route('/view_excel')
def view_excel():
    """显示Excel在线查看页面（AG-Grid版本）"""
    return render_template('view_excel_aggrid_new.html')

@app.route('/view_excel_old')
def view_excel_old():
    """显示Excel在线查看页面（旧版本）"""
    return render_template('view_excel.html')

@app.route('/view_excel_pro')
def view_excel_pro():
    """显示Excel在线查看页面（Handsontable专业版）"""
    return render_template('view_excel_handsontable.html')

@app.route('/view_excel_grid')
def view_excel_grid():
    """显示Excel在线查看页面（AG-Grid专业版 - 推荐）"""
    return render_template('view_excel_aggrid.html')

# ==================== 功能5：月报和季报统计 ====================

@app.route('/reports')
def reports_hub():
    """统计报表中心"""
    return render_template('reports_hub.html')

@app.route('/monthly_report')
def monthly_report():
    """月报统计页面"""
    return render_template('monthly_report.html')

@app.route('/monthly_report_aggrid')
def monthly_report_aggrid():
    """月报统计页面 (AG-Grid版本)"""
    return render_template('monthly_report_aggrid.html')

@app.route('/quarterly_report')
def quarterly_report():
    """季报统计页面"""
    return render_template('quarterly_report.html')

@app.route('/quarterly_report_aggrid')
def quarterly_report_aggrid():
    """季报统计页面 (AG-Grid版本)"""
    return render_template('quarterly_report_aggrid.html')

@app.route('/weekly_report_aggrid')
def weekly_report_aggrid():
    """周报统计页面 (AG-Grid版本)"""
    return render_template('weekly_report_aggrid.html')

@app.route('/api/get_weekly_stats')
def get_weekly_stats():
    """获取周度统计数据"""
    try:
        year = request.args.get('year', str(datetime.now().year), type=str)
        week = request.args.get('week', '1', type=str)
        
        excel_path = DATA_SOURCE_PATH
        if not os.path.exists(excel_path):
            return jsonify({'success': False, 'message': 'Excel文件不存在'})
        
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        
        # 读取表头（第4行）
        all_rows = list(ws.iter_rows(values_only=True))
        header = all_rows[3]
        data_rows = all_rows[4:]
        
        # 计算指定周的日期范围
        year_int = int(year)
        week_int = int(week)
        
        # 获取该年第一天
        jan_1 = datetime(year_int, 1, 1)
        # 计算到第一个周一的天数
        days_to_monday = (7 - jan_1.weekday()) % 7
        first_monday = jan_1 + timedelta(days=days_to_monday)
        
        # 计算指定周的开始和结束日期
        week_start = first_monday + timedelta(weeks=week_int - 1)
        week_end = week_start + timedelta(days=6)
        
        # 筛选指定周的数据
        weekly_data = []
        for row in data_rows:
            date_val = row[0]
            if isinstance(date_val, datetime):
                if week_start <= date_val <= week_end:
                    weekly_data.append(row)
        
        wb.close()
        
        # 格式化日期范围
        date_range = f"{week_start.strftime('%m月%d日')} - {week_end.strftime('%m月%d日')}"
        
        if not weekly_data:
            return jsonify({
                'success': True,
                'data': [],
                'summary': [],
                'stats': [],
                'date_range': date_range,
                'message': '该周暂无数据'
            })
        
        # 计算统计数据
        stats = []
        summary_data = [['指标名称'] + list(header[1:])]  # 表头
        
        # 统计指标
        stat_names = ['总计', '平均值', '最大值', '最小值']
        
        for stat_name in stat_names:
            row_data = [stat_name]
            for col_idx in range(1, len(header)):
                values = []
                for row in weekly_data:
                    if col_idx < len(row) and row[col_idx] is not None:
                        try:
                            val = float(row[col_idx])
                            if val != 0:  # 排除0值
                                values.append(val)
                        except (ValueError, TypeError):
                            pass
                
                if values:
                    if stat_name == '总计':
                        row_data.append(sum(values))
                    elif stat_name == '平均值':
                        row_data.append(sum(values) / len(values))
                    elif stat_name == '最大值':
                        row_data.append(max(values))
                    elif stat_name == '最小值':
                        row_data.append(min(values))
                else:
                    row_data.append(0)
            
            summary_data.append(row_data)
        
        # 生成统计卡片数据（取前4列的总计）
        if len(weekly_data) > 0:
            for col_idx in range(1, min(5, len(header))):
                col_name = header[col_idx] if col_idx < len(header) else f'列{col_idx}'
                values = []
                for row in weekly_data:
                    if col_idx < len(row) and row[col_idx] is not None:
                        try:
                            val = float(row[col_idx])
                            if val != 0:
                                values.append(val)
                        except (ValueError, TypeError):
                            pass
                
                if values:
                    stats.append({
                        'name': col_name,
                        'value': round(sum(values)),
                        'unit': '周累计 (m³)'
                    })
        
        # 返回数据
        return jsonify({
            'success': True,
            'data': [list(header)] + weekly_data,
            'summary': summary_data,
            'stats': stats,
            'date_range': date_range
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'查询失败: {str(e)}'
        })

@app.route('/api/get_monthly_stats')
def get_monthly_stats():
    """获取月度统计数据"""
    try:
        year = request.args.get('year', str(datetime.now().year), type=str)
        month = request.args.get('month', str(datetime.now().month), type=str)
        
        excel_path = DATA_SOURCE_PATH
        if not os.path.exists(excel_path):
            return jsonify({'success': False, 'message': 'Excel文件不存在'})
        
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        
        # 读取表头（第4行）
        all_rows = list(ws.iter_rows(values_only=True))
        header = all_rows[3]
        data_rows = all_rows[4:]
        
        # 筛选指定年月的数据
        year_int = int(year)
        month_int = int(month)
        monthly_data = []
        
        for row in data_rows:
            date_val = row[0]
            if isinstance(date_val, datetime) and date_val.year == year_int and date_val.month == month_int:
                monthly_data.append(row)
        
        wb.close()
        
        if not monthly_data:
            return jsonify({
                'success': True,
                'data': [],
                'header': list(header),
                'summary': {},
                'message': f'未找到{year}年{month}月的数据'
            })
        
        # 计算统计数据（对数值列求和、平均等）
        summary = {}
        for col_idx, col_name in enumerate(header):
            if col_idx == 0:  # 跳过日期列
                continue
            
            values = []
            for row in monthly_data:
                cell_value = row[col_idx]
                if cell_value and isinstance(cell_value, (int, float)):
                    values.append(cell_value)
            
            if values:
                summary[col_name] = {
                    'total': round(sum(values)),
                    'average': round(sum(values) / len(values)),
                    'max': round(max(values)),
                    'min': round(min(values)),
                    'count': len(values)
                }
        
        return jsonify({
            'success': True,
            'data': monthly_data[:31],  # 最多31天
            'header': list(header),
            'summary': summary,
            'year': year,
            'month': month
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取月报数据失败: {str(e)}'})

@app.route('/api/get_quarterly_stats')
def get_quarterly_stats():
    """获取季度统计数据"""
    try:
        year = request.args.get('year', str(datetime.now().year), type=str)
        quarter = request.args.get('quarter', '1', type=str)
        
        # 计算季度包含的月份
        quarter_months = {
            '1': ['1', '2', '3'],
            '2': ['4', '5', '6'],
            '3': ['7', '8', '9'],
            '4': ['10', '11', '12']
        }
        
        months = quarter_months.get(quarter, ['1', '2', '3'])
        
        excel_path = DATA_SOURCE_PATH
        if not os.path.exists(excel_path):
            return jsonify({'success': False, 'message': 'Excel文件不存在'})
        
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        
        # 读取表头和数据
        all_rows = list(ws.iter_rows(values_only=True))
        header = all_rows[3]
        data_rows = all_rows[4:]
        
        # 筛选季度数据
        year_int = int(year)
        months_int = [int(m) for m in months]
        quarterly_data = []
        
        for row in data_rows:
            date_val = row[0]
            if isinstance(date_val, datetime) and date_val.year == year_int and date_val.month in months_int:
                quarterly_data.append(row)
        
        wb.close()
        
        if not quarterly_data:
            return jsonify({
                'success': True,
                'data': [],
                'header': list(header),
                'summary': {},
                'message': f'未找到{year}年第{quarter}季度的数据'
            })
        
        # 计算季度统计
        summary = {}
        for col_idx, col_name in enumerate(header):
            if col_idx == 0:
                continue
            
            values = []
            for row in quarterly_data:
                cell_value = row[col_idx]
                if cell_value and isinstance(cell_value, (int, float)):
                    values.append(cell_value)
            
            if values:
                summary[col_name] = {
                    'total': round(sum(values)),
                    'average': round(sum(values) / len(values)),
                    'max': round(max(values)),
                    'min': round(min(values)),
                    'count': len(values)
                }
        
        return jsonify({
            'success': True,
            'data': quarterly_data[:100],  # 最多100条
            'header': list(header),
            'summary': summary,
            'year': year,
            'quarter': quarter,
            'months': months
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取季报数据失败: {str(e)}'})

@app.route('/api/get_excel_data')
def get_excel_data():
    """
    获取Excel数据（只读模式）
    默认只显示当前年份的数据，支持分页和搜索
    """
    try:
        excel_path = DATA_SOURCE_PATH
        
        if not os.path.exists(excel_path):
            return jsonify({
                'success': False,
                'message': f'Excel文件不存在: {excel_path}'
            })
        
        # 获取参数
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 200, type=int)
        search_text = request.args.get('search', '', type=str)
        year_filter = request.args.get('year', str(datetime.now().year), type=str)  # 默认当前年份
        
        # 读取Excel文件
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        
        # 获取所有数据
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            row_data = [cell if cell is not None else '' for cell in row]
            all_rows.append(row_data)
        
        wb.close()
        
        if len(all_rows) < 5:
            return jsonify({
                'success': False,
                'message': 'Excel文件数据不完整'
            })
        
        # Excel 结构：
        # 第1行: 标题（监控表流量明细）
        # 第2行: 空行
        # 第3行: 合并单元格标题（进水、片区过水/进水）
        # 第4行: 真正的表头
        # 第5行开始: 数据
        
        # 使用第4行作为表头（索引3）
        header = all_rows[3]
        
        # 从第5行开始作为数据（索引4）
        data_rows = all_rows[4:]
        
        # 组合成最终数据：[表头, 数据行1, 数据行2, ...]
        all_data = [header] + data_rows
        
        # 按年份过滤（如果不是"全部"）
        if year_filter != 'all':
            header = all_data[0]
            filtered_by_year = [header]  # 保留表头
            
            for row in all_data[1:]:
                # 检查第一列（日期列）是否包含指定年份
                date_str = str(row[0]) if row and row[0] else ''
                if year_filter in date_str:
                    filtered_by_year.append(row)
        else:
            filtered_by_year = all_data
        
        # 如果有搜索条件，进一步过滤
        if search_text:
            search_lower = search_text.lower()
            filtered_data = [filtered_by_year[0]]  # 保留表头
            for row in filtered_by_year[1:]:
                if any(search_lower in str(cell).lower() for cell in row):
                    filtered_data.append(row)
        else:
            filtered_data = filtered_by_year
        
        # 分离表头和数据行
        header = filtered_data[0]
        data_rows_only = filtered_data[1:]
        
        # 计算分页（只对数据行分页）
        total_data_rows = len(data_rows_only)
        total_pages = (total_data_rows + page_size - 1) // page_size if total_data_rows > 0 else 1
        start_idx = (page - 1) * page_size
        end_idx = min(start_idx + page_size, total_data_rows)
        
        # 获取当前页的数据行
        current_page_rows = data_rows_only[start_idx:end_idx]
        
        # 每页都返回：表头 + 当前页数据
        page_data = [header] + current_page_rows
        
        # 获取文件修改时间
        file_time = os.path.getmtime(excel_path)
        last_update = datetime.fromtimestamp(file_time).strftime('%Y-%m-%d %H:%M:%S')
        
        return jsonify({
            'success': True,
            'data': page_data,
            'total_rows': total_data_rows,  # 只计算数据行数，不含表头
            'total_cols': len(header) if header else 0,
            'current_page': page,
            'page_size': page_size,
            'total_pages': total_pages,
            'last_update': last_update,
            'file_name': '石滩供水服务部每日总供水情况.xlsx'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'读取Excel失败: {str(e)}'
        })

# ==================== 分区计量表查看 ====================

@app.route('/view_partition_meter')
def view_partition_meter():
    """显示分区计量表查看页面"""
    return render_template('view_partition_meter.html')

@app.route('/api/get_partition_meter_data')
def get_partition_meter_data():
    """获取分区计量表数据"""
    try:
        sheet_name = request.args.get('sheet', '石滩区')  # 默认显示"石滩区"工作表
        
        if not os.path.exists(EXCEL_PATH):
            return jsonify({
                'success': False,
                'message': f'Excel文件不存在: {EXCEL_PATH}'
            })
        
        # 读取Excel文件
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        
        # 获取所有工作表名称
        sheet_names = wb.sheetnames
        
        # 检查请求的工作表是否存在
        if sheet_name not in sheet_names:
            sheet_name = sheet_names[0]  # 如果不存在，使用第一个工作表
        
        ws = wb[sheet_name]
        
        # 读取所有数据
        data = []
        for row in ws.iter_rows(values_only=True):
            # 转换为列表，处理None值
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append('')
                elif isinstance(cell, (int, float)):
                    # 检查是否是Excel日期序列号
                    if isinstance(cell, int) and 40000 <= cell <= 50000:
                        try:
                            excel_start = datetime(1899, 12, 30)
                            actual_date = excel_start + timedelta(days=cell)
                            row_data.append(f"{actual_date.year}年{actual_date.month}月{actual_date.day}日")
                        except:
                            row_data.append(cell)
                    else:
                        row_data.append(cell)
                else:
                    row_data.append(str(cell))
            data.append(row_data)
        
        # 获取合并单元格信息
        merged_cells = []
        for merged_range in ws.merged_cells.ranges:
            merged_cells.append({
                'start_row': merged_range.min_row - 1,  # 转换为0索引
                'start_col': merged_range.min_col - 1,
                'end_row': merged_range.max_row - 1,
                'end_col': merged_range.max_col - 1
            })
        
        wb.close()
        
        # 获取文件最后修改时间
        file_time = os.path.getmtime(EXCEL_PATH)
        last_update = datetime.fromtimestamp(file_time).strftime('%Y-%m-%d %H:%M:%S')
        
        return jsonify({
            'success': True,
            'data': data,
            'sheet_names': sheet_names,
            'current_sheet': sheet_name,
            'merged_cells': merged_cells,
            'total_rows': len(data),
            'total_cols': len(data[0]) if data else 0,
            'last_update': last_update,
            'file_name': '石滩区分区计量.xlsx'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'读取Excel失败: {str(e)}'
        })

# ==================== 启动应用 ====================

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)

