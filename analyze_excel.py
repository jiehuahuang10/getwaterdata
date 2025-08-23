#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from datetime import datetime

def analyze_excel_file():
    """分析石滩供水服务部每日总供水情况.xlsx文件"""
    
    try:
        # 读取Excel文件
        wb = openpyxl.load_workbook('excel_exports/石滩供水服务部每日总供水情况.xlsx')
        sheet = wb.active
        
        print('=== Excel文件基本信息 ===')
        print(f'工作表名称: {sheet.title}')
        print(f'总行数: {sheet.max_row}')
        print(f'总列数: {sheet.max_column}')
        print()
        
        # 读取第4行作为列标题（根据之前的分析，第4行是实际的列标题）
        print('=== 列标题分析 ===')
        headers = []
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(4, col).value
            if header and str(header).strip():
                headers.append((col, str(header).strip()))
        
        print(f'共找到 {len(headers)} 个有效列标题:')
        for col_num, header in headers:
            print(f'  第{col_num}列: "{header}"')
        
        print()
        print('=== 水表列映射分析 ===')
        
        # 用户提供的映射关系
        mapping = {
            '荔新大道DN1200流量计': '荔新大道',
            '新城大道医院DN800流量计': '新城大道', 
            '三江新总表DN800（2190066）': '三江新总表',
            '宁西总表DN1200': '宁西2总表',
            '沙庄总表': '沙庄总表',
            '如丰大道600监控表': '如丰大道600监控表',
            '三棵树600监控表': '三棵树600监控表',
            '2501200108': '中山西路DN300流量计'
        }
        
        print('需要映射的水表:')
        found_columns = {}
        for system_name, excel_name in mapping.items():
            print(f'  系统名称: "{system_name}" -> Excel列名: "{excel_name}"')
            
            # 查找对应的列号
            found = False
            for col_num, header in headers:
                # 模糊匹配，处理可能的换行符和空格
                header_clean = header.replace('\n', '').replace(' ', '')
                excel_name_clean = excel_name.replace('\n', '').replace(' ', '')
                
                if excel_name_clean in header_clean or header_clean in excel_name_clean:
                    found_columns[system_name] = col_num
                    print(f'    ✓ 找到对应列: 第{col_num}列 "{header}"')
                    found = True
                    break
            
            if not found:
                print(f'    ✗ 未找到对应列')
        
        print()
        print('=== 日期范围分析 ===')
        
        # 查找日期范围（从第5行开始，因为前4行是标题）
        dates = []
        for row in range(5, min(50, sheet.max_row + 1)):
            cell_value = sheet.cell(row, 1).value  # 第1列是日期列
            if cell_value:
                if isinstance(cell_value, datetime):
                    dates.append(cell_value.strftime('%Y-%m-%d'))
                elif isinstance(cell_value, str) and len(cell_value) >= 8:
                    # 处理字符串日期
                    try:
                        # 尝试解析不同格式的日期
                        if '-' in cell_value:
                            dates.append(cell_value)
                    except:
                        pass
        
        if dates:
            print(f'日期范围: {dates[0]} 到 {dates[-1] if len(dates) > 1 else dates[0]}')
            print(f'样例日期（前10个）: {dates[:10]}')
            print(f'总共找到 {len(dates)} 个日期')
        
        print()
        print('=== 样例数据行 ===')
        # 显示几行实际数据
        for row in range(5, min(10, sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(10, sheet.max_column + 1)):  # 只显示前9列
                cell_value = sheet.cell(row, col).value
                if cell_value is None:
                    cell_value = ''
                elif isinstance(cell_value, datetime):
                    cell_value = cell_value.strftime('%Y-%m-%d')
                row_data.append(str(cell_value)[:15])  # 限制长度
            print(f'第{row}行: {row_data}')
        
        wb.close()
        
        return found_columns
        
    except Exception as e:
        print(f'分析Excel文件时出错: {e}')
        import traceback
        traceback.print_exc()
        return {}

if __name__ == '__main__':
    analyze_excel_file()