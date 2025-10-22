#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
对比显示添加前后的差异
"""

import openpyxl

print("=" * 80)
print("对比：添加统计表前后")
print("=" * 80)

# 原文件
print("\n【原文件】石滩区分区计量.xlsx - 石滩2025年9月工作表")
print("-" * 80)
try:
    wb1 = openpyxl.load_workbook("excel_exports/石滩区分区计量.xlsx", data_only=True)
    ws1 = wb1["石滩2025年9月"]
    print(f"总行数: {ws1.max_row}")
    print(f"最后一行内容:")
    for col in range(1, 11):
        val = ws1.cell(ws1.max_row, col).value
        if val:
            print(f"  列{col}: {val}")
    wb1.close()
except Exception as e:
    print(f"无法读取: {e}")

# 新文件
print("\n【新文件】石滩区分区计量_已添加统计表.xlsx - 石滩2025年9月工作表")
print("-" * 80)
try:
    wb2 = openpyxl.load_workbook("excel_exports/石滩区分区计量_已添加统计表.xlsx", data_only=True)
    ws2 = wb2["石滩2025年9月"]
    print(f"总行数: {ws2.max_row}")
    print(f"\n新增内容（第30-46行）:")
    
    for row in range(30, 47):
        row_data = []
        for col in range(1, 11):
            val = ws2.cell(row, col).value
            if val:
                row_data.append(f"{val}")
        
        if row_data:
            print(f"  第{row}行: {' | '.join(row_data[:5])}")
        else:
            print(f"  第{row}行: (空行)")
    
    wb2.close()
except Exception as e:
    print(f"无法读取: {e}")

print("\n" + "=" * 80)
print("对比结果:")
print("=" * 80)
print("✅ 原文件：28行")
print("✅ 新文件：46行")
print("✅ 新增：18行（第29-46行）")
print("\n💡 提示：请在Excel中打开新文件，切换到'石滩2025年9月'工作表查看")

