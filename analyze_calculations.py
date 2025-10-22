#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
深度分析石滩区分区计量表格的数据计算情况
"""

import openpyxl
from openpyxl.utils import get_column_letter
import re
from datetime import datetime

def analyze_formulas_in_sheet(ws, sheet_name):
    """分析工作表中的所有公式"""
    print(f"\n{'='*80}")
    print(f"工作表: {sheet_name}")
    print(f"{'='*80}")
    
    formulas = {}
    values = {}
    
    # 遍历所有单元格
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                
                # 如果是公式
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas[cell_ref] = {
                        'formula': cell.value,
                        'row': cell.row,
                        'col': cell.column,
                        'result': cell.value  # 实际计算结果
                    }
                # 如果是数值
                elif isinstance(cell.value, (int, float)):
                    values[cell_ref] = {
                        'value': cell.value,
                        'row': cell.row,
                        'col': cell.column
                    }
    
    return formulas, values

def analyze_monthly_summary_pattern(ws):
    """分析月度统计表的计算模式"""
    print(f"\n[分析] 查找月度统计表...")
    
    summaries = []
    
    # 查找所有包含"年"和"月"的标题行
    for row in range(1, ws.max_row + 1):
        for col in range(1, 11):
            cell_value = ws.cell(row, col).value
            if cell_value and isinstance(cell_value, str):
                if re.search(r'\d{4}年\d{1,2}月', cell_value):
                    summaries.append({
                        'title': cell_value,
                        'title_row': row,
                        'data_start_row': row + 2  # 标题后两行是数据
                    })
                    break
    
    print(f"[OK] 找到 {len(summaries)} 个月度统计表")
    
    return summaries

def analyze_calculation_pattern(ws, start_row):
    """分析特定统计表的计算模式"""
    print(f"\n--- 第{start_row}行统计表 ---")
    
    # 读取表头
    headers = {}
    header_row = start_row + 1
    for col in range(1, 11):
        val = ws.cell(header_row, col).value
        if val:
            headers[col] = val
    
    print(f"表头: {headers}")
    
    # 分析数据行
    data_rows = []
    for offset in range(2, 35):  # 假设最多31天 + 总计行
        row = start_row + offset
        row_data = {}
        has_data = False
        
        for col in range(1, 11):
            cell = ws.cell(row, col)
            cell_ref = f"{get_column_letter(col)}{row}"
            
            if cell.value:
                has_data = True
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    row_data[col] = {
                        'type': 'formula',
                        'formula': cell.value,
                        'cell_ref': cell_ref
                    }
                else:
                    row_data[col] = {
                        'type': 'value',
                        'value': cell.value,
                        'cell_ref': cell_ref
                    }
        
        if has_data:
            # 判断是否为总计行
            first_col_val = ws.cell(row, 1).value
            is_total = False
            if first_col_val and isinstance(first_col_val, str):
                if '总计' in first_col_val or '合计' in first_col_val:
                    is_total = True
            
            data_rows.append({
                'row': row,
                'is_total': is_total,
                'data': row_data
            })
        
        # 如果遇到总计行后的空行，停止
        if is_total and offset > 2:
            # 再检查后面3行是否都是空的
            all_empty = True
            for check_offset in range(1, 4):
                check_row = row + check_offset
                for check_col in range(1, 11):
                    if ws.cell(check_row, check_col).value:
                        all_empty = False
                        break
                if not all_empty:
                    break
            if all_empty:
                break
    
    print(f"数据行数: {len(data_rows)}")
    
    # 分析公式模式
    formula_patterns = {}
    for row_info in data_rows:
        row_num = row_info['row']
        row_label = "总计行" if row_info['is_total'] else f"数据行{row_num}"
        
        for col, cell_data in row_info['data'].items():
            if cell_data['type'] == 'formula':
                col_name = headers.get(col, f"列{col}")
                if col_name not in formula_patterns:
                    formula_patterns[col_name] = []
                
                formula_patterns[col_name].append({
                    'row': row_num,
                    'row_type': '总计' if row_info['is_total'] else '数据',
                    'formula': cell_data['formula'],
                    'cell_ref': cell_data['cell_ref']
                })
    
    # 输出公式模式
    print(f"\n[公式分析]")
    for col_name, formulas in formula_patterns.items():
        print(f"\n  {col_name}:")
        for f in formulas[:3]:  # 只显示前3个
            print(f"    {f['cell_ref']} ({f['row_type']}): {f['formula']}")
        if len(formulas) > 3:
            print(f"    ... 共{len(formulas)}个公式")
    
    return data_rows, formula_patterns

def main():
    excel_path = "excel_exports/石滩区分区计量.xlsx"
    
    print("="*80)
    print("[START] Analysis - Data Calculation Deep Analysis")
    print("="*80)
    print(f"File: {excel_path}")
    print(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        
        print(f"\n[基础信息]")
        print(f"工作表数量: {len(wb.sheetnames)}")
        print(f"工作表列表: {wb.sheetnames}")
        
        # 重点分析"石滩区"主表
        if "石滩区" in wb.sheetnames:
            ws = wb["石滩区"]
            print(f"\n{'='*80}")
            print(f"重点分析: 石滩区 主表")
            print(f"{'='*80}")
            print(f"总行数: {ws.max_row}")
            print(f"总列数: {ws.max_column}")
            
            # 1. 查找所有公式
            print(f"\n[步骤1] 扫描所有公式...")
            formulas, values = analyze_formulas_in_sheet(ws, "石滩区")
            print(f"[OK] 找到 {len(formulas)} 个公式单元格")
            print(f"[OK] 找到 {len(values)} 个数值单元格")
            
            # 显示前10个公式示例
            print(f"\n[公式示例] 前10个:")
            for i, (cell_ref, info) in enumerate(list(formulas.items())[:10]):
                print(f"  {cell_ref}: {info['formula']}")
            
            # 2. 分析月度统计表
            summaries = analyze_monthly_summary_pattern(ws)
            
            # 3. 深度分析第一个统计表的计算模式
            if summaries:
                first_summary = summaries[0]
                print(f"\n{'='*80}")
                print(f"深度分析第一个统计表: {first_summary['title']}")
                print(f"{'='*80}")
                
                data_rows, formula_patterns = analyze_calculation_pattern(
                    ws, 
                    first_summary['title_row']
                )
                
                # 4. 分析计算规则
                print(f"\n{'='*80}")
                print(f"计算规则总结")
                print(f"{'='*80}")
                
                # 分析供水量、售水量、损耗水量、水损耗的关系
                print(f"\n[关键列关系分析]")
                print(f"从公式模式推断:")
                
                for col_name, formulas_list in formula_patterns.items():
                    if formulas_list:
                        sample = formulas_list[0]
                        formula = sample['formula']
                        
                        print(f"\n  {col_name}:")
                        
                        # 分析公式类型
                        if 'SUM' in formula.upper():
                            print(f"    类型: 求和公式")
                            print(f"    示例: {formula}")
                        elif '+' in formula:
                            print(f"    类型: 加法运算")
                            print(f"    示例: {formula}")
                        elif '-' in formula:
                            print(f"    类型: 减法运算")
                            print(f"    示例: {formula}")
                        elif '/' in formula:
                            print(f"    类型: 除法运算")
                            print(f"    示例: {formula}")
                        elif '*' in formula:
                            print(f"    类型: 乘法运算")
                            print(f"    示例: {formula}")
                        else:
                            print(f"    类型: 其他")
                            print(f"    示例: {formula}")
                
                # 5. 检查最后一个统计表（最新的）
                if len(summaries) > 1:
                    last_summary = summaries[-1]
                    print(f"\n{'='*80}")
                    print(f"最新统计表: {last_summary['title']}")
                    print(f"{'='*80}")
                    
                    data_rows_last, formula_patterns_last = analyze_calculation_pattern(
                        ws, 
                        last_summary['title_row']
                    )
        
        # 分析其他工作表（如果有月度明细表）
        print(f"\n{'='*80}")
        print(f"其他工作表快速扫描")
        print(f"{'='*80}")
        
        for sheet_name in wb.sheetnames:
            if sheet_name != "石滩区":
                ws = wb[sheet_name]
                formulas, values = analyze_formulas_in_sheet(ws, sheet_name)
                print(f"\n{sheet_name}:")
                print(f"  公式数: {len(formulas)}")
                print(f"  数值数: {len(values)}")
                print(f"  总行数: {ws.max_row}")
                print(f"  总列数: {ws.max_column}")
                
                if formulas:
                    print(f"  公式示例:")
                    for i, (cell_ref, info) in enumerate(list(formulas.items())[:3]):
                        print(f"    {cell_ref}: {info['formula']}")
        
        print(f"\n{'='*80}")
        print(f"分析完成！")
        print(f"{'='*80}")
        
        # 生成分析报告文件
        report_path = "CALCULATION_ANALYSIS_REPORT.md"
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("# 石滩区分区计量表格 - 数据计算深度分析报告\n\n")
            f.write(f"**分析时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            f.write(f"**文件**: {excel_path}\n\n")
            f.write("---\n\n")
            
            f.write("## 1. 基础信息\n\n")
            f.write(f"- 工作表数量: {len(wb.sheetnames)}\n")
            f.write(f"- 工作表列表: {', '.join(wb.sheetnames)}\n\n")
            
            if "石滩区" in wb.sheetnames:
                ws = wb["石滩区"]
                f.write("## 2. 石滩区主表\n\n")
                f.write(f"- 总行数: {ws.max_row}\n")
                f.write(f"- 总列数: {ws.max_column}\n")
                f.write(f"- 公式单元格数: {len(formulas)}\n")
                f.write(f"- 数值单元格数: {len(values)}\n\n")
                
                f.write("### 2.1 月度统计表\n\n")
                f.write(f"找到 {len(summaries)} 个月度统计表:\n\n")
                for s in summaries:
                    f.write(f"- {s['title']} (第{s['title_row']}行)\n")
                
                f.write("\n### 2.2 公式模式分析\n\n")
                if formula_patterns:
                    for col_name, formulas_list in formula_patterns.items():
                        f.write(f"#### {col_name}\n\n")
                        if formulas_list:
                            f.write(f"公式数量: {len(formulas_list)}\n\n")
                            f.write("示例:\n\n")
                            for formula_info in formulas_list[:3]:
                                f.write(f"- `{formula_info['cell_ref']}` ({formula_info['row_type']}): `{formula_info['formula']}`\n")
                            f.write("\n")
        
        print(f"\n[OK] 详细报告已保存到: {report_path}")
        
    except Exception as e:
        print(f"\n[ERROR] 分析失败: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()

