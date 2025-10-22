#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
查找数据源中最后一个有效数据的日期（非0）
"""

import openpyxl
from datetime import datetime

def main():
    data_file = "excel_exports/石滩供水服务部每日总供水情况.xlsx"
    
    print("="*100)
    print("[Finding Last Valid Date with Non-Zero Data]")
    print("="*100)
    
    wb = openpyxl.load_workbook(data_file, data_only=True)
    ws = wb.worksheets[2]  # 第3个工作表
    
    header_row = 4
    date_col = 1
    data_start_row = 5
    
    # 监控点列
    check_columns = [7, 12, 14, 8, 15]  # 荔新大道、宁西2总表等
    
    print(f"\n[Scanning from end backwards...]")
    
    last_valid_date = None
    last_valid_row = None
    
    # 从后往前找，找到第一个有非零数据的行
    for row in range(ws.max_row, data_start_row - 1, -1):
        date_val = ws.cell(row, date_col).value
        
        if isinstance(date_val, datetime):
            # 检查这一行的监控点数据是否有非零值
            has_non_zero = False
            row_values = []
            
            for col in check_columns:
                val = ws.cell(row, col).value
                row_values.append(val if val else 0)
                if val and isinstance(val, (int, float)) and val != 0:
                    has_non_zero = True
            
            if has_non_zero:
                last_valid_date = date_val
                last_valid_row = row
                print(f"\n[Found] Last date with non-zero data:")
                print(f"  Date: {date_val.strftime('%Y-%m-%d')}")
                print(f"  Row: {row}")
                print(f"  Sample values: {[f'{v:,.0f}' if v else '0' for v in row_values[:3]]}")
                break
            elif date_val.year == 2025 and date_val.month == 10:
                # 显示10月的数据情况
                if date_val.day in [14, 15, 18, 19, 20, 24]:
                    print(f"  Row {row} - {date_val.strftime('%Y-%m-%d')}: {row_values[:3]}")
    
    if last_valid_date:
        print(f"\n{'='*100}")
        print(f"[Result]")
        print(f"{'='*100}")
        print(f"\nLast valid date: {last_valid_date.strftime('%Y-%m-%d')}")
        print(f"Row: {last_valid_row}")
        
        # 检查10月的数据情况
        print(f"\n[Checking October 2025 data range...]")
        
        oct_start = datetime(2025, 9, 25)
        oct_end = datetime(2025, 10, 24)
        
        if last_valid_date < oct_start:
            print(f"  Status: No October data yet")
        elif last_valid_date < oct_end:
            missing_days = (oct_end - last_valid_date).days
            coverage = ((last_valid_date - oct_start).days + 1) / 31 * 100
            print(f"  Status: PARTIAL October data")
            print(f"  Available: {oct_start.strftime('%Y-%m-%d')} to {last_valid_date.strftime('%Y-%m-%d')}")
            print(f"  Missing: {missing_days} days (until {oct_end.strftime('%Y-%m-%d')})")
            print(f"  Coverage: {coverage:.1f}%")
        else:
            print(f"  Status: COMPLETE October data available")
    
    print(f"\n{'='*100}")
    print(f"[Complete]")
    print(f"{'='*100}")

if __name__ == "__main__":
    main()

