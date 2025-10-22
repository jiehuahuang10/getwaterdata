#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查统计表是否已添加
"""

import openpyxl

excel_path = "excel_exports/石滩区分区计量_已添加统计表.xlsx"

print("=" * 80)
print("检查统计表添加情况")
print("=" * 80)

wb = openpyxl.load_workbook(excel_path, data_only=True)

# 检查石滩2025年9月工作表
sheet_name = "石滩2025年9月"
if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n工作表: {sheet_name}")
    print(f"总行数: {ws.max_row}")
    
    # 显示最后20行
    print(f"\n最后20行内容:")
    for row in range(max(1, ws.max_row - 19), ws.max_row + 1):
        row_data = []
        has_data = False
        for col in range(1, min(11, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if val:
                row_data.append(f"列{col}:{val}")
                has_data = True
        
        if has_data:
            print(f"第{row}行: {' | '.join(row_data)}")
        else:
            print(f"第{row}行: (空行)")
    
    # 查找"10月"关键字
    print(f"\n搜索'10月'关键字:")
    found = False
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row, col).value
            if val and "10月" in str(val):
                print(f"  找到: 第{row}行,列{col} = {val}")
                found = True
    
    if not found:
        print("  未找到'10月'关键字")

else:
    print(f"[ERROR] 工作表不存在: {sheet_name}")

wb.close()

