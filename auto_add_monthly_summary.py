#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
每月1号自动添加统计表
可以集成到GitHub Actions或定时任务中
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import copy
from datetime import date, datetime
import sys

def get_last_month_info():
    """
    获取上个月的信息
    """
    today = date.today()
    if today.month == 1:
        last_month = 12
        last_year = today.year - 1
    else:
        last_month = today.month - 1
        last_year = today.year
    
    # 当前月份（用于标题）
    current_month = today.month
    current_year = today.year
    
    return {
        'last_year': last_year,
        'last_month': last_month,
        'current_year': current_year,
        'current_month': current_month,
        'sheet_name': f"石滩{last_year}年{last_month}月",
        'title': f"石滩{current_month}月分区计量统计"
    }

def add_monthly_summary(excel_path, force_date=None):
    """
    添加月度统计表
    
    Args:
        excel_path: Excel文件路径
        force_date: 强制指定日期（用于测试），格式：datetime对象
    """
    # 获取月份信息
    if force_date:
        # 测试模式：使用指定日期
        month_info = {
            'last_year': 2025,
            'last_month': 9,
            'current_year': 2025,
            'current_month': 10,
            'sheet_name': "石滩2025年9月",
            'title': "石滩10月分区计量统计"
        }
    else:
        month_info = get_last_month_info()
    
    print("=" * 80)
    print(f"[START] 添加月度统计表")
    print("=" * 80)
    print(f"[INFO] 目标工作表: {month_info['sheet_name']}")
    print(f"[INFO] 新统计表标题: {month_info['title']}")
    
    # 加载工作簿
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        print(f"[ERROR] 无法加载工作簿: {e}")
        return False
    
    # 检查工作表是否存在
    if month_info['sheet_name'] not in wb.sheetnames:
        print(f"[ERROR] 工作表不存在: {month_info['sheet_name']}")
        print(f"[INFO] 可用工作表: {', '.join(wb.sheetnames)}")
        return False
    
    ws = wb[month_info['sheet_name']]
    print(f"[OK] 找到工作表，当前最大行: {ws.max_row}")
    
    # 检查是否已经添加过（避免重复添加）
    # 检查最后10行是否有当前月份的标题
    for row in range(max(1, ws.max_row - 10), ws.max_row + 1):
        for col in range(1, 10):
            cell_value = ws.cell(row, col).value
            if cell_value and month_info['title'] in str(cell_value):
                print(f"[WARN] 检测到已存在相同标题的统计表（第{row}行）")
                print(f"[WARN] 跳过添加，避免重复")
                return False
    
    # 开始添加统计表
    start_row = ws.max_row + 2  # 留一行空行
    
    # 1. 添加标题行
    title_row = start_row
    title_cell = ws.cell(title_row, 2)
    title_cell.value = month_info['title']
    
    # 复制格式（从第2行）
    if ws.cell(2, 2).value:
        source_title = ws.cell(2, 2)
        title_cell.font = copy.copy(source_title.font)
        title_cell.alignment = copy.copy(source_title.alignment)
        title_cell.fill = copy.copy(source_title.fill)
    
    # 合并单元格
    ws.merge_cells(f'B{title_row}:G{title_row}')
    print(f"[OK] 添加标题行: 第{title_row}行")
    
    # 2. 添加时间行
    time_row1 = start_row + 2
    ws.cell(time_row1, 2).value = "数据时间"
    ws.cell(time_row1, 3).value = "待填写"
    
    time_row2 = start_row + 3
    ws.cell(time_row2, 2).value = "数据读取值时间"
    ws.cell(time_row2, 3).value = "待填写"
    
    # 3. 添加表头行
    header_row = start_row + 5
    headers = ["", "", "售水量", "供水量", "损耗水量", "水损耗(百分比)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col)
        cell.value = header
        # 复制格式
        if ws.cell(7, col).value is not None or col >= 3:
            source_cell = ws.cell(7, col)
            cell.font = copy.copy(source_cell.font)
            cell.alignment = copy.copy(source_cell.alignment)
            cell.border = copy.copy(source_cell.border)
            cell.fill = copy.copy(source_cell.fill)
    
    print(f"[OK] 添加表头行: 第{header_row}行")
    
    # 4. 添加数据行
    data_rows = ["荔城", "嘉园", "石滩"]
    current_row = header_row + 1
    
    for area_name in data_rows:
        ws.cell(current_row, 2).value = area_name
        ws.cell(current_row, 3).value = 0
        ws.cell(current_row, 4).value = 0
        ws.cell(current_row, 5).value = 0
        ws.cell(current_row, 6).value = 0
        
        # 复制格式
        for col in range(2, 7):
            cell = ws.cell(current_row, col)
            source_cell = ws.cell(8, col)
            cell.font = copy.copy(source_cell.font)
            cell.alignment = copy.copy(source_cell.alignment)
            cell.border = copy.copy(source_cell.border)
            if col >= 3:
                cell.number_format = source_cell.number_format
        
        print(f"[OK] 添加数据行: 第{current_row}行 - {area_name}")
        current_row += 1
    
    # 5. 添加合计行
    summary_row = current_row
    ws.cell(summary_row, 2).value = "合计"
    ws.cell(summary_row, 3).value = 0
    ws.cell(summary_row, 4).value = 0
    ws.cell(summary_row, 5).value = 0
    ws.cell(summary_row, 6).value = 0
    
    # 复制格式
    for col in range(2, 7):
        cell = ws.cell(summary_row, col)
        source_cell = ws.cell(11, col)
        cell.font = copy.copy(source_cell.font)
        cell.alignment = copy.copy(source_cell.alignment)
        cell.border = copy.copy(source_cell.border)
        cell.fill = copy.copy(source_cell.fill)
        if col >= 3:
            cell.number_format = source_cell.number_format
    
    print(f"[OK] 添加合计行: 第{summary_row}行")
    
    # 6. 添加监控表部分
    separator_row = summary_row + 3
    ws.cell(separator_row, 3).value = "监控表供水量"
    
    # 添加表头
    header2_row = separator_row + 1
    headers2 = ["", "荔新大道", "宁西总表（插入式）DN1200", "如丰大道", "新城大道医院NB", "三樵仔", "供水量", "售水量", "损耗水量", "水损耗（百分比）"]
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(header2_row, col)
        cell.value = header
        # 复制格式
        if ws.cell(15, col).value is not None or col >= 2:
            source_cell = ws.cell(15, col)
            cell.font = copy.copy(source_cell.font)
            cell.alignment = copy.copy(source_cell.alignment)
            cell.border = copy.copy(source_cell.border)
            cell.fill = copy.copy(source_cell.fill)
    
    print(f"[OK] 添加监控表表头: 第{header2_row}行")
    
    # 添加数据行
    data2_rows = ["1区", "2区", "3区"]
    current_row = header2_row + 1
    
    for area_name in data2_rows:
        ws.cell(current_row, 1).value = area_name
        for col in range(2, 11):
            ws.cell(current_row, col).value = 0
            # 复制格式
            source_cell = ws.cell(16, col)
            ws.cell(current_row, col).font = copy.copy(source_cell.font)
            ws.cell(current_row, col).alignment = copy.copy(source_cell.alignment)
            ws.cell(current_row, col).border = copy.copy(source_cell.border)
            ws.cell(current_row, col).number_format = source_cell.number_format
        
        print(f"[OK] 添加监控表数据行: 第{current_row}行 - {area_name}")
        current_row += 1
    
    # 保存
    try:
        wb.save(excel_path)
        print(f"\n[SAVE] 成功保存工作簿")
    except Exception as e:
        print(f"\n[ERROR] 保存失败: {e}")
        return False
    
    print("\n" + "=" * 80)
    print("[SUCCESS] 月度统计表添加完成！")
    print("=" * 80)
    print(f"新统计表位置: 第{title_row}行 到 第{current_row-1}行")
    print(f"总共添加: {current_row - title_row} 行")
    
    return True

def main():
    """
    主函数
    """
    excel_path = "excel_exports/石滩区分区计量.xlsx"
    
    # 检查今天是否是1号
    today = date.today()
    
    if len(sys.argv) > 1 and sys.argv[1] == '--force':
        # 强制执行（用于测试）
        print("[INFO] 强制执行模式（测试）")
        success = add_monthly_summary(excel_path, force_date=today)
    elif today.day == 1:
        # 每月1号自动执行
        print(f"[INFO] 今天是{today.year}年{today.month}月1日，执行添加统计表")
        success = add_monthly_summary(excel_path)
    else:
        print(f"[INFO] 今天是{today.year}年{today.month}月{today.day}日，不是1号，跳过执行")
        print(f"[INFO] 如需测试，请运行: python auto_add_monthly_summary.py --force")
        return
    
    if success:
        print("\n[SUCCESS] 任务完成！")
    else:
        print("\n[FAILED] 任务失败！")

if __name__ == "__main__":
    main()

