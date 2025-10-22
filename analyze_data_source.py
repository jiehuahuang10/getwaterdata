#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
分析石滩供水服务部每日总供水情况.xlsx的数据结构
"""

import openpyxl
from datetime import datetime
import os

def main():
    # 文件路径
    data_file = "excel_exports/石滩供水服务部每日总供水情况.xlsx"
    
    print("="*100)
    print("[Data Source Analysis]")
    print("="*100)
    
    # 检查文件是否存在
    if not os.path.exists(data_file):
        print(f"[ERROR] File not found: {data_file}")
        
        # 列出目录中的所有Excel文件
        print("\nExcel files in excel_exports/:")
        for file in os.listdir("excel_exports"):
            if file.endswith(('.xlsx', '.xls')):
                print(f"  - {file}")
        return
    
    try:
        print(f"\n[Loading] {data_file}")
        wb = openpyxl.load_workbook(data_file, data_only=True)
        
        print(f"\n[Worksheets] Total: {len(wb.sheetnames)}")
        for name in wb.sheetnames:
            print(f"  - {name}")
        
        # 分析第一个工作表
        ws = wb.active
        print(f"\n[Active Sheet] {ws.title}")
        print(f"  Total Rows: {ws.max_row}")
        print(f"  Total Columns: {ws.max_column}")
        
        # 读取表头
        print(f"\n[Headers] First row:")
        headers = {}
        for col in range(1, min(20, ws.max_column + 1)):
            val = ws.cell(1, col).value
            if val:
                headers[col] = val
                print(f"  Column {col}: {val}")
        
        # 查找目标列
        print(f"\n[Target Columns Mapping]:")
        target_mapping = {
            "荔新大道": None,
            "宁西2总表": None,
            "如丰大道600监控表": None,
            "新城大道医院NB": None,
            "三棵树600监控表": None
        }
        
        for col, header in headers.items():
            for target_name in target_mapping.keys():
                if target_name in str(header):
                    target_mapping[target_name] = col
                    print(f"  [Found] {target_name} -> Column {col}")
        
        # 检查缺失的列
        missing = [k for k, v in target_mapping.items() if v is None]
        if missing:
            print(f"\n[Warning] Missing columns:")
            for name in missing:
                print(f"  - {name}")
                # 尝试模糊匹配
                print(f"    Similar headers:")
                for col, header in headers.items():
                    if header and any(word in str(header) for word in name.split()):
                        print(f"      Column {col}: {header}")
        
        # 读取前10行数据示例
        print(f"\n[Data Sample] First 10 rows:")
        for row in range(1, min(11, ws.max_row + 1)):
            # 读取日期列和前几个数据列
            date_val = ws.cell(row, 1).value
            print(f"\n  Row {row}:")
            print(f"    Date: {date_val}")
            
            # 显示目标列的值
            for target_name, col_idx in target_mapping.items():
                if col_idx:
                    val = ws.cell(row, col_idx).value
                    print(f"    {target_name}: {val}")
        
        # 查找2025年8月25日和9月24日的数据
        print(f"\n[Date Range Test] Looking for 2025-08-25 to 2025-09-24:")
        
        date_col = 1  # 假设第1列是日期
        found_dates = []
        
        for row in range(2, min(100, ws.max_row + 1)):
            date_val = ws.cell(row, date_col).value
            
            if date_val:
                # 尝试解析日期
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime('%Y-%m-%d')
                    if '2025-08' in date_str or '2025-09' in date_str:
                        found_dates.append({
                            'row': row,
                            'date': date_str
                        })
        
        if found_dates:
            print(f"  [OK] Found {len(found_dates)} dates in Aug-Sep 2025:")
            for item in found_dates[:5]:
                print(f"    Row {item['row']}: {item['date']}")
            if len(found_dates) > 5:
                print(f"    ... and {len(found_dates) - 5} more")
        else:
            print(f"  [Info] No dates found in Aug-Sep 2025, showing recent dates:")
            for row in range(2, min(12, ws.max_row + 1)):
                date_val = ws.cell(row, date_col).value
                print(f"    Row {row}: {date_val} (type: {type(date_val).__name__})")
        
        print(f"\n{'='*100}")
        print("[Analysis Complete]")
        print(f"{'='*100}")
        
    except Exception as e:
        print(f"\n[ERROR] {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()

