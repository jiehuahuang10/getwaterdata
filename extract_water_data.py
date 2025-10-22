#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
从石滩供水服务部每日总供水情况.xlsx提取指定月份的数据
"""

import openpyxl
from datetime import datetime, timedelta

# 列映射关系（从分析结果得出）
COLUMN_MAPPING = {
    "荔新大道": 7,      # 对应"荔新大道"列
    "宁西2总表": 12,     # 对应"宁西2总表"列
    "如丰大道600监控表": 14,  # 对应"如丰大道600监控表"列
    "新城大道医院NB": None,    # 需要进一步查找
    "三棵树600监控表": 15,    # 对应"三棵树600监控表"列
}

def find_column_by_name(ws, header_row, search_terms):
    """
    在表头行中查找列
    
    Args:
        ws: 工作表
        header_row: 表头行号
        search_terms: 搜索关键词列表
    
    Returns:
        列号，如果未找到返回None
    """
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(header_row, col).value
        if header_val:
            header_str = str(header_val).strip()
            for term in search_terms:
                if term in header_str:
                    return col
    return None

def extract_monthly_data(year, month, data_file="excel_exports/石滩供水服务部每日总供水情况.xlsx"):
    """
    提取指定月份的数据
    
    时间范围：上月25日 - 本月24日
    
    Args:
        year: 年份
        month: 月份
        data_file: 数据源文件路径
    
    Returns:
        dict: 包含各监控点的总和数据
    """
    print(f"\n{'='*100}")
    print(f"[Extracting Data] {year}年{month}月")
    print(f"{'='*100}")
    
    # 计算日期范围
    if month == 1:
        start_date = datetime(year - 1, 12, 25)
    else:
        start_date = datetime(year, month - 1, 25)
    
    end_date = datetime(year, month, 24)
    
    print(f"\n[Date Range]")
    print(f"  From: {start_date.strftime('%Y-%m-%d')}")
    print(f"  To:   {end_date.strftime('%Y-%m-%d')}")
    print(f"  Days: {(end_date - start_date).days + 1}")
    
    # 加载数据文件
    print(f"\n[Loading] {data_file}")
    wb = openpyxl.load_workbook(data_file, data_only=True)
    ws = wb.worksheets[2]  # 第3个工作表
    
    header_row = 4
    date_col = 1
    data_start_row = 5
    
    # 更新列映射（查找可能缺失的列）
    print(f"\n[Column Mapping]")
    column_mapping = {
        "荔新大道": find_column_by_name(ws, header_row, ["荔新大道", "荔新"]),
        "宁西2总表": find_column_by_name(ws, header_row, ["宁西2总表", "宁西2"]),
        "如丰大道600监控表": find_column_by_name(ws, header_row, ["如丰大道600", "如丰大道"]),
        "新城大道医院NB": find_column_by_name(ws, header_row, ["新城大道医院NB", "新城大道医院", "新城大道"]),
        "三棵树600监控表": find_column_by_name(ws, header_row, ["三棵树600", "三棵树", "三棵竹600"]),
    }
    
    for name, col in column_mapping.items():
        if col:
            header_val = ws.cell(header_row, col).value
            print(f"  {name:20s} -> Column {col:2d} ({header_val})")
        else:
            print(f"  {name:20s} -> NOT FOUND!")
    
    # 查找日期范围内的数据
    print(f"\n[Searching Data Rows]...")
    matched_rows = []
    
    for row in range(data_start_row, ws.max_row + 1):
        date_val = ws.cell(row, date_col).value
        
        if isinstance(date_val, datetime):
            if start_date <= date_val <= end_date:
                matched_rows.append(row)
    
    print(f"  [OK] Found {len(matched_rows)} rows")
    
    if matched_rows:
        print(f"  First row: {matched_rows[0]} - {ws.cell(matched_rows[0], date_col).value.strftime('%Y-%m-%d')}")
        print(f"  Last row: {matched_rows[-1]} - {ws.cell(matched_rows[-1], date_col).value.strftime('%Y-%m-%d')}")
    
    # 计算总和
    print(f"\n[Calculating Totals]...")
    totals = {}
    
    for name, col in column_mapping.items():
        if col is None:
            totals[name] = 0
            print(f"  {name:20s} -> Column not found, set to 0")
            continue
        
        total = 0
        valid_count = 0
        
        for row in matched_rows:
            val = ws.cell(row, col).value
            if val and isinstance(val, (int, float)):
                total += val
                valid_count += 1
        
        totals[name] = total
        print(f"  {name:20s} -> {total:15,.0f} (from {valid_count} valid values)")
    
    return {
        "year": year,
        "month": month,
        "date_range": f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
        "days": len(matched_rows),
        "totals": totals
    }

def main():
    print("="*100)
    print("[Water Data Extraction Tool]")
    print("="*100)
    
    # 测试：提取2025年9月的数据
    result = extract_monthly_data(2025, 9)
    
    print(f"\n{'='*100}")
    print(f"[Result Summary]")
    print(f"{'='*100}")
    print(f"\nMonth: {result['year']}年{result['month']}月")
    print(f"Date Range: {result['date_range']}")
    print(f"Days of Data: {result['days']}")
    print(f"\nTotals:")
    for name, value in result['totals'].items():
        print(f"  {name:25s}: {value:15,.0f}")
    
    print(f"\n{'='*100}")
    print(f"[Complete]")
    print(f"{'='*100}")

if __name__ == "__main__":
    main()

