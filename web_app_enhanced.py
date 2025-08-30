#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# type: ignore
"""
水务数据获取Web界面 - 增强版
保留完整功能，同时兼容云部署
"""

from flask import Flask, render_template, jsonify, request, send_file
import json
import os
import glob
from datetime import datetime, timedelta
import threading
import time

# 尝试导入完整功能模块
try:
    import requests
    from bs4 import BeautifulSoup
    import hashlib
    HTTP_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ HTTP功能模块不可用: {e}")
    HTTP_AVAILABLE = False
    # 创建替代对象避免IDE错误
    class MockModule:
        def Session(self): return None
        def get(self, *args, **kwargs): return None
        def post(self, *args, **kwargs): return None
        def md5(self, *args, **kwargs): return type('obj', (object,), {'hexdigest': lambda: ''})
    requests = MockModule()
    BeautifulSoup = lambda *args, **kwargs: None
    hashlib = MockModule()

# 尝试导入Excel功能
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ Excel功能模块不可用: {e}")
    EXCEL_AVAILABLE = False
    # 创建替代对象避免IDE错误
    class MockExcel:
        def __init__(self, *args, **kwargs): 
            self.active = None
        def save(self, *args, **kwargs): pass
    Workbook = MockExcel
    Font = lambda *args, **kwargs: None
    PatternFill = lambda *args, **kwargs: None
    Alignment = lambda *args, **kwargs: None
    Border = lambda *args, **kwargs: None
    Side = lambda *args, **kwargs: None

# 尝试导入数据处理功能
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

app = Flask(__name__)

# 全局变量存储任务状态
task_status = {
    'running': False,
    'progress': 0,
    'message': '准备就绪',
    'data': None,
    'error': None,
    'start_time': None,
    'end_time': None
}

def calculate_recent_7days():
    """计算最近7天的日期范围"""
    today = datetime.now()
    end_date = today - timedelta(days=1)  # 昨天
    start_date = end_date - timedelta(days=7)  # 昨天往前推7天
    
    return start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')

def enhanced_water_data_fetch():
    """增强版水务数据获取功能"""
    global task_status
    
    if not HTTP_AVAILABLE:
        task_status['error'] = '网络功能不可用'
        task_status['message'] = '缺少HTTP请求模块'
        return
    
    try:
        task_status['running'] = True
        task_status['progress'] = 0
        task_status['message'] = '开始获取数据...'
        task_status['error'] = None
        task_status['start_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        task_status['data'] = None
        
        # 实际数据获取过程
        task_status['progress'] = 10
        task_status['message'] = '正在连接水务系统...'
        time.sleep(1)
        
        session = requests.Session()  # type: ignore
        
        # 登录步骤
        task_status['progress'] = 30
        task_status['message'] = '正在登录水务系统...'
        
        login_url = "http://axwater.dmas.cn/Login.aspx"
        try:
            login_page = session.get(login_url, timeout=10)
            
            soup = BeautifulSoup(login_page.text, 'html.parser')  # type: ignore
            form = soup.find('form')
            
            if not form:
                raise Exception("无法找到登录表单")
            
            form_data = {}
            for input_elem in form.find_all('input'):  # type: ignore
                name = input_elem.get('name')
                value = input_elem.get('value', '')
                if name:
                    form_data[name] = value
            
            # 填入登录信息
            username = os.environ.get('USERNAME', '13509288500')
            password = os.environ.get('PASSWORD', '288500')
            
            form_data['user'] = username
            form_data['pwd'] = hashlib.md5(password.encode('utf-8')).hexdigest()
            
            login_response = session.post(login_url, data=form_data, timeout=10)
            
            if "window.location='frmMain.aspx'" in login_response.text:
                task_status['progress'] = 60
                task_status['message'] = '登录成功，正在获取数据...'
                
                # 访问主页面
                main_url = "http://axwater.dmas.cn/frmMain.aspx"
                main_response = session.get(main_url, timeout=10)
                
                if main_response.status_code == 200:
                    # 访问报表页面
                    task_status['progress'] = 70
                    task_status['message'] = '正在访问报表页面...'
                    
                    report_url = "http://axwater.dmas.cn/reports/FluxRpt.aspx"
                    report_response = session.get(report_url, timeout=10)
                    
                    if '登录超时' not in report_response.text:
                        # 获取API数据
                        task_status['progress'] = 80
                        task_status['message'] = '正在调用数据API...'
                        
                        start_date, end_date = calculate_recent_7days()
                        
                        # 8个水表的ID
                        meter_ids = [
                            '1261181000263', '1261181000300', '1262330402331',
                            '2190066', '2190493', '2501200108', '2520005', '2520006'
                        ]
                        
                        formatted_node_ids = "'" + "','".join(meter_ids) + "'"
                        
                        api_url = "http://axwater.dmas.cn/reports/ashx/getRptWaterYield.ashx"
                        api_params = {
                            'nodeId': formatted_node_ids,
                            'startDate': start_date,
                            'endDate': end_date,
                            'meterType': '-1',
                            'statisticsType': 'flux',
                            'type': 'dayRpt'
                        }
                        
                        api_headers = {
                            'Referer': report_url,
                            'X-Requested-With': 'XMLHttpRequest',
                            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
                            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                        }
                        
                        api_response = session.post(api_url, data=api_params, headers=api_headers, timeout=15)
                        
                        if api_response.text and len(api_response.text) > 10:
                            try:
                                data = api_response.json()
                                
                                # 保存数据
                                timestamp = time.strftime('%Y%m%d_%H%M%S')
                                filename = f"ENHANCED_WEB_DATA_{timestamp}.json"
                                
                                output_data = {
                                    'timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
                                    'source': 'enhanced_web_interface',
                                    'success': True,
                                    'data_type': 'real_api_data',
                                    'calculation_date': datetime.now().strftime('%Y-%m-%d'),
                                    'date_range': {
                                        'start': start_date,
                                        'end': end_date,
                                        'description': '最近7天的真实数据'
                                    },
                                    'meter_count': len(meter_ids),
                                    'data': data,
                                    'excel_available': EXCEL_AVAILABLE,
                                    'pandas_available': PANDAS_AVAILABLE,
                                    'note': f'通过增强版Web界面获取的{datetime.now().strftime("%Y年%m月%d日")}真实数据'
                                }
                                
                                with open(filename, 'w', encoding='utf-8') as f:
                                    json.dump(output_data, f, ensure_ascii=False, indent=2)
                                
                                task_status['data'] = output_data
                                task_status['progress'] = 100
                                task_status['message'] = f'数据获取成功！共获取{len(data.get("rows", []))}个水表的真实数据'
                                
                            except json.JSONDecodeError:
                                task_status['error'] = 'API返回的不是有效的JSON数据'
                                task_status['message'] = '数据解析失败'
                        else:
                            task_status['error'] = 'API返回空响应'
                            task_status['message'] = '数据获取失败'
                    else:
                        task_status['error'] = '访问报表页面时显示登录超时'
                        task_status['message'] = '会话已过期'
                else:
                    task_status['error'] = '无法访问主页面'
                    task_status['message'] = '系统访问失败'
            else:
                task_status['error'] = '登录失败，请检查账号密码'
                task_status['message'] = '登录失败'
                
        except Exception as e:
            task_status['error'] = f'网络连接失败: {str(e)}'
            task_status['message'] = '无法连接到水务系统'
            
    except Exception as e:
        task_status['error'] = str(e)
        task_status['message'] = f'发生错误: {str(e)}'
    finally:
        task_status['running'] = False
        task_status['end_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def create_excel_export(data):
    """创建Excel导出功能"""
    if not EXCEL_AVAILABLE:
        return None
        
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "水表数据"
        
        # 设置表头
        headers = ['序号', '水表名称', '用水量', '单位', '状态', '采集时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # 填入数据
        if data and 'data' in data and 'rows' in data['data']:
            for row_idx, meter_data in enumerate(data['data']['rows'], 2):
                if isinstance(meter_data, dict):
                    ws.cell(row=row_idx, column=1, value=row_idx-1)
                    ws.cell(row=row_idx, column=2, value=meter_data.get('Name', 'N/A'))
                    
                    # 获取最新日期的数据
                    yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
                    value = meter_data.get(yesterday, 'N/A')
                    ws.cell(row=row_idx, column=3, value=value)
                    ws.cell(row=row_idx, column=4, value='立方米')
                    ws.cell(row=row_idx, column=5, value='正常')
                    ws.cell(row=row_idx, column=6, value=data.get('timestamp', ''))
        
        # 保存文件
        filename = f"水表数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        return filename
    except Exception as e:
        print(f"Excel导出失败: {e}")
        return None

@app.route('/')
def index():
    """增强版主页面"""
    excel_status = "✅ 可用" if EXCEL_AVAILABLE else "❌ 不可用"
    pandas_status = "✅ 可用" if PANDAS_AVAILABLE else "❌ 不可用" 
    http_status = "✅ 可用" if HTTP_AVAILABLE else "❌ 不可用"
    
    return f"""
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>水务数据获取系统 - 增强版</title>
        <style>
            body {{ font-family: 'Microsoft YaHei', sans-serif; margin: 0; padding: 20px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; }}
            .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 15px; box-shadow: 0 10px 30px rgba(0,0,0,0.3); }}
            h1 {{ color: #2c3e50; text-align: center; margin-bottom: 30px; font-size: 28px; }}
            .status {{ padding: 15px; margin: 20px 0; border-radius: 8px; }}
            .success {{ background: linear-gradient(135deg, #d4edda, #c3e6cb); color: #155724; border: 1px solid #c3e6cb; }}
            .error {{ background: linear-gradient(135deg, #f8d7da, #f5c6cb); color: #721c24; border: 1px solid #f5c6cb; }}
            .info {{ background: linear-gradient(135deg, #d1ecf1, #bee5eb); color: #0c5460; border: 1px solid #bee5eb; }}
            .warning {{ background: linear-gradient(135deg, #fff3cd, #ffeaa7); color: #856404; border: 1px solid #ffeaa7; }}
            .btn {{ padding: 12px 24px; margin: 10px 5px; border: none; border-radius: 8px; cursor: pointer; font-size: 16px; font-weight: bold; transition: all 0.3s ease; }}
            .btn:hover {{ transform: translateY(-2px); box-shadow: 0 5px 15px rgba(0,0,0,0.2); }}
            .btn-primary {{ background: linear-gradient(135deg, #007bff, #0056b3); color: white; }}
            .btn-success {{ background: linear-gradient(135deg, #28a745, #1e7e34); color: white; }}
            .btn-info {{ background: linear-gradient(135deg, #17a2b8, #117a8b); color: white; }}
            .btn-warning {{ background: linear-gradient(135deg, #ffc107, #e0a800); color: black; }}
            .progress {{ width: 100%; height: 25px; background: #e9ecef; border-radius: 15px; overflow: hidden; margin: 15px 0; }}
            .progress-bar {{ height: 100%; background: linear-gradient(90deg, #007bff, #28a745); transition: width 0.3s ease; border-radius: 15px; }}
            .data-display {{ background: #f8f9fa; padding: 25px; border-radius: 10px; margin: 20px 0; border-left: 5px solid #007bff; }}
            .meter {{ background: white; padding: 20px; margin: 15px 0; border-radius: 10px; border-left: 4px solid #28a745; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
            .meter-name {{ font-weight: bold; color: #2c3e50; font-size: 18px; }}
            .meter-value {{ font-size: 20px; color: #28a745; margin: 10px 0; }}
            .system-info {{ background: #f8f9fa; padding: 20px; border-radius: 10px; margin: 20px 0; }}
            .feature-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 20px; margin: 20px 0; }}
            .feature-card {{ background: white; padding: 20px; border-radius: 10px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); border-top: 4px solid #007bff; }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🌊 水务数据获取系统 - 增强版</h1>
            
            <div class="success status">
                <strong>🚀 系统状态：</strong>增强版已成功部署到云端！<br>
                <strong>📅 部署时间：</strong>{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br>
                <strong>🌐 访问地址：</strong>https://getwaterdata.onrender.com<br>
                <strong>💡 版本特色：</strong>完整功能 + 云端优化 + 智能降级
            </div>
            
            <div class="system-info">
                <h3>📊 系统功能状态</h3>
                <div class="feature-grid">
                    <div class="feature-card">
                        <h4>🌐 网络功能</h4>
                        <p>状态: {http_status}</p>
                        <p>支持真实水务数据获取</p>
                    </div>
                    <div class="feature-card">
                        <h4>📊 Excel导出</h4>
                        <p>状态: {excel_status}</p>
                        <p>支持专业Excel文件生成</p>
                    </div>
                    <div class="feature-card">
                        <h4>📈 数据分析</h4>
                        <p>状态: {pandas_status}</p>
                        <p>支持高级数据处理</p>
                    </div>
                </div>
            </div>
            
            <div id="status-display"></div>
            
            <div style="text-align: center; margin: 30px 0;">
                <button class="btn btn-primary" onclick="startDataFetch()">🔄 获取水表数据</button>
                <button class="btn btn-success" onclick="refreshStatus()">📊 刷新状态</button>
                <button class="btn btn-info" onclick="exportExcel()">📁 导出Excel</button>
                <button class="btn btn-warning" onclick="viewHistory()">📈 查看历史</button>
            </div>
            
            <div id="progress-container" style="display: none;">
                <div class="progress">
                    <div class="progress-bar" id="progress-bar" style="width: 0%;"></div>
                </div>
                <div id="progress-message">准备中...</div>
            </div>
            
            <div id="data-container"></div>
        </div>
        
        <script>
            function startDataFetch() {{
                fetch('/api/start-fetch', {{method: 'POST'}})
                    .then(response => response.json())
                    .then(data => {{
                        console.log('Start fetch response:', data);
                        document.getElementById('progress-container').style.display = 'block';
                        checkProgress();
                    }})
                    .catch(error => console.error('Error:', error));
            }}
            
            function exportExcel() {{
                fetch('/api/export-excel', {{method: 'POST'}})
                    .then(response => {{
                        if (response.ok) {{
                            return response.blob();
                        }}
                        throw new Error('Excel导出失败');
                    }})
                    .then(blob => {{
                        const url = window.URL.createObjectURL(blob);
                        const a = document.createElement('a');
                        a.style.display = 'none';
                        a.href = url;
                        a.download = '水表数据_' + new Date().toISOString().slice(0,10) + '.xlsx';
                        document.body.appendChild(a);
                        a.click();
                        window.URL.revokeObjectURL(url);
                    }})
                    .catch(error => {{
                        alert('Excel导出功能不可用: ' + error.message);
                    }});
            }}
            
            function checkProgress() {{
                fetch('/api/status')
                    .then(response => response.json())
                    .then(data => {{
                        updateProgress(data);
                        if (data.running) {{
                            setTimeout(checkProgress, 1000);
                        }}
                    }})
                    .catch(error => console.error('Error:', error));
            }}
            
            function updateProgress(status) {{
                const progressBar = document.getElementById('progress-bar');
                const progressMessage = document.getElementById('progress-message');
                const dataContainer = document.getElementById('data-container');
                
                progressBar.style.width = status.progress + '%';
                progressMessage.textContent = status.message;
                
                if (status.error) {{
                    document.getElementById('status-display').innerHTML = 
                        '<div class="error status"><strong>❌ 错误：</strong>' + status.error + '</div>';
                }}
                
                if (status.data && status.data.data && status.data.data.rows) {{
                    let html = '<div class="data-display"><h3>📊 获取的数据：</h3>';
                    html += '<p><strong>数据来源：</strong>' + status.data.source + '</p>';
                    html += '<p><strong>数据时间：</strong>' + status.data.timestamp + '</p>';
                    html += '<p><strong>数据类型：</strong>' + status.data.data_type + '</p>';
                    
                    status.data.data.rows.forEach(meter => {{
                        if (typeof meter === 'object' && meter.Name) {{
                            html += '<div class="meter">';
                            html += '<div class="meter-name">' + meter.Name + '</div>';
                            
                            // 查找最新日期的数据
                            const yesterday = new Date();
                            yesterday.setDate(yesterday.getDate() - 1);
                            const dateStr = yesterday.toISOString().slice(0,10);
                            const value = meter[dateStr] || 'N/A';
                            
                            html += '<div class="meter-value">用水量: ' + value + ' 立方米</div>';
                            html += '</div>';
                        }}
                    }});
                    html += '</div>';
                    dataContainer.innerHTML = html;
                }}
            }}
            
            function refreshStatus() {{
                fetch('/api/status')
                    .then(response => response.json())
                    .then(data => updateProgress(data))
                    .catch(error => console.error('Error:', error));
            }}
            
            function viewHistory() {{
                fetch('/api/history')
                    .then(response => response.json())
                    .then(data => {{
                        alert('历史数据功能：共找到 ' + data.count + ' 条历史记录');
                    }})
                    .catch(error => console.error('Error:', error));
            }}
            
            // 页面加载时获取状态
            window.onload = function() {{
                refreshStatus();
            }}
        </script>
    </body>
    </html>
    """

@app.route('/api/start-fetch', methods=['POST'])
def start_fetch():
    """启动数据获取"""
    if task_status['running']:
        return jsonify({'success': False, 'message': '数据获取正在进行中'})
    
    # 在后台线程中运行数据获取
    thread = threading.Thread(target=enhanced_water_data_fetch)
    thread.daemon = True
    thread.start()
    
    return jsonify({'success': True, 'message': '数据获取已启动'})

@app.route('/api/status')
def get_status():
    """获取当前状态"""
    return jsonify(task_status)

@app.route('/api/export-excel', methods=['POST'])
def export_excel():
    """Excel导出功能"""
    if not EXCEL_AVAILABLE:
        return jsonify({'success': False, 'error': 'Excel功能不可用'}), 400
    
    if not task_status.get('data'):
        return jsonify({'success': False, 'error': '没有可导出的数据'}), 400
    
    try:
        filename = create_excel_export(task_status['data'])
        if filename and os.path.exists(filename):
            return send_file(filename, as_attachment=True, download_name=filename)
        else:
            return jsonify({'success': False, 'error': 'Excel文件生成失败'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/history')
def get_history():
    """获取历史数据"""
    try:
        # 查找所有JSON数据文件
        json_files = glob.glob("*_WEB_DATA_*.json")
        return jsonify({
            'success': True,
            'count': len(json_files),
            'files': json_files[-10:]  # 返回最近10个文件
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

if __name__ == '__main__':
    print("🌐 启动水务数据获取Web界面（增强版）...")
    print("📱 功能特色: 完整功能 + 云端优化 + 智能降级")
    print("🔧 Excel导出:", "✅ 可用" if EXCEL_AVAILABLE else "❌ 降级模式")
    print("📊 数据分析:", "✅ 可用" if PANDAS_AVAILABLE else "❌ 降级模式") 
    print("🌐 网络功能:", "✅ 可用" if HTTP_AVAILABLE else "❌ 离线模式")
    
    # 支持云部署的端口配置
    import os
    port = int(os.environ.get('PORT', 5000))
    debug_mode = os.environ.get('FLASK_ENV') != 'production'
    
    app.run(debug=debug_mode, host='0.0.0.0', port=port)