#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
智能提取水数据 - 处理当前月份数据不完整的情况
"""

import openpyxl
from datetime import datetime, timedelta

def find_column_by_name(ws, header_row, search_terms):
    """在表头行中查找列"""
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(header_row, col).value
        if header_val:
            header_str = str(header_val).strip()
            for term in search_terms:
                if term in header_str:
                    return col
    return None

def extract_monthly_data_smart(year, month, data_file="excel_exports/石滩供水服务部每日总供水情况.xlsx"):
    """
    智能提取指定月份的数据
    
    逻辑：
    1. 理想时间范围：上月25日 - 本月24日
    2. 实际时间范围：上月25日 - 数据源中的最后可用日期
    
    Args:
        year: 年份
        month: 月份
        data_file: 数据源文件路径
    
    Returns:
        dict: 包含各监控点的总和数据
    """
    print(f"\n{'='*100}")
    print(f"[Smart Extracting Data] {year}年{month}月")
    print(f"{'='*100}")
    
    # 计算理想的开始日期
    if month == 1:
        start_date = datetime(year - 1, 12, 25)
    else:
        start_date = datetime(year, month - 1, 25)
    
    # 计算理想的结束日期
    ideal_end_date = datetime(year, month, 24)
    
    # 加载数据文件
    print(f"\n[Loading] {data_file}")
    wb = openpyxl.load_workbook(data_file, data_only=True)
    ws = wb.worksheets[2]  # 第3个工作表
    
    header_row = 4
    date_col = 1
    data_start_row = 5
    
    # 查找数据源中的最后可用日期
    print(f"\n[Finding Last Available Date]...")
    last_available_date = None
    last_available_row = None
    
    for row in range(ws.max_row, data_start_row - 1, -1):
        date_val = ws.cell(row, date_col).value
        if isinstance(date_val, datetime):
            last_available_date = date_val
            last_available_row = row
            break
    
    if last_available_date:
        print(f"  Last available date in data source: {last_available_date.strftime('%Y-%m-%d')} (Row {last_available_row})")
    else:
        raise Exception("No valid date found in data source")
    
    # 确定实际的结束日期
    # 如果理想结束日期超过最后可用日期，使用最后可用日期
    if ideal_end_date > last_available_date:
        actual_end_date = last_available_date
        print(f"  [INFO] Ideal end date ({ideal_end_date.strftime('%Y-%m-%d')}) is beyond available data")
        print(f"  [INFO] Using last available date: {actual_end_date.strftime('%Y-%m-%d')}")
        is_partial = True
    else:
        actual_end_date = ideal_end_date
        is_partial = False
    
    # 检查开始日期是否在数据源范围内
    if start_date > last_available_date:
        raise Exception(f"Start date {start_date.strftime('%Y-%m-%d')} is beyond available data")
    
    print(f"\n[Date Range]")
    print(f"  From: {start_date.strftime('%Y-%m-%d')}")
    print(f"  To:   {actual_end_date.strftime('%Y-%m-%d')}")
    if is_partial:
        print(f"  Status: [PARTIAL] Data is incomplete (expected until {ideal_end_date.strftime('%Y-%m-%d')})")
        missing_days = (ideal_end_date - actual_end_date).days
        print(f"  Missing: {missing_days} days of data")
    else:
        print(f"  Status: [COMPLETE] Full month data available")
    print(f"  Days: {(actual_end_date - start_date).days + 1}")
    
    # 更新列映射
    print(f"\n[Column Mapping]")
    column_mapping = {
        "荔新大道": find_column_by_name(ws, header_row, ["荔新大道", "荔新"]),
        "宁西2总表": find_column_by_name(ws, header_row, ["宁西2总表", "宁西2"]),
        "如丰大道600监控表": find_column_by_name(ws, header_row, ["如丰大道600", "如丰大道"]),
        "新城大道医院NB": find_column_by_name(ws, header_row, ["新城大道医院NB", "新城大道医院", "新城大道"]),
        "三棵树600监控表": find_column_by_name(ws, header_row, ["三棵树600", "三棵树", "三棵竹600"]),
    }
    
    for name, col in column_mapping.items():
        if col:
            header_val = ws.cell(header_row, col).value
            print(f"  {name:20s} -> Column {col:2d} ({header_val})")
        else:
            print(f"  {name:20s} -> NOT FOUND!")
    
    # 查找日期范围内的数据
    print(f"\n[Searching Data Rows]...")
    matched_rows = []
    
    for row in range(data_start_row, ws.max_row + 1):
        date_val = ws.cell(row, date_col).value
        
        if isinstance(date_val, datetime):
            if start_date <= date_val <= actual_end_date:
                matched_rows.append(row)
    
    print(f"  [OK] Found {len(matched_rows)} rows")
    
    if matched_rows:
        print(f"  First row: {matched_rows[0]} - {ws.cell(matched_rows[0], date_col).value.strftime('%Y-%m-%d')}")
        print(f"  Last row: {matched_rows[-1]} - {ws.cell(matched_rows[-1], date_col).value.strftime('%Y-%m-%d')}")
    
    # 计算总和
    print(f"\n[Calculating Totals]...")
    totals = {}
    
    for name, col in column_mapping.items():
        if col is None:
            totals[name] = 0
            print(f"  {name:20s} -> Column not found, set to 0")
            continue
        
        total = 0
        valid_count = 0
        
        for row in matched_rows:
            val = ws.cell(row, col).value
            if val and isinstance(val, (int, float)):
                total += val
                valid_count += 1
        
        totals[name] = total
        print(f"  {name:20s} -> {total:15,.0f} (from {valid_count} valid values)")
    
    return {
        "year": year,
        "month": month,
        "start_date": start_date.strftime('%Y-%m-%d'),
        "end_date": actual_end_date.strftime('%Y-%m-%d'),
        "ideal_end_date": ideal_end_date.strftime('%Y-%m-%d'),
        "is_partial": is_partial,
        "days": len(matched_rows),
        "expected_days": (ideal_end_date - start_date).days + 1,
        "totals": totals
    }

def main():
    print("="*100)
    print("[Smart Water Data Extraction Tool]")
    print("="*100)
    
    # 测试场景1：完整月份（9月）
    print("\n[Test 1] Full month data (September 2025)")
    result1 = extract_monthly_data_smart(2025, 9)
    
    print(f"\n[Result 1]")
    print(f"  Month: {result1['year']}年{result1['month']}月")
    print(f"  Date Range: {result1['start_date']} to {result1['end_date']}")
    print(f"  Status: {'PARTIAL' if result1['is_partial'] else 'COMPLETE'}")
    print(f"  Days: {result1['days']}/{result1['expected_days']}")
    
    # 测试场景2：当前月份（可能不完整）
    print(f"\n{'='*100}")
    print("\n[Test 2] Current month (may be partial)")
    try:
        result2 = extract_monthly_data_smart(2025, 10)
        
        print(f"\n[Result 2]")
        print(f"  Month: {result2['year']}年{result2['month']}月")
        print(f"  Date Range: {result2['start_date']} to {result2['end_date']}")
        print(f"  Status: {'PARTIAL' if result2['is_partial'] else 'COMPLETE'}")
        print(f"  Days: {result2['days']}/{result2['expected_days']}")
        
        if result2['is_partial']:
            print(f"\n  [WARNING] This is partial data!")
            print(f"  Expected end date: {result2['ideal_end_date']}")
            print(f"  Actual end date: {result2['end_date']}")
            print(f"  Missing days: {result2['expected_days'] - result2['days']}")
    except Exception as e:
        print(f"\n[ERROR] {e}")
    
    print(f"\n{'='*100}")
    print(f"[Complete]")
    print(f"{'='*100}")

if __name__ == "__main__":
    main()

