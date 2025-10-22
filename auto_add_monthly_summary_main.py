#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
每月1号自动在"石滩区"主工作表添加统计表
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import copy
from datetime import date
import sys

def add_monthly_summary_to_main(excel_path, force_date=None):
    """
    在"石滩区"主工作表底部添加月度统计表
    
    Args:
        excel_path: Excel文件路径
        force_date: 强制指定日期（用于测试）
    """
    print("=" * 80)
    print("[START] 在'石滩区'主工作表添加月度统计表")
    print("=" * 80)
    
    # 加载工作簿
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        print(f"[ERROR] 无法加载工作簿: {e}")
        return False
    
    # 检查工作表是否存在
    if "石滩区" not in wb.sheetnames:
        print(f"[ERROR] 工作表不存在: 石滩区")
        return False
    
    ws = wb["石滩区"]
    print(f"[OK] 找到工作表，当前最大行: {ws.max_row}")
    
    # 获取下个月信息
    today = date.today() if not force_date else force_date
    next_month = today.month + 1 if today.month < 12 else 1
    next_year = today.year if today.month < 12 else today.year + 1
    month_title = f"{next_year}年{next_month}月"
    
    print(f"[INFO] 将添加: {month_title}")
    
    # 检查是否已经添加过（避免重复）
    for row in range(max(1, ws.max_row - 10), ws.max_row + 1):
        for col in range(1, 10):
            cell_value = ws.cell(row, col).value
            if cell_value and month_title in str(cell_value):
                print(f"[WARN] 检测到已存在 {month_title} 的统计表（第{row}行）")
                print(f"[WARN] 跳过添加，避免重复")
                return False
    
    # 开始添加统计表
    start_row = ws.max_row + 2  # 留一行空行
    
    # 1. 添加月份标题
    month_row = start_row
    month_cell = ws.cell(month_row, 5)  # E列
    month_cell.value = month_title
    
    # 复制格式
    if ws.cell(51, 5).value:
        source_cell = ws.cell(51, 5)
        month_cell.font = copy.copy(source_cell.font)
        month_cell.alignment = copy.copy(source_cell.alignment)
        month_cell.fill = copy.copy(source_cell.fill)
    
    print(f"[OK] 添加月份标题: 第{month_row}行 - {month_title}")
    
    # 2. 添加表头行
    header_row = start_row + 1
    headers = {
        2: "荔新大道",
        3: "宁西总表（插入式）DN1200",
        4: "如丰大道",
        5: "新城大道医院NB",
        6: "三棵竹",
        7: "供水量",
        8: "售水量",
        9: "损耗水量",
        10: "水损耗（百分比）"
    }
    
    for col, header in headers.items():
        cell = ws.cell(header_row, col)
        cell.value = header
        # 复制格式
        if ws.cell(53, col).value is not None or col >= 2:
            source_cell = ws.cell(53, col)
            cell.font = copy.copy(source_cell.font)
            cell.alignment = copy.copy(source_cell.alignment)
            cell.border = copy.copy(source_cell.border)
            cell.fill = copy.copy(source_cell.fill)
    
    print(f"[OK] 添加表头行: 第{header_row}行")
    
    # 3. 添加数据行
    data_rows = ["1区", "2区", "3区"]
    current_row = header_row + 1
    
    for area_name in data_rows:
        ws.cell(current_row, 1).value = area_name
        
        for col in range(2, 11):
            ws.cell(current_row, col).value = 0
            source_cell = ws.cell(54, col)
            ws.cell(current_row, col).font = copy.copy(source_cell.font)
            ws.cell(current_row, col).alignment = copy.copy(source_cell.alignment)
            ws.cell(current_row, col).border = copy.copy(source_cell.border)
            if col >= 2:
                ws.cell(current_row, col).number_format = source_cell.number_format
        
        source_cell = ws.cell(54, 1)
        ws.cell(current_row, 1).font = copy.copy(source_cell.font)
        ws.cell(current_row, 1).alignment = copy.copy(source_cell.alignment)
        ws.cell(current_row, 1).border = copy.copy(source_cell.border)
        
        print(f"[OK] 添加数据行: 第{current_row}行 - {area_name}")
        current_row += 1
    
    # 4. 添加合计行
    summary_row = current_row
    ws.cell(summary_row, 6).value = "合计:"
    
    for col in range(7, 11):
        ws.cell(summary_row, col).value = 0
        source_cell = ws.cell(57, col)
        ws.cell(summary_row, col).font = copy.copy(source_cell.font)
        ws.cell(summary_row, col).alignment = copy.copy(source_cell.alignment)
        ws.cell(summary_row, col).border = copy.copy(source_cell.border)
        ws.cell(summary_row, col).fill = copy.copy(source_cell.fill)
        ws.cell(summary_row, col).number_format = source_cell.number_format
    
    source_cell = ws.cell(57, 6)
    ws.cell(summary_row, 6).font = copy.copy(source_cell.font)
    ws.cell(summary_row, 6).alignment = copy.copy(source_cell.alignment)
    ws.cell(summary_row, 6).border = copy.copy(source_cell.border)
    ws.cell(summary_row, 6).fill = copy.copy(source_cell.fill)
    
    print(f"[OK] 添加合计行: 第{summary_row}行")
    
    # 保存
    try:
        wb.save(excel_path)
        print(f"\n[SAVE] 成功保存工作簿")
    except Exception as e:
        print(f"\n[ERROR] 保存失败: {e}")
        return False
    
    print("\n" + "=" * 80)
    print("[SUCCESS] 月度统计表添加完成！")
    print("=" * 80)
    print(f"新统计表位置: 第{month_row}行 到 第{summary_row}行")
    print(f"月份: {month_title}")
    
    return True

def main():
    """
    主函数
    """
    excel_path = "excel_exports/石滩区分区计量.xlsx"
    
    # 检查今天是否是1号
    today = date.today()
    
    if len(sys.argv) > 1 and sys.argv[1] == '--force':
        # 强制执行（用于测试）
        print("[INFO] 强制执行模式（测试）")
        success = add_monthly_summary_to_main(excel_path, force_date=today)
    elif today.day == 1:
        # 每月1号自动执行
        print(f"[INFO] 今天是{today.year}年{today.month}月1日，执行添加统计表")
        success = add_monthly_summary_to_main(excel_path)
    else:
        print(f"[INFO] 今天是{today.year}年{today.month}月{today.day}日，不是1号，跳过执行")
        print(f"[INFO] 如需测试，请运行: python auto_add_monthly_summary_main.py --force")
        return
    
    if success:
        print("\n[SUCCESS] 任务完成！")
    else:
        print("\n[FAILED] 任务失败！")

if __name__ == "__main__":
    main()

