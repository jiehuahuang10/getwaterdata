#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
每月1号自动在石滩区分区计量表格中添加统计表
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import copy

class MonthlySummaryAdder:
    """
    月度统计表添加器
    """
    
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.wb = None
        self.ws = None
    
    def load_workbook(self):
        """加载工作簿"""
        try:
            self.wb = openpyxl.load_workbook(self.excel_path)
            print(f"[OK] 成功加载工作簿: {self.excel_path}")
            return True
        except Exception as e:
            print(f"[ERROR] 加载工作簿失败: {e}")
            return False
    
    def get_current_month_sheet(self):
        """
        获取当前月份的工作表
        例如：现在是10月，需要在9月的表格中添加统计
        """
        # 获取上个月的年月
        today = date.today()
        if today.month == 1:
            last_month = 12
            last_year = today.year - 1
        else:
            last_month = today.month - 1
            last_year = today.year
        
        # 构建工作表名称
        sheet_name = f"石滩{last_year}年{last_month}月"
        
        print(f"[INFO] 当前日期: {today}")
        print(f"[INFO] 目标工作表: {sheet_name}")
        
        if sheet_name in self.wb.sheetnames:
            self.ws = self.wb[sheet_name]
            print(f"[OK] 找到工作表: {sheet_name}")
            return True
        else:
            print(f"[ERROR] 未找到工作表: {sheet_name}")
            print(f"可用的工作表: {', '.join(self.wb.sheetnames)}")
            return False
    
    def analyze_existing_summary(self):
        """
        分析现有的统计表格式
        从截图看，统计表有以下特征：
        - 包含"合计:"标签
        - 有多列数据（供水量、售水量、损耗水量、水损耗等）
        - 有特定的格式和样式
        """
        print("\n[ANALYZE] 分析现有统计表格式...")
        
        # 查找"合计:"所在的行
        summary_row = None
        for row in range(self.ws.max_row, 0, -1):
            cell_value = self.ws.cell(row, 1).value
            if cell_value and "合计" in str(cell_value):
                summary_row = row
                print(f"[OK] 找到现有统计行: 第{row}行")
                break
        
        if not summary_row:
            print("[WARN] 未找到现有统计行，将创建新的统计格式")
            return None
        
        # 分析统计行的格式
        summary_info = {
            'row': summary_row,
            'cells': []
        }
        
        # 获取统计行的所有单元格信息
        for col in range(1, self.ws.max_column + 1):
            cell = self.ws.cell(summary_row, col)
            cell_info = {
                'column': col,
                'value': cell.value,
                'font': copy.copy(cell.font),
                'alignment': copy.copy(cell.alignment),
                'border': copy.copy(cell.border),
                'fill': copy.copy(cell.fill),
                'number_format': cell.number_format
            }
            summary_info['cells'].append(cell_info)
        
        print(f"[INFO] 统计行包含 {len(summary_info['cells'])} 个单元格")
        
        # 显示统计行的关键信息
        print("\n统计行内容:")
        for i, cell_info in enumerate(summary_info['cells'][:10], 1):
            if cell_info['value']:
                print(f"  列{i}: {cell_info['value']}")
        
        return summary_info
    
    def add_new_summary_row(self, template_info=None):
        """
        添加新的统计行
        """
        print("\n[ADD] 添加新的统计行...")
        
        # 在最后一行下面添加新行
        new_row = self.ws.max_row + 1
        print(f"[INFO] 新统计行位置: 第{new_row}行")
        
        if template_info:
            # 使用模板格式创建新统计行
            print("[OK] 使用现有格式创建统计行")
            
            for cell_info in template_info['cells']:
                col = cell_info['column']
                new_cell = self.ws.cell(new_row, col)
                
                # 复制值（如果是"合计:"则保留，数字则清零）
                if cell_info['value']:
                    if "合计" in str(cell_info['value']):
                        new_cell.value = cell_info['value']
                    elif isinstance(cell_info['value'], (int, float)):
                        new_cell.value = 0  # 数字初始化为0
                    else:
                        new_cell.value = cell_info['value']
                
                # 复制格式
                new_cell.font = cell_info['font']
                new_cell.alignment = cell_info['alignment']
                new_cell.border = cell_info['border']
                new_cell.fill = cell_info['fill']
                new_cell.number_format = cell_info['number_format']
        
        else:
            # 创建默认格式的统计行
            print("[WARN] 使用默认格式创建统计行")
            
            # 根据截图的格式创建
            # 第1列: 空
            # 第2-6列: 监控表供水量相关
            # 第7列: 合计:
            # 第8-12列: 损耗统计相关
            
            # 设置"合计:"标签
            合计_cell = self.ws.cell(new_row, 7)
            合计_cell.value = "合计:"
            合计_cell.font = Font(name='宋体', size=11, bold=True)
            合计_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 为所有列设置边框和初始值0
            for col in range(1, self.ws.max_column + 1):
                cell = self.ws.cell(new_row, col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 数值列初始化为0
                if col >= 2 and col != 7:
                    cell.value = 0
                    cell.number_format = '0'
        
        print(f"[OK] 成功添加新统计行到第{new_row}行")
        return new_row
    
    def save_workbook(self):
        """保存工作簿"""
        try:
            self.wb.save(self.excel_path)
            print(f"\n[SAVE] 成功保存工作簿: {self.excel_path}")
            return True
        except Exception as e:
            print(f"\n[ERROR] 保存工作簿失败: {e}")
            return False
    
    def run(self):
        """
        执行添加统计表的完整流程
        """
        print("=" * 80)
        print("[START] 开始执行月度统计表添加任务")
        print("=" * 80)
        
        # 1. 加载工作簿
        if not self.load_workbook():
            return False
        
        # 2. 获取目标工作表
        if not self.get_current_month_sheet():
            return False
        
        # 3. 分析现有统计表格式
        template_info = self.analyze_existing_summary()
        
        # 4. 添加新的统计行
        new_row = self.add_new_summary_row(template_info)
        
        if not new_row:
            return False
        
        # 5. 保存工作簿
        if not self.save_workbook():
            return False
        
        print("\n" + "=" * 80)
        print("[SUCCESS] 月度统计表添加完成！")
        print("=" * 80)
        print(f"[INFO] 工作表: {self.ws.title}")
        print(f"[INFO] 新统计行: 第{new_row}行")
        print(f"[INFO] 提示: 数据已初始化为0，等待后续填充")
        
        return True


def test_add_summary():
    """
    测试添加统计表功能
    """
    excel_path = "excel_exports/石滩区分区计量.xlsx"
    
    adder = MonthlySummaryAdder(excel_path)
    success = adder.run()
    
    if success:
        print("\n[SUCCESS] 测试成功！")
    else:
        print("\n[FAILED] 测试失败！")
    
    return success


if __name__ == "__main__":
    test_add_summary()

