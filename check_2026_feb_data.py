#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
from datetime import datetime

data_file = "excel_exports/石滩供水服务部每日总供水情况.xlsx"
wb = openpyxl.load_workbook(data_file, data_only=True)
ws = wb.worksheets[2]

print("Checking for 2026-02-25 data...")

target_date = datetime(2026, 2, 25)
found = False

for row in range(2900, 3200):
    date_val = ws.cell(row, 1).value
    if isinstance(date_val, datetime):
        if date_val.year == 2026 and date_val.month == 2:
            if date_val.day == 25:
                print(f"\nFound 2026-02-25 at row {row}")
                print(f"  Col7 (荔新大道): {ws.cell(row, 7).value}")
                print(f"  Col12 (宁西2总表): {ws.cell(row, 12).value}")
                found = True
                break
            elif date_val.day in [20, 24, 25, 26]:
                print(f"Row {row}: {date_val.strftime('%Y-%m-%d')}, Col7={ws.cell(row, 7).value}")

if not found:
    print("\n2026-02-25 NOT found or has no data")

