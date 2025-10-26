#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
纯Python计算Excel公式
用于GitHub Actions等没有Excel的环境
"""

import openpyxl
from datetime import datetime

def calculate_water_formulas(excel_path):
    """
    手动计算Excel中的公式列
    
    公式说明：
    1. 石滩供水服务部日供水 = 荔新大道 + 新城大道 + 三江新总表 + 边界过水 + 边界过水(请11) + 宁西2总表 + 沙庄总表
    2. 环比差值 = 当前行石滩供水服务部日供水 - 上一行石滩供水服务部日供水
    3. 石滩 = 荔新大道 + 新城大道
    4. 三江 = 三江新总表
    5. 沙庄 = 沙庄总表
    """
    print(f"[INFO] 开始计算公式: {excel_path}")
    
    try:
        # 打开Excel文件（不使用data_only，以保留公式）
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb.active
        
        print(f"[INFO] 工作表: {ws.title}")
        print(f"[INFO] 总行数: {ws.max_row}")
        
        # 读取表头（第4行，索引3）
        header_row = 4
        headers = [cell.value for cell in ws[header_row]]
        
        print(f"[INFO] 表头: {headers[:10]}...")  # 只打印前10列
        
        # 找到列索引（移除换行符和空格）
        # 注意：只保留第一次出现的列（因为Excel中有重复的列名）
        col_indices = {}
        for idx, header in enumerate(headers, start=1):
            if header:
                # 移除所有换行符和前后空格
                clean_header = str(header).replace('\n', '').replace('\r', '').strip()
                # 只在第一次出现时记录（避免被后面的重复列覆盖）
                if clean_header not in col_indices:
                    col_indices[clean_header] = idx
        
        # 关键列
        key_columns = [
            '荔新大道', '新城大道', '三江新总表', '边界过水', 
            '边界过水(请11)', '宁西2总表', '沙庄总表',
            '石滩供水服务部日供水', '环比差值', '石滩', '三江', '沙庄'
        ]
        
        print(f"\n[INFO] 关键列索引:")
        for col in key_columns:
            if col in col_indices:
                print(f"  {col}: 列{col_indices[col]}")
        
        # 定义辅助函数
        def get_cell_value(row, col_name):
            """安全获取单元格值，处理列不存在的情况"""
            col_idx = col_indices.get(col_name, 0)
            if col_idx == 0:
                return 0
            try:
                value = ws.cell(row, col_idx).value
                return value if value is not None else 0
            except:
                return 0
        
        def to_float(val):
            """将值转换为浮点数"""
            try:
                return float(val) if val not in (None, '', '-') else 0
            except:
                return 0
        
        # 从第5行开始计算（数据行）
        data_start_row = 5
        prev_total = None
        
        updated_count = 0
        
        # 遍历所有行，但跳过空行
        for row_idx in range(data_start_row, ws.max_row + 1):
            try:
                # 检查日期列是否有值（第1列），如果没有则跳过
                date_value = ws.cell(row_idx, 1).value
                if not date_value:
                    continue
                
                # 获取关键列索引（仅用于检查）
                lixin_col = col_indices.get('荔新大道', 0)
                xincheng_col = col_indices.get('新城大道', 0)
                sanjiang_col = col_indices.get('三江新总表', 0)
                
                # 检查关键列是否存在
                if lixin_col == 0 or xincheng_col == 0 or sanjiang_col == 0:
                    if updated_count == 0:
                        print(f"[ERROR] 关键列索引未找到，停止处理")
                        print(f"  荔新大道: {lixin_col}, 新城大道: {xincheng_col}, 三江新总表: {sanjiang_col}")
                    break
                
                # 使用安全方法获取所有列的值并转换为数值
                lixin = to_float(get_cell_value(row_idx, '荔新大道'))
                xincheng = to_float(get_cell_value(row_idx, '新城大道'))
                sanjiang = to_float(get_cell_value(row_idx, '三江新总表'))
                bianjie1 = to_float(get_cell_value(row_idx, '边界过水'))
                bianjie2 = to_float(get_cell_value(row_idx, '边界过水(请11)'))
                ningxi = to_float(get_cell_value(row_idx, '宁西2总表'))
                shazhuang = to_float(get_cell_value(row_idx, '沙庄总表'))
                
                # 计算公式
                # 1. 石滩供水服务部日供水
                total = lixin + xincheng + sanjiang + bianjie1 + bianjie2 + ningxi + shazhuang
                
                # 2. 环比差值
                huanbi = total - prev_total if prev_total is not None else 0
                
                # 3. 石滩
                shitan = lixin + xincheng
                
                # 4. 三江（直接使用三江新总表）
                sanjiang_val = sanjiang
                
                # 5. 沙庄（直接使用沙庄总表）
                shazhuang_val = shazhuang
                
                # 写入计算结果
                if '石滩供水服务部日供水' in col_indices:
                    ws.cell(row_idx, col_indices['石滩供水服务部日供水']).value = round(total, 2)
                
                if '环比差值' in col_indices:
                    ws.cell(row_idx, col_indices['环比差值']).value = round(huanbi, 2)
                
                if '石滩' in col_indices:
                    ws.cell(row_idx, col_indices['石滩']).value = round(shitan, 2)
                
                if '三江' in col_indices:
                    ws.cell(row_idx, col_indices['三江']).value = round(sanjiang_val, 2)
                
                if '沙庄' in col_indices:
                    ws.cell(row_idx, col_indices['沙庄']).value = round(shazhuang_val, 2)
                
                prev_total = total
                updated_count += 1
                
                # 每100行打印一次进度
                if updated_count % 100 == 0:
                    print(f"[INFO] 已处理 {updated_count} 行...")
                
            except Exception as row_error:
                # 只在前100次错误时打印，避免日志过长
                if updated_count < 100:
                    print(f"[WARNING] 第{row_idx}行计算失败: {row_error}")
                continue
        
        # 保存文件
        print(f"\n[INFO] 保存文件...")
        wb.save(excel_path)
        wb.close()
        
        print(f"[SUCCESS] 公式计算完成！共处理 {updated_count} 行数据")
        return True
        
    except Exception as e:
        print(f"[ERROR] 计算失败: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    import sys
    
    excel_path = 'excel_exports/石滩供水服务部每日总供水情况.xlsx'
    
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    
    success = calculate_water_formulas(excel_path)
    
    if success:
        print("\n[SUCCESS] 公式计算成功!")
        sys.exit(0)
    else:
        print("\n[ERROR] 公式计算失败!")
        sys.exit(1)

