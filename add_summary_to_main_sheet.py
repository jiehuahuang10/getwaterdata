#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
在"石滩区"主工作表底部添加统计表
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import copy
from datetime import date

def add_summary_to_main_sheet():
    """
    在"石滩区"主工作表底部添加新的统计表
    """
    excel_path = "excel_exports/石滩区分区计量.xlsx"
    
    print("=" * 80)
    print("[START] 在'石滩区'主工作表添加统计表")
    print("=" * 80)
    
    # 加载工作簿
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["石滩区"]
    
    print(f"[INFO] 当前工作表: {ws.title}")
    print(f"[INFO] 当前最大行: {ws.max_row}")
    print(f"[INFO] 当前最大列: {ws.max_column}")
    
    # 分析现有结构
    print("\n[ANALYZE] 分析表格结构...")
    print("最后10行内容:")
    for row in range(max(1, ws.max_row - 9), ws.max_row + 1):
        row_data = []
        for col in range(1, min(15, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if val:
                row_data.append(f"列{col}:{val}")
        if row_data:
            print(f"  第{row}行: {' | '.join(row_data[:8])}")
    
    # 在最后一行下面添加新的统计表
    # 根据您截图的格式（2025年8月和9月的格式）
    start_row = ws.max_row + 2  # 留一行空行
    
    print(f"\n[ADD] 新统计表起始行: {start_row}")
    
    # 1. 添加月份标题（例如：2025年10月）
    month_row = start_row
    month_cell = ws.cell(month_row, 5)  # E列
    
    # 获取下个月
    today = date.today()
    next_month = today.month + 1 if today.month < 12 else 1
    next_year = today.year if today.month < 12 else today.year + 1
    
    month_cell.value = f"{next_year}年{next_month}月"
    
    # 复制格式（从第51行的"2025年8月"）
    if ws.cell(51, 5).value:
        source_cell = ws.cell(51, 5)
        month_cell.font = copy.copy(source_cell.font)
        month_cell.alignment = copy.copy(source_cell.alignment)
        month_cell.fill = copy.copy(source_cell.fill)
    
    print(f"[OK] 添加月份标题: 第{month_row}行 - {next_year}年{next_month}月")
    
    # 2. 添加表头行（监控表供水量）
    header_row = start_row + 1
    
    # 根据截图的表头结构
        headers = {
            2: "荔新大道",
            3: "宁西总表（插入式）DN1200",
            4: "如丰大道",
            5: "新城大道医院NB",
            6: "三棵竹",
            7: "供水量",
            8: "售水量",
            9: "损耗水量",
            10: "水损耗（百分比）"
        }
    
    for col, header in headers.items():
        cell = ws.cell(header_row, col)
        cell.value = header
        # 复制第53行的格式
        if ws.cell(53, col).value is not None or col >= 2:
            source_cell = ws.cell(53, col)
            cell.font = copy.copy(source_cell.font)
            cell.alignment = copy.copy(source_cell.alignment)
            cell.border = copy.copy(source_cell.border)
            cell.fill = copy.copy(source_cell.fill)
    
    print(f"[OK] 添加表头行: 第{header_row}行")
    
    # 3. 添加数据行（1区、2区、3区）
    data_rows = ["1区", "2区", "3区"]
    current_row = header_row + 1
    
    for area_name in data_rows:
        # A列：区域名称
        ws.cell(current_row, 1).value = area_name
        
        # B-J列：数据（初始化为0）
        for col in range(2, 11):
            ws.cell(current_row, col).value = 0
            
            # 复制第54行的格式
            source_cell = ws.cell(54, col)
            ws.cell(current_row, col).font = copy.copy(source_cell.font)
            ws.cell(current_row, col).alignment = copy.copy(source_cell.alignment)
            ws.cell(current_row, col).border = copy.copy(source_cell.border)
            if col >= 2:
                ws.cell(current_row, col).number_format = source_cell.number_format
        
        # A列格式
        source_cell = ws.cell(54, 1)
        ws.cell(current_row, 1).font = copy.copy(source_cell.font)
        ws.cell(current_row, 1).alignment = copy.copy(source_cell.alignment)
        ws.cell(current_row, 1).border = copy.copy(source_cell.border)
        
        print(f"[OK] 添加数据行: 第{current_row}行 - {area_name}")
        current_row += 1
    
    # 4. 添加合计行
    summary_row = current_row
    ws.cell(summary_row, 6).value = "合计:"
    
    # 合计行的数据列
    for col in range(7, 11):
        ws.cell(summary_row, col).value = 0
        # 复制第57行的格式
        source_cell = ws.cell(57, col)
        ws.cell(summary_row, col).font = copy.copy(source_cell.font)
        ws.cell(summary_row, col).alignment = copy.copy(source_cell.alignment)
        ws.cell(summary_row, col).border = copy.copy(source_cell.border)
        ws.cell(summary_row, col).fill = copy.copy(source_cell.fill)
        ws.cell(summary_row, col).number_format = source_cell.number_format
    
    # "合计:"标签格式
    source_cell = ws.cell(57, 6)
    ws.cell(summary_row, 6).font = copy.copy(source_cell.font)
    ws.cell(summary_row, 6).alignment = copy.copy(source_cell.alignment)
    ws.cell(summary_row, 6).border = copy.copy(source_cell.border)
    ws.cell(summary_row, 6).fill = copy.copy(source_cell.fill)
    
    print(f"[OK] 添加合计行: 第{summary_row}行")
    
    # 保存到新文件
    new_path = "excel_exports/石滩区分区计量_主表已添加.xlsx"
    wb.save(new_path)
    print(f"\n[SAVE] 成功保存到: {new_path}")
    
    print("\n" + "=" * 80)
    print("[SUCCESS] 统计表添加完成！")
    print("=" * 80)
    print(f"新统计表位置: 第{month_row}行 到 第{summary_row}行")
    print(f"总共添加: {summary_row - month_row + 1} 行")
    print(f"\n请打开文件查看: {new_path}")
    print(f"工作表: 石滩区")
    
    return True

if __name__ == "__main__":
    try:
        add_summary_to_main_sheet()
        print("\n[SUCCESS] 测试成功！")
    except Exception as e:
        print(f"\n[ERROR] 测试失败: {e}")
        import traceback
        traceback.print_exc()

