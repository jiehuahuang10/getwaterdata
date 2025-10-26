#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl

excel_path = 'excel_exports/石滩区分区计量.xlsx'

wb = openpyxl.load_workbook(excel_path)
ws = wb['石滩区']

print("=== 检查包含'月'字的所有单元格 ===\n")

found_count = 0
for row_idx, row in enumerate(ws.iter_rows(), start=1):
    for col_idx, cell in enumerate(row, start=1):
        if cell.value:
            cell_str = str(cell.value)
            if '月' in cell_str or '年' in cell_str:
                print(f"第{row_idx}行, 第{col_idx}列:")
                print(f"  值: {repr(cell.value)}")
                print(f"  类型: {type(cell.value)}")
                print(f"  字符串: {cell_str}")
                print()
                found_count += 1
                if found_count >= 20:  # 只显示前20个
                    print("... (显示前20个)")
                    break
    if found_count >= 20:
        break

if found_count == 0:
    print("没有找到包含'年'或'月'的单元格！")
    print("\n=== 显示前10行的所有非空单元格 ===\n")
    for row_idx in range(1, min(11, ws.max_row + 1)):
        for col_idx in range(1, min(11, ws.max_column + 1)):
            cell = ws.cell(row_idx, col_idx)
            if cell.value:
                print(f"第{row_idx}行, 第{col_idx}列: {repr(cell.value)} (类型: {type(cell.value).__name__})")

wb.close()

