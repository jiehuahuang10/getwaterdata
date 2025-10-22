#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
网页版：点击按钮添加月度统计表
"""

from flask import Flask, render_template, jsonify, request
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import copy
from datetime import datetime
import os
from extract_water_data import extract_monthly_data

app = Flask(__name__)

EXCEL_PATH = "excel_exports/石滩区分区计量.xlsx"

def add_monthly_summary_to_main(month_offset=1, use_real_data=False, sale_values=None):
    """
    在"石滩区"主工作表底部添加月度统计表
    
    Args:
        month_offset: 月份偏移量（1=在最后月份基础上+1，2=在最后月份基础上+2...）
        use_real_data: 是否使用真实数据（从石滩供水服务部每日总供水情况.xlsx提取）
        sale_values: 售水量列表 [1区售水量, 2区售水量, 3区售水量]
    
    Returns:
        dict: 结果信息
    """
    try:
        # 加载工作簿
        wb = openpyxl.load_workbook(EXCEL_PATH)
        
        if "石滩区" not in wb.sheetnames:
            return {"success": False, "message": "工作表不存在: 石滩区"}
        
        ws = wb["石滩区"]
        
        # 查找最后一个月份
        last_month = None
        last_year = None
        last_row = None
        
        import re
        from datetime import datetime, timedelta
        
        for row in range(ws.max_row, max(0, ws.max_row - 30), -1):
            for col in range(1, 10):
                cell_value = ws.cell(row, col).value
                
                # 方式1：文本格式 "2025年11月"
                if cell_value and isinstance(cell_value, str):
                    match = re.match(r'(\d{4})年(\d{1,2})月', cell_value)
                    if match:
                        last_year = int(match.group(1))
                        last_month = int(match.group(2))
                        last_row = row
                        break
                
                # 方式2：整数格式（Excel日期序列号）
                elif cell_value and isinstance(cell_value, int):
                    # Excel日期从1900年开始，但实际是1899-12-30
                    if 40000 < cell_value < 50000:  # 合理的日期范围
                        try:
                            base = datetime(1899, 12, 30)
                            date_obj = base + timedelta(days=cell_value)
                            last_year = date_obj.year
                            last_month = date_obj.month
                            last_row = row
                            break
                        except:
                            pass
            
            if last_month:
                break
        
        # 如果找到了最后的月份，在此基础上加
        if last_month and last_year:
            target_month = last_month + month_offset
            target_year = last_year
            
            while target_month > 12:
                target_month -= 12
                target_year += 1
            
            month_title = f"{target_year}年{target_month}月"
            base_info = f"基于最后月份 {last_year}年{last_month}月（第{last_row}行）"
        else:
            # 如果没找到，使用当前日期
            today = datetime.now()
            target_month = today.month + month_offset
            target_year = today.year
            
            while target_month > 12:
                target_month -= 12
                target_year += 1
            
            month_title = f"{target_year}年{target_month}月"
            base_info = f"未找到最后月份，基于当前日期"
        
        # 检查数据源是否有该月份的完整数据
        # 需要检查数据源中是否有目标月份的24日数据（即统计周期的结束日期）
        if use_real_data:
            try:
                data_source_file = "excel_exports/石滩供水服务部每日总供水情况.xlsx"
                wb_data = openpyxl.load_workbook(data_source_file, data_only=True)
                ws_data = wb_data.worksheets[2]
                
                # 需要检查的日期：目标月份的24日（统计周期结束日期）
                from datetime import datetime as dt
                required_end_date = dt(target_year, target_month, 24)
                
                # 检查数据源中是否有这个日期（且有非零数据）
                date_col = 1
                check_columns = [7, 12, 14, 8, 15]  # 监控点列
                has_required_data = False
                
                for row in range(5, ws_data.max_row + 1):
                    date_val = ws_data.cell(row, date_col).value
                    if isinstance(date_val, dt) and date_val.date() == required_end_date.date():
                        # 检查是否有非零数据
                        for col in check_columns:
                            val = ws_data.cell(row, col).value
                            if val and isinstance(val, (int, float)) and val != 0:
                                has_required_data = True
                                break
                        break
                
                if not has_required_data:
                    return {
                        "success": False,
                        "message": f"[提示] 数据源中找不到完整数据！\n\n目标月份：{month_title}\n需要数据：{required_end_date.strftime('%Y年%m月%d日')}\n\n数据源文件中找不到{target_month}月24日的数据，或该日期数据为空。\n\n统计周期为上月25日至本月24日，必须等到本月24日数据更新后才能添加统计表。"
                    }
                
                wb_data.close()
            except Exception as e:
                print(f"[WARNING] Failed to check data source: {e}")
                # 如果检查失败，继续执行（降级到模拟数据）
        
        # 检查是否已经添加过
        for row in range(max(1, ws.max_row - 10), ws.max_row + 1):
            for col in range(1, 10):
                cell_value = ws.cell(row, col).value
                if cell_value and month_title in str(cell_value):
                    return {
                        "success": False, 
                        "message": f"已存在 {month_title} 的统计表（第{row}行），请勿重复添加"
                    }
        
        # 开始添加统计表
        start_row = ws.max_row + 2
        
        # 1. 添加月份标题（合并单元格）
        month_row = start_row
        
        # 定义统一的边框样式
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # 合并单元格（从A到J列，覆盖整行）
        ws.merge_cells(start_row=month_row, start_column=1, end_row=month_row, end_column=10)
        
        month_cell = ws.cell(month_row, 1)
        month_cell.value = month_title
        
        # 先复制格式
        if ws.cell(51, 5).value:
            source_cell = ws.cell(51, 5)
            month_cell.font = copy.copy(source_cell.font)
            month_cell.fill = copy.copy(source_cell.fill)
        
        # 给合并区域的所有单元格设置边框（从A到J列）
        for col in range(1, 11):
            ws.cell(month_row, col).border = thin_border
        
        # 最后设置居中对齐（确保不被覆盖）
        month_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 2. 添加分类标题行（监控表供水量、损耗统计）
        category_row = start_row + 1
        
        # 添加"监控表供水量"标题（合并B到F列）
        ws.merge_cells(start_row=category_row, start_column=2, end_row=category_row, end_column=6)
        category_cell1 = ws.cell(category_row, 2)
        category_cell1.value = "监控表供水量"
        category_cell1.font = copy.copy(ws.cell(52, 2).font)
        category_cell1.alignment = Alignment(horizontal='center', vertical='center')
        category_cell1.border = thin_border
        category_cell1.fill = copy.copy(ws.cell(52, 2).fill)
        
        # 添加"损耗统计"标题（合并G到J列）
        ws.merge_cells(start_row=category_row, start_column=7, end_row=category_row, end_column=10)
        category_cell2 = ws.cell(category_row, 7)
        category_cell2.value = "损耗统计"
        category_cell2.font = copy.copy(ws.cell(52, 7).font)
        category_cell2.alignment = Alignment(horizontal='center', vertical='center')
        category_cell2.border = thin_border
        category_cell2.fill = copy.copy(ws.cell(52, 7).fill)
        
        # 给合并区域的其他单元格也设置边框
        for col in range(2, 7):
            ws.cell(category_row, col).border = thin_border
        for col in range(7, 11):
            ws.cell(category_row, col).border = thin_border
        
        # A列也设置边框（分类标题行）
        ws.cell(category_row, 1).border = thin_border
        
        # 3. 添加详细表头行
        header_row = start_row + 2
        
        # A列设置边框（详细表头行）
        ws.cell(header_row, 1).border = thin_border
        ws.cell(header_row, 1).alignment = Alignment(horizontal='center', vertical='center')
        
        headers = {
            2: "荔新大道",
            3: "宁西总表（插入式）DN1200",
            4: "如丰大道",
            5: "新城大道医院NB",
            6: "三棵竹",
            7: "供水量",
            8: "售水量",
            9: "损耗水量",
            10: "水损耗（百分比）"
        }
        
        for col, header in headers.items():
            cell = ws.cell(header_row, col)
            cell.value = header
            if ws.cell(53, col).value is not None or col >= 2:
                source_cell = ws.cell(53, col)
                cell.font = copy.copy(source_cell.font)
                # 设置居中对齐
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # 确保有边框
                cell.border = thin_border
                cell.fill = copy.copy(source_cell.fill)
        
        # 4. 添加数据行（包含模拟数据或真实数据和公式）
        import random
        
        # 如果使用真实数据，先提取
        real_data_totals = None
        if use_real_data:
            try:
                print(f"[INFO] Extracting real data for {target_year}-{target_month}...")
                real_data_result = extract_monthly_data(target_year, target_month)
                real_data_totals = real_data_result['totals']
                print(f"[OK] Real data extracted successfully")
            except Exception as e:
                print(f"[WARNING] Failed to extract real data: {e}")
                print(f"[INFO] Falling back to simulated data")
                real_data_totals = None
        
        # 定义三种数据行类型（参考9月真实数据）
        row_types = [
            {
                'label': '1 区',
                'formula': 'subtract',  # B-C-D
                'b_range': (4200000, 4300000),
                'c_range': (3400000, 3500000),
                'd_range': (250000, 260000),
                'sale_range': (200000, 210000)
            },
            {
                'label': '2 区',
                'formula': 'add_subtract',  # D+E-F
                'd_range': (250000, 260000),
                'e_range': (500000, 510000),
                'f_range': (260000, 270000),
                'sale_range': (500000, 510000)
            },
            {
                'label': '3 区',
                'formula': 'direct',  # =F
                'f_range': (260000, 270000),
                'sale_range': (290000, 295000)
            }
        ]
        
        current_row = header_row + 1
        data_start_row = current_row
        
        for idx, row_type in enumerate(row_types):
            # A列：标签（加边框和居中）
            a_cell = ws.cell(current_row, 1)
            a_cell.value = row_type['label']
            a_cell.border = thin_border
            a_cell.alignment = Alignment(horizontal='center', vertical='center')
            # 复制A列的格式
            if ws.cell(61 + idx, 1).value:
                source_a_cell = ws.cell(61 + idx, 1)
                a_cell.font = copy.copy(source_a_cell.font)
                if source_a_cell.fill:
                    a_cell.fill = copy.copy(source_a_cell.fill)
            
            # 生成监控点数据（真实数据或模拟数据）和公式
            if row_type['formula'] == 'subtract':
                # 第一种类型: 供水量 = B-C-D
                if real_data_totals:
                    b_val = round(real_data_totals.get("荔新大道", 0))
                    c_val = round(real_data_totals.get("宁西2总表", 0))
                    d_val = round(real_data_totals.get("如丰大道600监控表", 0))
                else:
                    b_val = random.randint(*row_type['b_range'])
                    c_val = random.randint(*row_type['c_range'])
                    d_val = random.randint(*row_type['d_range'])
                
                ws.cell(current_row, 2).value = b_val  # B列: 荔新大道
                ws.cell(current_row, 3).value = c_val  # C列: 宁西总表
                ws.cell(current_row, 4).value = d_val  # D列: 如丰大道
                ws.cell(current_row, 5).value = 0      # E列
                ws.cell(current_row, 6).value = 0      # F列
                
                # G列：供水量 = B-C-D
                ws.cell(current_row, 7).value = f"=B{current_row}-C{current_row}-D{current_row}"
                
            elif row_type['formula'] == 'add_subtract':
                # 第二种类型: 供水量 = D+E-F
                if real_data_totals:
                    d_val = round(real_data_totals.get("如丰大道600监控表", 0))
                    e_val = round(real_data_totals.get("新城大道医院NB", 0))
                    f_val = round(real_data_totals.get("三棵树600监控表", 0))
                else:
                    d_val = random.randint(*row_type['d_range'])
                    e_val = random.randint(*row_type['e_range'])
                    f_val = random.randint(*row_type['f_range'])
                
                ws.cell(current_row, 2).value = 0      # B列
                ws.cell(current_row, 3).value = 0      # C列
                ws.cell(current_row, 4).value = d_val  # D列: 如丰大道
                ws.cell(current_row, 5).value = e_val  # E列: 新城大道医院NB
                ws.cell(current_row, 6).value = f_val  # F列: 三棵竹
                
                # G列：供水量 = D+E-F
                ws.cell(current_row, 7).value = f"=D{current_row}+E{current_row}-F{current_row}"
                
            elif row_type['formula'] == 'direct':
                # 第三种类型: 供水量 = F
                if real_data_totals:
                    f_val = round(real_data_totals.get("三棵树600监控表", 0))
                else:
                    f_val = random.randint(*row_type['f_range'])
                
                ws.cell(current_row, 2).value = 0      # B列
                ws.cell(current_row, 3).value = 0      # C列
                ws.cell(current_row, 4).value = 0      # D列
                ws.cell(current_row, 5).value = 0      # E列
                ws.cell(current_row, 6).value = f_val  # F列: 三棵竹
                
                # G列：供水量 = F
                ws.cell(current_row, 7).value = f"=F{current_row}"
            
            # H列：售水量（从用户输入获取）
            if sale_values and len(sale_values) > idx:
                ws.cell(current_row, 8).value = sale_values[idx]
            # 否则留空，等待用户手动输入
            
            # I列：损耗水量 = G-H
            ws.cell(current_row, 9).value = f"=G{current_row}-H{current_row}"
            
            # J列：水损耗率 = I/G
            ws.cell(current_row, 10).value = f"=I{current_row}/G{current_row}"
            
            # 复制格式并确保有边框
            source_row = 61 + idx  # 参考9月的对应行
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            for col in range(1, 11):
                source_cell = ws.cell(source_row, col)
                target_cell = ws.cell(current_row, col)
                target_cell.font = copy.copy(source_cell.font)
                # 设置居中对齐
                target_cell.alignment = Alignment(horizontal='center', vertical='center')
                # 确保有边框
                target_cell.border = thin_border
                if source_cell.fill:
                    target_cell.fill = copy.copy(source_cell.fill)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
            
            # 特别设置J列（水损耗%）为百分比格式
            ws.cell(current_row, 10).number_format = '0.00%'
            
            current_row += 1
        
        # 5. 添加合计行（包含SUM公式）
        summary_row = current_row
        data_end_row = current_row - 1
        
        # A-E列：设置边框（合计标签之前的空单元格）
        for col in range(1, 6):
            ws.cell(summary_row, col).border = thin_border
            ws.cell(summary_row, col).alignment = Alignment(horizontal='center', vertical='center')
        
        # F列：合计标签
        ws.cell(summary_row, 6).value = "合计:"
        
        # G列：供水量总计 = SUM(G数据行)
        ws.cell(summary_row, 7).value = f"=SUM(G{data_start_row}:G{data_end_row})"
        
        # H列：售水量总计 = SUM(H数据行)
        ws.cell(summary_row, 8).value = f"=SUM(H{data_start_row}:H{data_end_row})"
        
        # I列：损耗水量总计 = G合计 - H合计
        ws.cell(summary_row, 9).value = f"=G{summary_row}-H{summary_row}"
        
        # J列：水损耗率总计 = I合计 / G合计
        ws.cell(summary_row, 10).value = f"=I{summary_row}/G{summary_row}"
        
        # 复制格式（从9月的合计行 - 第64行）并确保有边框
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        for col in range(6, 11):
            source_cell = ws.cell(64, col)
            target_cell = ws.cell(summary_row, col)
            target_cell.font = copy.copy(source_cell.font)
            # 设置居中对齐
            target_cell.alignment = Alignment(horizontal='center', vertical='center')
            # 确保有边框
            target_cell.border = thin_border
            if source_cell.fill:
                target_cell.fill = copy.copy(source_cell.fill)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
        
        # 特别设置合计行J列（水损耗%）为百分比格式
        ws.cell(summary_row, 10).number_format = '0.00%'
        
        # 保存
        wb.save(EXCEL_PATH)
        
        return {
            "success": True,
            "message": f"成功添加 {month_title} 统计表",
            "month": month_title,
            "base_info": base_info,
            "row_start": month_row,
            "row_end": summary_row,
            "total_rows": summary_row - month_row + 1
        }
        
    except Exception as e:
        return {
            "success": False,
            "message": f"添加失败: {str(e)}"
        }

@app.route('/')
def index():
    """主页"""
    return render_template('add_summary.html')

@app.route('/add_summary', methods=['POST'])
def add_summary():
    """添加统计表API"""
    try:
        # 获取参数
        data = request.get_json()
        month_offset = data.get('month_offset', 1) if data else 1
        use_real_data = data.get('use_real_data', True) if data else True  # 默认使用真实数据
        sale_values = data.get('sale_values', None) if data else None  # 获取售水量数组
        
        result = add_monthly_summary_to_main(month_offset, use_real_data, sale_values)
        return jsonify(result)
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"服务器错误: {str(e)}"
        })

@app.route('/get_info', methods=['GET'])
def get_info():
    """获取当前表格信息"""
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb["石滩区"]
        
        # 查找最后一个月份
        last_month = None
        for row in range(ws.max_row, max(0, ws.max_row - 20), -1):
            for col in range(1, 10):
                cell_value = ws.cell(row, col).value
                if cell_value and "年" in str(cell_value) and "月" in str(cell_value):
                    last_month = str(cell_value)
                    break
            if last_month:
                break
        
        wb.close()
        
        return jsonify({
            "success": True,
            "total_rows": ws.max_row,
            "last_month": last_month,
            "excel_path": EXCEL_PATH
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"获取信息失败: {str(e)}"
        })

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5001))
    print("=" * 80)
    print("月度统计表添加工具 - 网页版")
    print("=" * 80)
    print(f"Excel文件: {EXCEL_PATH}")
    print(f"访问地址: http://0.0.0.0:{port}")
    print("=" * 80)
    app.run(debug=False, host='0.0.0.0', port=port)

