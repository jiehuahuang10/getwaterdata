#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
使用openpyxl重新保存Excel文件，触发公式重新计算
注意：openpyxl本身不能计算公式，但可以通过重新保存来保留公式
在有Excel的环境中打开时会自动计算
"""

import openpyxl
import os

def recalculate_excel_formulas(excel_path):
    """
    重新保存Excel文件，确保公式被保留
    
    Args:
        excel_path: Excel文件路径
    
    Returns:
        bool: 是否成功
    """
    try:
        print(f"[INFO] 正在处理文件: {excel_path}")
        
        # 以非只读模式打开，data_only=False保留公式
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb.active
        
        print(f"[INFO] 工作表: {ws.title}")
        print(f"[INFO] 行数: {ws.max_row}, 列数: {ws.max_column}")
        
        # 设置工作簿为需要完全计算
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True
        
        # 保存文件
        wb.save(excel_path)
        wb.close()
        
        print("[SUCCESS] 文件已保存，公式将在Excel中打开时自动计算")
        return True
        
    except Exception as e:
        print(f"[ERROR] 处理失败: {e}")
        return False

if __name__ == '__main__':
    excel_path = 'excel_exports/石滩供水服务部每日总供水情况.xlsx'
    
    if os.path.exists(excel_path):
        recalculate_excel_formulas(excel_path)
    else:
        print(f"[ERROR] 文件不存在: {excel_path}")

