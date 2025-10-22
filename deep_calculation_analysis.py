#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
深度分析石滩区分区计量表格的数据计算情况
重点分析公式和数据关系
"""

import openpyxl
from openpyxl.utils import get_column_letter
import re
from datetime import datetime

def main():
    excel_path = "excel_exports/石滩区分区计量.xlsx"
    
    print("="*100)
    print("[START] Deep Calculation Analysis")
    print("="*100)
    
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb["石滩区"]
    
    print(f"\n[Basic Info]")
    print(f"  Total Rows: {ws.max_row}")
    print(f"  Total Columns: {ws.max_column}")
    
    # 分析第一个完整的月度统计表 (2025年10月 - 第67行开始)
    print(f"\n{'='*100}")
    print(f"[Analysis Target] Monthly Summary Table (Row 67 - 2025年10月)")
    print(f"{'='*100}")
    
    start_row = 67
    
    # 1. 读取表头
    print(f"\n[Step 1] Headers (Row {start_row + 1}):")
    headers = {}
    for col in range(1, 11):
        val = ws.cell(start_row + 1, col).value
        if val:
            headers[col] = val
            col_letter = get_column_letter(col)
            print(f"  Column {col_letter} (#{col}): {val}")
    
    # 2. 分析数据行的结构
    print(f"\n[Step 2] Data Rows Structure:")
    
    # 先找到有数据的行
    data_row = start_row + 2  # 第一行数据
    
    print(f"\n  Analyzing Row {data_row} (First Data Row):")
    for col in range(1, 11):
        cell = ws.cell(data_row, col)
        col_letter = get_column_letter(col)
        col_name = headers.get(col, f"Col{col}")
        
        if cell.value is not None:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                print(f"    {col_letter}{data_row} ({col_name}): FORMULA = {cell.value}")
            else:
                print(f"    {col_letter}{data_row} ({col_name}): VALUE = {cell.value}")
        else:
            print(f"    {col_letter}{data_row} ({col_name}): EMPTY")
    
    # 3. 查找总计行
    print(f"\n[Step 3] Looking for Total Row...")
    total_row = None
    for row in range(start_row + 2, start_row + 40):
        first_cell = ws.cell(row, 1).value
        if first_cell and isinstance(first_cell, str) and ('总计' in first_cell or '合计' in first_cell):
            total_row = row
            print(f"  [OK] Found Total Row at {row}")
            break
    
    if total_row:
        print(f"\n  Total Row {total_row} Analysis:")
        for col in range(1, 11):
            cell = ws.cell(total_row, col)
            col_letter = get_column_letter(col)
            col_name = headers.get(col, f"Col{col}")
            
            if cell.value is not None:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"    {col_letter}{total_row} ({col_name}): FORMULA = {cell.value}")
                else:
                    print(f"    {col_letter}{total_row} ({col_name}): VALUE = {cell.value}")
    
    # 4. 分析列之间的计算关系
    print(f"\n{'='*100}")
    print(f"[Key Calculation Relationships]")
    print(f"{'='*100}")
    
    # 从公式推断关系
    print(f"\n[Analysis] Based on formulas in the table:")
    
    # 检查各列的公式模式
    relationships = {}
    
    for col in range(2, 11):
        col_letter = get_column_letter(col)
        col_name = headers.get(col, f"Col{col}")
        
        # 检查这一列的公式
        formulas_found = []
        for row in range(start_row + 2, start_row + 35):
            cell = ws.cell(row, col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas_found.append({
                    'row': row,
                    'formula': cell.value
                })
        
        if formulas_found:
            relationships[col_name] = formulas_found
    
    # 输出关系
    for col_name, formulas in relationships.items():
        print(f"\n  [{col_name}]:")
        
        # 分析第一个公式
        if formulas:
            first_formula = formulas[0]['formula']
            print(f"    Formula Example: {first_formula}")
            
            # 解析公式含义
            if 'SUM' in first_formula.upper():
                print(f"    Type: SUM (求和)")
                # 提取求和范围
                match = re.search(r'SUM\((.*?)\)', first_formula, re.IGNORECASE)
                if match:
                    range_str = match.group(1)
                    print(f"    Sum Range: {range_str}")
            elif '+' in first_formula:
                print(f"    Type: Addition (加法)")
                # 提取加法项
                parts = re.findall(r'[A-Z]+\d+', first_formula)
                if parts:
                    print(f"    Terms: {' + '.join(parts)}")
            elif '-' in first_formula:
                print(f"    Type: Subtraction (减法)")
                # 提取减法项
                parts = re.findall(r'[A-Z]+\d+', first_formula)
                if parts and len(parts) >= 2:
                    print(f"    Formula: {parts[0]} - {' - '.join(parts[1:])}")
            elif '/' in first_formula:
                print(f"    Type: Division (除法)")
                # 提取除法项
                parts = re.findall(r'[A-Z]+\d+', first_formula)
                if parts and len(parts) >= 2:
                    print(f"    Formula: {parts[0]} / {parts[1]}")
            elif '*' in first_formula:
                print(f"    Type: Multiplication (乘法)")
    
    # 5. 推断业务逻辑
    print(f"\n{'='*100}")
    print(f"[Business Logic Inference]")
    print(f"{'='*100}")
    
    print(f"\n根据表头和公式分析，推断以下业务逻辑：")
    print(f"\n1. 【供水监控表】（列 B-F）:")
    print(f"   - B列: 荔新大道")
    print(f"   - C列: 宁西总表（插入式）DN1200")
    print(f"   - D列: 如丰大道")
    print(f"   - E列: 新城大道医院NB")
    print(f"   - F列: 三棵竹")
    print(f"   这5个列是监控点的数据")
    
    print(f"\n2. 【供水量】（列 G）:")
    if '供水量' in relationships:
        formulas = relationships['供水量']
        if formulas:
            example = formulas[0]['formula']
            print(f"   公式示例: {example}")
            
            # 分析公式逻辑
            if 'B' in example and 'C' in example and 'D' in example:
                if '-' in example:
                    print(f"   逻辑: 可能是 B - C - D 或其他减法组合")
                elif '+' in example:
                    print(f"   逻辑: 可能是 B + C + D + ... 的总和")
            elif 'SUM' in example.upper():
                print(f"   逻辑: 对某些列求和")
    
    print(f"\n3. 【售水量】（列 H）:")
    print(f"   这一列通常是手动输入的实际售水数据")
    
    print(f"\n4. 【损耗水量】（列 I）:")
    if '损耗水量' in relationships:
        formulas = relationships['损耗水量']
        if formulas:
            example = formulas[0]['formula']
            print(f"   公式示例: {example}")
            print(f"   逻辑: 通常是 供水量 - 售水量 = 损耗水量")
            print(f"   即: G列 - H列 = I列")
    
    print(f"\n5. 【水损耗（百分比）】（列 J）:")
    if '水损耗（百分比）' in relationships:
        formulas = relationships['水损耗（百分比）']
        if formulas:
            example = formulas[0]['formula']
            print(f"   公式示例: {example}")
            print(f"   逻辑: 通常是 损耗水量 / 供水量 = 损耗率")
            print(f"   即: I列 / G列 = J列")
    
    # 6. 检查其他月度明细表
    print(f"\n{'='*100}")
    print(f"[Other Monthly Detail Sheets]")
    print(f"{'='*100}")
    
    monthly_sheets = [name for name in wb.sheetnames if '2025年' in name]
    print(f"\nFound {len(monthly_sheets)} monthly detail sheets:")
    
    for sheet_name in monthly_sheets[:3]:  # 只分析前3个
        ws_monthly = wb[sheet_name]
        print(f"\n  [{sheet_name}]:")
        print(f"    Rows: {ws_monthly.max_row}, Columns: {ws_monthly.max_column}")
        
        # 查找表头
        header_row = None
        for row in range(1, 10):
            val = ws_monthly.cell(row, 1).value
            if val and isinstance(val, str) and '日期' in val:
                header_row = row
                break
        
        if header_row:
            print(f"    Header Row: {header_row}")
            headers_monthly = []
            for col in range(1, min(15, ws_monthly.max_column + 1)):
                val = ws_monthly.cell(header_row, col).value
                if val:
                    headers_monthly.append(str(val))
            print(f"    Headers: {', '.join(headers_monthly[:10])}")
            
            # 检查是否有公式
            formula_count = 0
            for row in range(header_row + 1, min(header_row + 32, ws_monthly.max_row + 1)):
                for col in range(1, min(15, ws_monthly.max_column + 1)):
                    cell = ws_monthly.cell(row, col)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
            
            print(f"    Formulas Found: {formula_count}")
            
            # 显示第一个公式示例
            for row in range(header_row + 1, min(header_row + 32, ws_monthly.max_row + 1)):
                for col in range(1, min(15, ws_monthly.max_column + 1)):
                    cell = ws_monthly.cell(row, col)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        col_letter = get_column_letter(col)
                        print(f"    Formula Example: {col_letter}{row} = {cell.value}")
                        break
                else:
                    continue
                break
    
    # 7. 总结
    print(f"\n{'='*100}")
    print(f"[SUMMARY]")
    print(f"{'='*100}")
    
    print(f"\n本表格的数据计算逻辑：")
    print(f"\n1. 数据源：")
    print(f"   - 月度明细表（如'石滩2025年9月'）记录每日的原始数据")
    print(f"   - 主表（'石滩区'）汇总月度统计")
    
    print(f"\n2. 主表月度统计表结构：")
    print(f"   - 第1行: 月份标题（如'2025年10月'）")
    print(f"   - 第2行: 表头（监控点名称 + 汇总列）")
    print(f"   - 第3行起: 每日数据（可能有公式引用月度明细表）")
    print(f"   - 最后: 总计行（对当月所有数据求和）")
    
    print(f"\n3. 关键计算公式：")
    print(f"   - 供水量 = f(监控点数据)  // 可能是加减运算")
    print(f"   - 损耗水量 = 供水量 - 售水量")
    print(f"   - 水损耗率 = 损耗水量 / 供水量")
    
    print(f"\n4. 数据流向：")
    print(f"   原始数据 → 月度明细表 → 主表月度统计 → 总计")
    
    print(f"\n{'='*100}")
    print(f"[OK] Analysis Complete!")
    print(f"{'='*100}")
    
    # 生成详细报告
    report_path = "DETAILED_CALCULATION_REPORT.md"
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("# 石滩区分区计量表格 - 数据计算详细分析报告\n\n")
        f.write(f"**分析时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write("---\n\n")
        
        f.write("## 一、表格结构概述\n\n")
        f.write(f"### 1.1 工作表列表\n\n")
        f.write(f"共 {len(wb.sheetnames)} 个工作表：\n\n")
        for name in wb.sheetnames:
            sheet = wb[name]
            f.write(f"- **{name}** ({sheet.max_row} 行 × {sheet.max_column} 列)\n")
        
        f.write("\n### 1.2 主表（石滩区）\n\n")
        f.write("主表包含多个月度统计表，每个统计表的结构相同：\n\n")
        f.write("```\n")
        f.write("行N:   YYYY年MM月          (月份标题)\n")
        f.write("行N+1: 监控点表头 | 汇总列   (表头)\n")
        f.write("行N+2: 日期1的数据         (数据行)\n")
        f.write("行N+3: 日期2的数据\n")
        f.write("...\n")
        f.write("行N+M: 总计               (总计行)\n")
        f.write("```\n\n")
        
        f.write("## 二、数据列说明\n\n")
        f.write("### 2.1 监控点列（输入数据）\n\n")
        for col_num in [2, 3, 4, 5, 6]:
            if col_num in headers:
                f.write(f"- **列{get_column_letter(col_num)}**: {headers[col_num]}\n")
        
        f.write("\n### 2.2 汇总计算列（公式列）\n\n")
        for col_num in [7, 8, 9, 10]:
            if col_num in headers:
                col_name = headers[col_num]
                f.write(f"#### {col_name} (列{get_column_letter(col_num)})\n\n")
                
                if col_name in relationships and relationships[col_name]:
                    example_formula = relationships[col_name][0]['formula']
                    f.write(f"**公式示例**: `{example_formula}`\n\n")
                    
                    if col_name == '供水量':
                        f.write("**计算逻辑**: 根据各监控点数据计算总供水量\n\n")
                    elif col_name == '售水量':
                        f.write("**计算逻辑**: 实际售出的水量（可能是手动输入或引用其他数据）\n\n")
                    elif col_name == '损耗水量':
                        f.write("**计算逻辑**: `损耗水量 = 供水量 - 售水量`\n\n")
                    elif col_name == '水损耗（百分比）':
                        f.write("**计算逻辑**: `水损耗率 = 损耗水量 / 供水量`\n\n")
                else:
                    f.write("**类型**: 输入列（无公式）\n\n")
        
        f.write("## 三、公式详解\n\n")
        
        for col_name, formulas in relationships.items():
            if formulas:
                f.write(f"### 3.{list(relationships.keys()).index(col_name) + 1} {col_name}\n\n")
                f.write(f"**公式数量**: {len(formulas)}\n\n")
                f.write("**公式示例**:\n\n")
                
                for i, formula_info in enumerate(formulas[:5]):
                    row_num = formula_info['row']
                    formula = formula_info['formula']
                    f.write(f"{i+1}. 第{row_num}行: `{formula}`\n")
                
                if len(formulas) > 5:
                    f.write(f"\n... 还有 {len(formulas) - 5} 个公式\n")
                
                f.write("\n")
        
        f.write("## 四、数据流向图\n\n")
        f.write("```\n")
        f.write("┌─────────────────┐\n")
        f.write("│  原始数据录入    │\n")
        f.write("└────────┬────────┘\n")
        f.write("         │\n")
        f.write("         ↓\n")
        f.write("┌─────────────────┐\n")
        f.write("│ 月度明细表       │  (如: 石滩2025年9月)\n")
        f.write("│ - 每日数据       │\n")
        f.write("│ - 监控点读数     │\n")
        f.write("└────────┬────────┘\n")
        f.write("         │\n")
        f.write("         ↓\n")
        f.write("┌─────────────────┐\n")
        f.write("│ 主表月度统计     │  (石滩区工作表)\n")
        f.write("│ - 供水量计算     │\n")
        f.write("│ - 损耗计算       │\n")
        f.write("│ - 损耗率计算     │\n")
        f.write("└────────┬────────┘\n")
        f.write("         │\n")
        f.write("         ↓\n")
        f.write("┌─────────────────┐\n")
        f.write("│  月度总计        │\n")
        f.write("│  (SUM公式)       │\n")
        f.write("└─────────────────┘\n")
        f.write("```\n\n")
        
        f.write("## 五、自动化建议\n\n")
        f.write("### 5.1 现有功能\n\n")
        f.write("- ✅ 自动添加新月度统计表结构\n")
        f.write("- ✅ 自动复制格式和样式\n")
        f.write("- ✅ Web界面手动触发\n\n")
        
        f.write("### 5.2 待实现功能\n\n")
        f.write("1. **数据自动填充**\n")
        f.write("   - 从月度明细表自动提取数据\n")
        f.write("   - 自动计算供水量、损耗量、损耗率\n\n")
        
        f.write("2. **公式自动生成**\n")
        f.write("   - 为新添加的统计表自动添加计算公式\n")
        f.write("   - 自动引用对应月份的明细表数据\n\n")
        
        f.write("3. **数据校验**\n")
        f.write("   - 检查数据完整性\n")
        f.write("   - 检查计算结果准确性\n\n")
        
        f.write("---\n\n")
        f.write("*报告生成完成*\n")
    
    print(f"\n[OK] Detailed report saved to: {report_path}")

if __name__ == "__main__":
    main()

