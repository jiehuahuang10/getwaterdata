#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
import re
from datetime import datetime, timedelta

wb = openpyxl.load_workbook('excel_exports/石滩区分区计量.xlsx', data_only=True)
ws = wb['石滩区']

print("Checking last months in Excel:")
for row in range(50, min(100, ws.max_row + 1)):
    for col in range(1, 10):
        val = ws.cell(row, col).value
        if val:
            if isinstance(val, str) and re.search(r'\d{4}年\d{1,2}月', val):
                print(f"Row {row}, Col {col}: {val}")
                break
            elif isinstance(val, int) and 40000 < val < 50000:
                try:
                    base = datetime(1899, 12, 30)
                    date_obj = base + timedelta(days=val)
                    print(f"Row {row}, Col {col}: {val} (date: {date_obj.year}年{date_obj.month}月)")
                    break
                except:
                    pass

