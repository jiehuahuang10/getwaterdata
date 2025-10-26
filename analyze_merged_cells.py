#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl

excel_path = 'excel_exports/石滩区分区计量.xlsx'

wb = openpyxl.load_workbook(excel_path)
ws = wb['石滩区']

print("=== 合并单元格信息 ===\n")

for idx, merged_range in enumerate(ws.merged_cells.ranges, 1):
    cell = ws.cell(merged_range.min_row, merged_range.min_col)
    print(f"{idx}. 区域: {merged_range}")
    print(f"   起始: 第{merged_range.min_row}行, 第{merged_range.min_col}列")
    print(f"   结束: 第{merged_range.max_row}行, 第{merged_range.max_col}列")
    print(f"   跨度: {merged_range.max_row - merged_range.min_row + 1}行 x {merged_range.max_col - merged_range.min_col + 1}列")
    print(f"   内容: {repr(cell.value)}")
    print()

print(f"\n总共 {len(ws.merged_cells.ranges)} 个合并单元格")

# 显示前10行的数据结构
print("\n=== 前10行数据 ===\n")
for row_idx in range(1, min(11, ws.max_row + 1)):
    print(f"第{row_idx}行:")
    for col_idx in range(1, min(11, ws.max_column + 1)):
        cell = ws.cell(row_idx, col_idx)
        if cell.value:
            print(f"  列{col_idx}: {repr(cell.value)}")
    print()

wb.close()

